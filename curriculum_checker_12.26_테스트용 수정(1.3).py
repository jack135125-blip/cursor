# -*- coding: utf-8 -*-
"""
교육과정 편성표 확인 프로그램 (Tkinter + openpyxl)
- 단일 파일(모듈 분리 없음)
- .xlsx / .xlsm 지원 (openpyxl)


사용 방법
1) pip install openpyxl

"""

import os
import re
import difflib
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from io import BytesIO


from openpyxl import load_workbook

try:
    import requests  # type: ignore
    import urllib3  # type: ignore
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False


# =========================
# 유틸
# =========================

EPS = 1e-9

# 구글 스프레드시트 URL
GOOGLE_SHEET_ID = "1BaTm1J34hep9QV8fswwPfcfCZX-geGtanLwX9BkhCyU"


def load_workbook_from_google_sheet(spreadsheet_id: str, timeout=10):
    """
    구글 스프레드시트에서 Excel 형식으로 다운로드하여 workbook 반환
    return: (wb_data_only, wb_formula, error_msg) 또는 (None, None, error_msg)
    """
    if not REQUESTS_AVAILABLE:
        return None, None, "온라인 데이터 참조 기능을 사용할 수 없습니다."
    
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    
    try:
        response = requests.get(export_url, timeout=timeout, verify=False)
        if response.status_code == 200:
            # data_only=True (값만)
            wb_v = load_workbook(BytesIO(response.content), data_only=True)
            # data_only=False (수식 포함)
            wb_f = load_workbook(BytesIO(response.content), data_only=False)
            return wb_v, wb_f, None
        else:
            return None, None, f"구글 스프레드시트 다운로드 실패 (HTTP 상태 코드: {response.status_code})"
    except requests.exceptions.Timeout:
        return None, None, f"구글 스프레드시트 연결 시간 초과 (타임아웃: {timeout}초). 인터넷 연결을 확인해주세요."
    except requests.exceptions.ConnectionError:
        return None, None, "구글 스프레드시트에 연결할 수 없습니다. 인터넷 연결을 확인해주세요."
    except requests.exceptions.RequestException as e:
        return None, None, f"구글 스프레드시트 요청 중 오류 발생: {str(e)}"
    except Exception as e:
        return None, None, f"구글 스프레드시트 처리 중 예상치 못한 오류 발생: {str(e)}"
    
    return None, None, "알 수 없는 오류가 발생했습니다."


def normalize_course_name(name: str) -> str:
    """괄호( ) 안 내용을 제거하고, 양끝 공백만 제거(내부 공백은 유지)."""
    if name is None:
        return ""
    s = str(name)
    s = re.sub(r"\([^)]*\)", "", s)  # ( ... ) 제거
    return s.strip()


def split_bidirectional(name: str):
    """'음악↔미술' 같은 문자열을 ['음악','미술']로 분해(양쪽 공백 제거)."""
    if name is None:
        return []
    s = str(name)
    if "↔" not in s:
        return []
    parts = [p.strip() for p in s.split("↔")]
    return [p for p in parts if p != ""]


def is_error_token(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().upper()
    return s in ("#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def to_number(value):
    """숫자 변환(정수/실수). 실패 시 None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value):  # NaN
            return None
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def find_sheet_for_year(sheetnames, year: int):
    """
    '2026 입학생...' 또는 '2026학년도 입학생...' 등:
    - 시트명이 year로 시작하고 '입학생'을 포함하면 매칭
    - '예시'가 포함된 시트는 제외
    """
    y = str(year)
    candidates = [n for n in sheetnames if n.startswith(y) and ("입학생" in n) and ("예시" not in n)]
    if candidates:
        return candidates[0]
    for n in sheetnames:
        if "예시" in n:
            continue
        n2 = n.replace(" ", "")
        if n2.startswith(f"{y}입학생"):
            return n
    return None


def find_hidden_sheet_name(sheetnames):
    """'숨김' 시트를 찾음(정확 일치 우선, 포함은 차선)."""
    if "숨김" in sheetnames:
        return "숨김"
    for n in sheetnames:
        if "숨김" in n:
            return n
    return None


def find_all_grades_sheet(sheetnames):
    """'2026 전학년' 시트를 찾음. '예시'가 포함된 시트는 제외."""
    for n in sheetnames:
        if "예시" in n:
            continue
        n_normalized = n.replace(" ", "")
        if "2026" in n and "전학년" in n_normalized:
            return n
    return None


def load_reference_sheets_from_google():
    """
    구글 스프레드시트에서 '숨김', '전문교과목록' 시트 로드
    return: (wb_v, wb_f, success, error_msg)
    """
    wb_v, wb_f, error_msg = load_workbook_from_google_sheet(GOOGLE_SHEET_ID, timeout=10)
    
    if wb_v is None or wb_f is None:
        return None, None, False, error_msg or "온라인 데이터에 연결할 수 없습니다."
    
    sheetnames = wb_v.sheetnames
    has_hidden = "숨김" in sheetnames
    has_vocational = "전문교과목록" in sheetnames
    
    if not has_hidden:
        return None, None, False, "온라인 데이터에 '숨김' 시트가 없습니다."
    
    return wb_v, wb_f, True, None


def build_merged_lookup(ws):
    """
    셀 좌표(행,열) -> (min_row, min_col, max_row, max_col) 매핑을 만든다.
    """
    lookup = {}
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = (min_row, min_col, max_row, max_col)
    return lookup


def get_value_with_merge(ws_values, ws_formula, merged_lookup, row, col):
    """
    data_only 값(ws_values) 기준으로:
    - 해당 셀이 병합 영역이면 top-left 값으로 보정
    - 동시에 수식(ws_formula)의 존재도 반환
    return: (value, formula_str_or_none, (used_row, used_col))
    """
    used_row, used_col = row, col
    key = (row, col)

    if key in merged_lookup:
        min_row, min_col, _, _ = merged_lookup[key]
        used_row, used_col = min_row, min_col

    v = ws_values.cell(used_row, used_col).value
    f = ws_formula.cell(used_row, used_col).value
    formula = f if isinstance(f, str) and f.startswith("=") else None
    return v, formula, (used_row, used_col)


def find_hidden_header_row(ws_values, ws_formula, merged_lookup):
    """숨김 시트에서 '과목명' 헤더 행 찾기(기본 2행)."""
    for r in range(1, 21):
        v, _, _ = get_value_with_merge(ws_values, ws_formula, merged_lookup, r, 2)  # B열
        if v is not None and str(v).strip() == "과목명":
            return r
    return 2


def safe_strip(v):
    if v is None:
        return ""
    return str(v).strip()


# =========================
# 핵심 검사 로직
# =========================

def find_marker_row(ws_values, ws_formula, merged_lookup, marker_text, search_col=1):
    """
    특정 텍스트를 포함하는 행 찾기 (A열 기본)
    marker_text: 찾을 텍스트 (예: '학교지정과목교과편성', '학생선택과목교과편성')
    """
    for r in range(1, ws_values.max_row + 1):
        v, _, _ = get_value_with_merge(ws_values, ws_formula, merged_lookup, r, search_col)
        if v is not None:
            v_str = str(v).strip().replace(" ", "")
            marker_normalized = marker_text.replace(" ", "")
            if marker_normalized in v_str:
                return r
    return None


def get_column_name(col_num, year=None):
    """
    열 번호를 한글 이름으로 변환
    
    2025/2026 시트:
    - B열(2): 교과(군)
    - C열(3): 과목유형
    - D열(4): 과목명
    - E열(5): 기본학점
    - F열(6): 운영학점
    - O열(15): 성적처리
    - G~L열(7~12): 학기별 학점
    
    2024 시트:
    - C열(3): 교과(군)
    - D열(4): 과목유형
    - E열(5): 과목명
    - F열(6): 기본학점
    - G열(7): 운영학점
    - P열(16): 성적처리
    - H~M열(8~13): 학기별 학점
    """
    if year == 2024:
        # 2024 시트 열 구조
        col_names_2024 = {
            3: "교과(군)",
            4: "과목유형",
            5: "과목명",
            6: "기본학점",
            7: "운영학점",
            16: "성적처리"
        }
        
        # H~M 열은 학기 학점 (구체적인 학년/학기 정보 포함)
        if col_num == 8:  # H열
            return "1학년 1학기"
        elif col_num == 9:  # I열
            return "1학년 2학기"
        elif col_num == 10:  # J열
            return "2학년 1학기"
        elif col_num == 11:  # K열
            return "2학년 2학기"
        elif col_num == 12:  # L열
            return "3학년 1학기"
        elif col_num == 13:  # M열
            return "3학년 2학기"
        
        return col_names_2024.get(col_num, f"{chr(64 + col_num)}열")
    else:
        # 2025/2026 시트 열 구조
        col_names = {
            2: "교과(군)",
            3: "과목유형",
            4: "과목명",
            5: "기본학점",
            6: "운영학점",
            15: "성적처리"
        }
        
        # G~L 열은 학기 학점 (구체적인 학년/학기 정보 포함)
        if col_num == 7:  # G열
            return "1학년 1학기"
        elif col_num == 8:  # H열
            return "1학년 2학기"
        elif col_num == 9:  # I열
            return "2학년 1학기"
        elif col_num == 10:  # J열
            return "2학년 2학기"
        elif col_num == 11:  # K열
            return "3학년 1학기"
        elif col_num == 12:  # L열
            return "3학년 2학기"
        
        return col_names.get(col_num, f"{chr(64 + col_num)}열")


def format_number(num):
    """
    숫자를 정수 형태로 포맷팅 (소수점 없이)
    """
    if num is None:
        return "None"
    if isinstance(num, (int, float)):
        # 정수로 변환 가능하면 정수로 표시
        if num == int(num):
            return str(int(num))
        else:
            return str(int(num))  # 강제로 정수로 변환
    return str(num)


def check_all_grades_sheet(wb_v, wb_f, targets, issues, hidden=None, vocational_courses=None, new_courses=None, hidden_list_norm=None):
    """
    '2026 전학년' 시트 검증
    - 전학년 시트와 2026 입학생 시트: G, H열 비교 (1학년)
    - 전학년 시트와 2025 입학생 시트: I, J열 비교 (2학년)
    - 전학년 시트 K, L열과 2024 입학생 시트 L, M열 비교 (3학년)
    - 전학년 시트의 '증배' 관련 과목 검증
    """
    sheetnames = wb_v.sheetnames
    all_grades_sheet = find_all_grades_sheet(sheetnames)
    
    if not all_grades_sheet:
        issues.append({
            "severity": "WARNING",
            "sheet": "2026 전학년",
            "row": "-",
            "message": "'2026 전학년' 시트를 찾을 수 없습니다."
        })
        return
    
    # '2026 전학년' 시트 로드
    ws_all_v = wb_v[all_grades_sheet]
    ws_all_f = wb_f[all_grades_sheet]
    merge_all = build_merged_lookup(ws_all_f)
    
    # 각 입학생 시트 로드
    sheets_data = {}
    for year in [2026, 2025, 2024]:
        sname = targets.get(year)
        if not sname:
            continue
        ws_v = wb_v[sname]
        ws_f = wb_f[sname]
        merge = build_merged_lookup(ws_f)
        sheets_data[year] = {
            "name": sname,
            "ws_v": ws_v,
            "ws_f": ws_f,
            "merge": merge
        }
    
    # ===== 1. '학교 지정 과목 교과~' 위쪽 검증 =====
    marker_row_all = find_marker_row(ws_all_v, ws_all_f, merge_all, "학교지정과목교과편성")
    
    if marker_row_all:
        # '2026 전학년' 시트의 교과목 수집 (marker_row_all 위쪽)
        # 화살표가 있는 과목의 경우 여러 행이 있을 수 있으므로 리스트로 저장
        all_grades_courses = {}  # {과목명_정규화: [{row, B~L열, O열 값}, ...]}
        
        for r in range(5, marker_row_all):
            # A열에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
            a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, r, 1)
            if a_col_value and '증배' in str(a_col_value):
                continue
            
            course_raw, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, r, 4)  # D열
            if not course_raw or str(course_raw).strip() == "":
                continue
            
            course_norm = normalize_course_name(course_raw)
            if not course_norm:
                continue
            
            # B~L열(2~12), O열(15) 값 수집
            row_data = {"row": r}
            for col in range(2, 13):  # B~L
                v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, r, col)
                row_data[col] = safe_strip(v) if col in [2, 3] else to_number(v)
            v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, r, 15)  # O열
            row_data[15] = safe_strip(v)
            
            # 같은 과목명이 여러 행에 있을 수 있음 (화살표 과목)
            if course_norm not in all_grades_courses:
                all_grades_courses[course_norm] = []
            all_grades_courses[course_norm].append(row_data)
        
        # 각 입학생 시트 검증
        for year in [2026, 2025, 2024]:
            if year not in sheets_data:
                continue
            
            data = sheets_data[year]
            ws_v, ws_f, merge = data["ws_v"], data["ws_f"], data["merge"]
            sname = data["name"]
            
            # 해당 시트의 '학교 지정 과목 교과~' 찾기
            marker_row_src = find_marker_row(ws_v, ws_f, merge, "학교지정과목교과편성")
            if not marker_row_src:
                continue
            
            # 검사할 열 결정
            if year == 2026:
                check_cols = [7, 8]  # G, H
                src_cols = list(range(2, 13)) + [15]  # B~L, O
                dst_cols = list(range(2, 13)) + [15]  # B~L, O
                course_col = 4  # D열
            elif year == 2025:
                check_cols = [9, 10]  # I, J
                src_cols = list(range(2, 13)) + [15]  # B~L, O
                dst_cols = list(range(2, 13)) + [15]  # B~L, O
                course_col = 4  # D열
            else:  # 2024
                check_cols = [12, 13]  # L, M (2024 입학생 시트)
                src_cols = list(range(3, 14)) + [16]  # C~M, P
                dst_cols = list(range(2, 13)) + [15]  # B~L, O (2026 전학년 기준)
                course_col = 5  # E열 (2024는 과목명이 E열에 있음)
            
            # 행별 검사 - 같은 과목명의 행 순서를 추적
            course_row_index = {}  # {과목명: 현재 인덱스}
            
            # A열(또는 B열) 확인용
            a_col = 1 if year in [2026, 2025] else 2
            
            for r in range(5, marker_row_src):
                # A열(또는 B열)에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
                a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                if a_col_value and '증배' in str(a_col_value):
                    continue
                
                course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, course_col)
                if not course_raw or str(course_raw).strip() == "":
                    continue
                
                # check_cols에 숫자가 있는지 확인
                has_number = False
                for col in check_cols:
                    v, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, col)
                    if to_number(v) is not None:
                        has_number = True
                        break
                
                if not has_number:
                    continue
                
                course_norm = normalize_course_name(course_raw)
                
                # '2026 전학년' 시트에 있는지 확인
                if course_norm not in all_grades_courses:
                    msg_line1 = f"'{sname}' 시트 {r}행의 '{course_norm}' 과목이 '2026 전학년' 시트에 없습니다."
                    msg_line2 = "      선택 미달 등으로 개설되지 않은 경우도 2026 전학년 시트에 추가하고 개설여부에 X해주세요."
                    issues.append({
                        "severity": "ERROR",
                        "sheet": all_grades_sheet,
                        "row": "-",
                        "message": msg_line1 + "\n" + msg_line2
                    })
                    continue
                
                # 같은 과목명의 몇 번째 행인지 확인 (화살표 과목 대응)
                if course_norm not in course_row_index:
                    course_row_index[course_norm] = 0
                else:
                    course_row_index[course_norm] += 1
                
                idx = course_row_index[course_norm]
                all_data_list = all_grades_courses[course_norm]
                
                # 인덱스가 범위를 벗어나면 마지막 데이터 사용
                if idx >= len(all_data_list):
                    idx = len(all_data_list) - 1
                
                all_data = all_data_list[idx]
                for i, src_col in enumerate(src_cols):
                    dst_col = dst_cols[i]
                    
                    src_val, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, src_col)
                    dst_val = all_data.get(dst_col)
                    
                    # B열은 병합 고려
                    if src_col in [2, 3] or dst_col in [2, 3]:  # 문자열 비교
                        src_str = safe_strip(src_val)
                        dst_str = safe_strip(dst_val) if isinstance(dst_val, str) else str(dst_val) if dst_val is not None else ""
                        
                        if src_str != dst_str:
                            src_col_name = get_column_name(src_col, year)
                            dst_col_name = get_column_name(dst_col)  # 전학년 시트는 year 없음
                            issues.append({
                                "severity": "ERROR",
                                "sheet": all_grades_sheet,
                                "row": all_data["row"],
                                "message": f"'{sname}' 시트 {r}행의 {src_col_name}('{src_str}')과 '2026 전학년' 시트의 {dst_col_name}('{dst_str}')이 일치하지 않습니다."
                            })
                    elif src_col == 15 or dst_col == 15 or src_col == 16 or dst_col == 16:  # O열/P열 (문자열)
                        src_str = safe_strip(src_val)
                        dst_str = safe_strip(dst_val) if isinstance(dst_val, str) else str(dst_val) if dst_val is not None else ""
                        
                        if src_str != dst_str:
                            src_col_name = get_column_name(src_col, year)
                            dst_col_name = get_column_name(dst_col)  # 전학년 시트는 year 없음
                            issues.append({
                                "severity": "ERROR",
                                "sheet": all_grades_sheet,
                                "row": all_data["row"],
                                "message": f"'{sname}' 시트 {r}행의 {src_col_name}('{src_str}')과 '2026 전학년' 시트의 {dst_col_name}('{dst_str}')이 일치하지 않습니다."
                            })
                    else:  # 숫자 비교
                        src_num = to_number(src_val)
                        dst_num = dst_val if isinstance(dst_val, (int, float)) else to_number(dst_val)
                        
                        if src_num is not None and dst_num is not None:
                            if abs(src_num - dst_num) > EPS:
                                src_col_name = get_column_name(src_col, year)
                                dst_col_name = get_column_name(dst_col)  # 전학년 시트는 year 없음
                                src_num_str = format_number(src_num)
                                dst_num_str = format_number(dst_num)
                                issues.append({
                                    "severity": "ERROR",
                                    "sheet": all_grades_sheet,
                                    "row": all_data["row"],
                                    "message": f"'{sname}' 시트 {r}행의 {src_col_name}({src_num_str})과 '2026 전학년' 시트의 {dst_col_name}({dst_num_str})이 일치하지 않습니다."
                                })
                        elif src_num is not None or dst_num is not None:
                            src_col_name = get_column_name(src_col, year)
                            dst_col_name = get_column_name(dst_col)  # 전학년 시트는 year 없음
                            src_num_str = format_number(src_num)
                            dst_num_str = format_number(dst_num)
                            issues.append({
                                "severity": "ERROR",
                                "sheet": all_grades_sheet,
                                "row": all_data["row"],
                                "message": f"'{sname}' 시트 {r}행의 {src_col_name}({src_num_str})과 '2026 전학년' 시트의 {dst_col_name}({dst_num_str})이 일치하지 않습니다."
                            })
        
        # 역방향 검증: '2026 전학년' 시트에만 있고 입학생 시트에 없는 경우
        for course_norm, data_list in all_grades_courses.items():
            # data_list는 리스트이므로 첫 번째 항목 사용
            if not data_list:
                continue
            
            # 전학년 시트의 첫 번째 행에서 A열에 '증배'가 있는지 확인 (교차 점검 제외)
            first_data = data_list[0]
            a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, first_data["row"], 1)
            if a_col_value and '증배' in str(a_col_value):
                continue
            
            found_in_any = False
            
            for year in [2026, 2025, 2024]:
                if year not in sheets_data:
                    continue
                
                sheet_data = sheets_data[year]
                ws_v, ws_f, merge = sheet_data["ws_v"], sheet_data["ws_f"], sheet_data["merge"]
                sname = sheet_data["name"]
                
                marker_row_src = find_marker_row(ws_v, ws_f, merge, "학교지정과목교과편성")
                if not marker_row_src:
                    continue
                
                # 해당 학년의 열 확인
                if year == 2026:
                    check_cols = [7, 8]
                    rev_course_col = 4  # D열
                elif year == 2025:
                    check_cols = [9, 10]
                    rev_course_col = 4  # D열
                else:  # 2024
                    check_cols = [12, 13]  # L, M열 (2024는 H~M이 학기별 열이므로 3학년은 L, M)
                    rev_course_col = 5  # E열 (2024는 과목명이 E열에 있음)
                
                # A열(또는 B열) 확인용
                a_col = 1 if year in [2026, 2025] else 2
                
                # 입학생 시트에서 해당 과목이 check_cols에 숫자가 있는지 확인
                for r in range(5, marker_row_src):
                    # A열(또는 B열)에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
                    a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                    if a_col_value and '증배' in str(a_col_value):
                        continue
                    
                    course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, rev_course_col)
                    if not course_raw:
                        continue
                    
                    if normalize_course_name(course_raw) == course_norm:
                        # check_cols에 숫자가 있는지 확인
                        for col in check_cols:
                            v, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, col)
                            if to_number(v) is not None:
                                found_in_any = True
                                break
                    
                    if found_in_any:
                        break
                
                if found_in_any:
                    break
            
            if not found_in_any:
                # data_list의 첫 번째 항목에서 행 번호 가져오기
                first_data = data_list[0]
                issues.append({
                    "severity": "ERROR",
                    "sheet": all_grades_sheet,
                    "row": first_data["row"],
                    "message": f"'2026 전학년' 시트의 '{course_norm}' 과목이 해당 입학생 시트의 학년 열에 없습니다."
                })
    
    # ===== 2. '학생 선택 과목 교과~' 위쪽 검증 =====
    marker_row_student = find_marker_row(ws_all_v, ws_all_f, merge_all, "학생선택과목교과편성")
    
    if marker_row_student and marker_row_all:
        # '2026 전학년' 시트의 교과목 수집 (marker_row_all ~ marker_row_student 사이)
        # A열 병합 구간별로 과목 수집
        student_courses_by_year = {
            2026: {
                'G': {},  # G열(7)에 숫자가 있는 과목 (1학년 1학기) - 리스트로 저장
                'H': {}   # H열(8)에 숫자가 있는 과목 (1학년 2학기) - 리스트로 저장
            },
            2025: {
                'I': {},  # I열(9)에 숫자가 있는 과목 (2학년 1학기) - 리스트로 저장
                'J': {}   # J열(10)에 숫자가 있는 과목 (2학년 2학기) - 리스트로 저장
            },
            2024: {
                'K': {},  # K열(11)에 숫자가 있는 과목 (3학년 1학기) - 리스트로 저장
                'L': {}   # L열(12)에 숫자가 있는 과목 (3학년 2학기) - 리스트로 저장
            }
        }
        
        processed_merges_all = set()
        
        for r in range(marker_row_all + 1, marker_row_student):
            # 먼저 A열 병합 확인 (전학년 시트는 A열 사용)
            key = (r, 1)  # A열
            if key in merge_all:
                min_row, _, max_row, _ = merge_all[key]
                # A열 병합에 '증배'가 포함되어 있는지 확인
                a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, r, 1)
                if a_col_value and '증배' in str(a_col_value):
                    # '증배'가 있으면 교차 점검 제외
                    continue
                else:
                    merge_key = (min_row, max_row)
            else:
                # A열에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
                a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, r, 1)
                if a_col_value and '증배' in str(a_col_value):
                    continue
                merge_key = (r, r)
            
            if merge_key in processed_merges_all:
                continue
            
            processed_merges_all.add(merge_key)
            
            # 병합 구간 내에서 각 학년별 열에 숫자가 하나라도 있는지 확인
            has_2026_G = False  # G열 (7) - 1학년 1학기
            has_2026_H = False  # H열 (8) - 1학년 2학기
            has_2025_I = False  # I열 (9) - 2학년 1학기
            has_2025_J = False  # J열 (10) - 2학년 2학기
            has_2024_K = False  # K열 (11) - 3학년 1학기
            has_2024_L = False  # L열 (12) - 3학년 2학기
            
            for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student)):
                # 2026 (1학년): G열 (7) - 1학기
                v_G, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 7)
                if to_number(v_G) is not None:
                    has_2026_G = True
                
                # 2026 (1학년): H열 (8) - 2학기
                v_H, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 8)
                if to_number(v_H) is not None:
                    has_2026_H = True
                
                # 2025 (2학년): I열 (9) - 1학기
                v_I, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 9)
                if to_number(v_I) is not None:
                    has_2025_I = True
                
                # 2025 (2학년): J열 (10) - 2학기
                v_J, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 10)
                if to_number(v_J) is not None:
                    has_2025_J = True
                
                # 2024 (3학년): K열 (11) - 1학기
                v_K, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 11)
                if to_number(v_K) is not None:
                    has_2024_K = True
                
                # 2024 (3학년): L열 (12) - 2학기
                v_L, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 12)
                if to_number(v_L) is not None:
                    has_2024_L = True
            
            # 숫자가 없으면 건너뛰기
            if not (has_2026_G or has_2026_H or has_2025_I or has_2025_J or has_2024_K or has_2024_L):
                continue
            
            # 해당 병합 구간 내에서 각 학년별로 숫자가 있는 행만 과목 수집
            for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student)):
                c_raw, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 4)  # D열
                if not c_raw:
                    continue
                
                # 총계 행 같은 키워드가 포함된 경우 제외
                if any(keyword in str(c_raw) for keyword in ["편성학점", "총교과", "창의적체험", "편성학점수"]):
                    continue
                
                course_norm_rr = normalize_course_name(c_raw)
                if not course_norm_rr:  # 빈 문자열이 아닌 경우만 추가
                    continue
                
                # 병합 구간 전체에서 해당 학년 열에 숫자가 하나라도 있으면, 이 행의 과목도 수집
                # (각 행의 개별 숫자 확인이 아니라, 병합 구간 전체의 숫자 확인 결과 사용)
                if has_2026_G or has_2026_H or has_2025_I or has_2025_J or has_2024_K or has_2024_L:
                    # B~L열, O열 값 수집
                    row_data = {"row": rr}
                    for col in range(2, 13):  # B~L열 (2~12)
                        v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col)
                        row_data[col] = safe_strip(v) if col in [2, 3] else to_number(v)
                    v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 15)
                    row_data[15] = safe_strip(v)
                    
                    # 병합 정보 저장
                    row_data["merge_start"] = merge_key[0]
                    row_data["merge_end"] = merge_key[1]
                    
                    # 각 학년별로 분류 (리스트로 추가)
                    # 병합 구간 전체에서 해당 학년 열에 숫자가 있으면 추가
                    if has_2026_G:
                        if course_norm_rr not in student_courses_by_year[2026]['G']:
                            student_courses_by_year[2026]['G'][course_norm_rr] = []
                        student_courses_by_year[2026]['G'][course_norm_rr].append(row_data)
                    if has_2026_H:
                        if course_norm_rr not in student_courses_by_year[2026]['H']:
                            student_courses_by_year[2026]['H'][course_norm_rr] = []
                        student_courses_by_year[2026]['H'][course_norm_rr].append(row_data)
                    if has_2025_I:
                        if course_norm_rr not in student_courses_by_year[2025]['I']:
                            student_courses_by_year[2025]['I'][course_norm_rr] = []
                        student_courses_by_year[2025]['I'][course_norm_rr].append(row_data)
                    if has_2025_J:
                        if course_norm_rr not in student_courses_by_year[2025]['J']:
                            student_courses_by_year[2025]['J'][course_norm_rr] = []
                        student_courses_by_year[2025]['J'][course_norm_rr].append(row_data)
                    if has_2024_K:
                        if course_norm_rr not in student_courses_by_year[2024]['K']:
                            student_courses_by_year[2024]['K'][course_norm_rr] = []
                        student_courses_by_year[2024]['K'][course_norm_rr].append(row_data)
                    if has_2024_L:
                        if course_norm_rr not in student_courses_by_year[2024]['L']:
                            student_courses_by_year[2024]['L'][course_norm_rr] = []
                        student_courses_by_year[2024]['L'][course_norm_rr].append(row_data)
        
        # 각 입학생 시트 검증
        for year in [2026, 2025, 2024]:
            if year not in sheets_data:
                continue
            
            data = sheets_data[year]
            ws_v, ws_f, merge = data["ws_v"], data["ws_f"], data["merge"]
            sname = data["name"]
            
            # 해당 시트의 마커 찾기
            marker_row_src = find_marker_row(ws_v, ws_f, merge, "학교지정과목교과편성")
            if not marker_row_src:
                continue
            
            # '학생 선택 과목 편성' 마커 찾기 (이 위쪽으로만 비교)
            marker_row_student_end = find_marker_row(ws_v, ws_f, merge, "학생선택과목교과편성")
            if not marker_row_student_end:
                # 마커를 못 찾으면 시트 끝까지
                marker_row_student_end = ws_v.max_row + 1
            
            # 검사할 열 결정 (입학생 시트에서 확인할 열)
            if year == 2026:
                # 2026년은 G열과 H열을 각각 별도로 처리
                check_cols_list = [[7], [8]]  # G열, H열을 각각 별도로
                student_course_col = 4  # D열
            elif year == 2025:
                # 2025년은 I열과 J열을 각각 별도로 처리
                check_cols_list = [[9], [10]]  # I열, J열을 각각 별도로
                student_course_col = 4  # D열
            else:  # 2024
                # 2024년은 L열과 M열을 각각 별도로 처리 (입학생 시트 기준)
                check_cols_list = [[12], [13]]  # L열, M열을 각각 별도로
                student_course_col = 5  # E열 (2024는 과목명이 E열에 있음)
            
            a_col = 1 if year in [2026, 2025] else 2  # '증배' 확인용
            
            # 모든 학년을 각 열별로 처리
            if year in [2026, 2025, 2024]:
                for col_idx, check_cols in enumerate(check_cols_list):
                    check_col = check_cols[0]  # 각 학년별 열
                    # col_key 결정: 전학년 시트의 열 키
                    if year == 2026:
                        col_key = 'G' if check_col == 7 else 'H'  # G열(7) 또는 H열(8)
                    elif year == 2025:
                        col_key = 'I' if check_col == 9 else 'J'  # I열(9) 또는 J열(10)
                    else:  # 2024
                        # 입학생 시트 L(12) -> 전학년 시트 K(11), 입학생 시트 M(13) -> 전학년 시트 L(12)
                        col_key = 'K' if check_col == 12 else 'L'  # K열(전학년 11) 또는 L열(전학년 12)
                    
                    processed_merges = set()
                    
                    for r in range(marker_row_src + 1, marker_row_student_end):
                        # A열(또는 B열) 병합 확인
                        key = (r, a_col)
                        if key in merge:
                            min_row, _, max_row, _ = merge[key]
                            # A열 병합에 '증배'가 포함되어 있는지 확인
                            a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                            if a_col_value and '증배' in str(a_col_value):
                                # '증배'가 있으면 교차 점검 제외
                                continue
                            else:
                                merge_key = (min_row, max_row)
                        else:
                            # A열(또는 B열)에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
                            a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                            if a_col_value and '증배' in str(a_col_value):
                                continue
                            merge_key = (r, r)
                        
                        course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, student_course_col)
                        if not course_raw or str(course_raw).strip() == "":
                            continue
                        
                        # 해당 열에 숫자가 있는지 확인
                        v, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, check_col)
                        if to_number(v) is None:
                            continue
                        
                        course_norm = normalize_course_name(course_raw)
                        if not course_norm:  # 빈 문자열인 경우 건너뛰기
                            continue
                        
                        # 총계 행 같은 키워드가 포함된 경우 제외
                        if any(keyword in str(course_raw) for keyword in ["편성학점", "총교과", "창의적체험", "편성학점수"]):
                            continue
                        
                        if merge_key in processed_merges:
                            continue
                        
                        processed_merges.add(merge_key)
                        
                        # 해당 병합 구간의 모든 과목 수집
                        courses_in_merge = []  # (과목명, 행번호) 튜플 리스트
                        
                        # 먼저 병합 구간에서 check_col에 숫자가 있는지 확인
                        has_number_in_merge = False
                        for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student_end)):
                            v, _, _ = get_value_with_merge(ws_v, ws_f, merge, rr, check_col)
                            if to_number(v) is not None:
                                has_number_in_merge = True
                                break
                        
                        # 병합 구간 전체에서 숫자가 하나라도 있으면, 병합 구간의 모든 행에서 과목 수집
                        if has_number_in_merge:
                            for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student_end)):
                                c_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge, rr, student_course_col)
                                if not c_raw:
                                    continue
                                
                                # 총계 행 같은 키워드가 포함된 경우 제외
                                if any(keyword in str(c_raw) for keyword in ["편성학점", "총교과", "창의적체험", "편성학점수"]):
                                    continue
                                
                                normalized = normalize_course_name(c_raw)
                                if normalized:  # 빈 문자열이 아닌 경우만 추가
                                    courses_in_merge.append((normalized, rr))  # 행 번호와 함께 저장
                        
                        # '2026 전학년' 시트에 해당 과목들이 모두 있는지 확인 (해당 열 기준)
                        student_courses = student_courses_by_year.get(year, {}).get(col_key, {})
                        
                        # courses_in_merge는 이미 (과목명, 행번호) 튜플 리스트로 순서가 유지됨
                        course_occurrence_count = {}  # 각 과목명이 몇 번째 나왔는지 추적
                        
                        for cn, course_row in courses_in_merge:
                            if cn not in student_courses:
                                # 각 열별로 적절한 열 이름 표시
                                if year == 2026:
                                    col_name = "G열(1학년 1학기)" if col_key == 'G' else "H열(1학년 2학기)"
                                elif year == 2025:
                                    col_name = "I열(2학년 1학기)" if col_key == 'I' else "J열(2학년 2학기)"
                                else:  # 2024
                                    col_name = "K열(3학년 1학기)" if col_key == 'K' else "L열(3학년 2학기)"
                                msg_line1 = f"'{sname}' 시트의 '{cn}' 과목({course_row}행)이 '2026 전학년' 시트의 {col_name}에 없습니다."
                                msg_line2 = "선택 미달 등으로 개설되지 않은 경우도 2026 전학년 시트에 추가하고 개설여부에 X해주세요."
                                issues.append({
                                    "severity": "ERROR",
                                    "sheet": all_grades_sheet,
                                    "row": "-",
                                    "message": msg_line1 + "\n" + msg_line2
                                })
                            else:
                                # 과목이 여러 번 나오는 경우를 고려하여 순서대로 매칭
                                all_data_list = student_courses[cn]  # 리스트
                                
                                # 현재 과목명의 몇 번째 출현인지 확인
                                if cn not in course_occurrence_count:
                                    course_occurrence_count[cn] = 0
                                occurrence_index = course_occurrence_count[cn]
                                course_occurrence_count[cn] += 1
                                
                                # 리스트 범위 확인
                                if occurrence_index >= len(all_data_list):
                                    continue
                                
                                all_data = all_data_list[occurrence_index]
                                
                                # A열에 '증배'가 포함되어 있는지 확인
                                a_col_for_check = 1 if year in [2026, 2025] else 2
                                a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, course_row, a_col_for_check)
                                has_jeungbae = a_col_value and '증배' in str(a_col_value)
                                
                                # 비교할 열 결정
                                if year == 2026:
                                    # 2026년: G(7), H(8) 열 비교
                                    src_cols = [2, 3, 5, 6, check_col, 15]
                                    dst_cols = [2, 3, 5, 6, check_col, 15]
                                elif year == 2025:
                                    # 2025년: I(9), J(10) 열 비교
                                    src_cols = [2, 3, 5, 6, check_col, 15]
                                    dst_cols = [2, 3, 5, 6, check_col, 15]
                                else:  # 2024
                                    # 2024년: 입학생 시트 L(12), M(13) -> 전학년 시트 K(11), L(12)
                                    dst_check_col = 11 if check_col == 12 else 12  # 전학년 시트 열
                                    src_cols = [3, 4, 6, 7, check_col, 16]
                                    dst_cols = [2, 3, 5, 6, dst_check_col, 15]
                                
                                # 각 열 비교
                                for i, src_col in enumerate(src_cols):
                                    dst_col = dst_cols[i]
                                    
                                    src_val, _, _ = get_value_with_merge(ws_v, ws_f, merge, course_row, src_col)
                                    dst_val = all_data.get(dst_col)
                                    
                                    # 문자열 비교 (교과군, 과목유형)
                                    if src_col in [2, 3] or dst_col in [2, 3]:
                                        # '증배'가 있고 교과(군) 열인 경우 비교 건너뛰기
                                        if has_jeungbae and (src_col == 2 or dst_col == 2 or (year == 2024 and (src_col == 3 or dst_col == 3))):
                                            is_gyogwa_col = (src_col == 2 or dst_col == 2) if year in [2026, 2025] else (src_col == 3 or dst_col == 3)
                                            if is_gyogwa_col:
                                                continue
                                        
                                        src_str = safe_strip(src_val)
                                        dst_str = safe_strip(dst_val) if isinstance(dst_val, str) else str(dst_val) if dst_val is not None else ""
                                        
                                        if src_str != dst_str:
                                            src_col_name = get_column_name(src_col, year)
                                            dst_col_name = get_column_name(dst_col)
                                            issues.append({
                                                "severity": "ERROR",
                                                "sheet": all_grades_sheet,
                                                "row": all_data["row"],
                                                "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}('{src_str}')과 '2026 전학년' 시트의 {dst_col_name}('{dst_str}')이 일치하지 않습니다."
                                            })
                                    # 성적처리 비교
                                    elif src_col == 15 or dst_col == 15:
                                        src_str = safe_strip(src_val)
                                        dst_str = safe_strip(dst_val) if isinstance(dst_val, str) else str(dst_val) if dst_val is not None else ""
                                        
                                        if src_str != dst_str:
                                            src_col_name = get_column_name(src_col, year)
                                            dst_col_name = get_column_name(dst_col)
                                            issues.append({
                                                "severity": "ERROR",
                                                "sheet": all_grades_sheet,
                                                "row": all_data["row"],
                                                "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}('{src_str}')과 '2026 전학년' 시트의 {dst_col_name}('{dst_str}')이 일치하지 않습니다."
                                            })
                                    # 숫자 비교 (기본학점, 운영학점, 학기학점)
                                    else:
                                        src_num = to_number(src_val)
                                        dst_num = dst_val if isinstance(dst_val, (int, float)) else to_number(dst_val)
                                        
                                        if src_num is not None and dst_num is not None:
                                            if abs(src_num - dst_num) > EPS:
                                                src_col_name = get_column_name(src_col, year)
                                                dst_col_name = get_column_name(dst_col)
                                                src_num_str = format_number(src_num)
                                                dst_num_str = format_number(dst_num)
                                                issues.append({
                                                    "severity": "ERROR",
                                                    "sheet": all_grades_sheet,
                                                    "row": all_data["row"],
                                                    "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}({src_num_str})과 '2026 전학년' 시트의 {dst_col_name}({dst_num_str})이 일치하지 않습니다."
                                                })
                                        elif src_num is not None or dst_num is not None:
                                            src_col_name = get_column_name(src_col, year)
                                            dst_col_name = get_column_name(dst_col)
                                            src_num_str = format_number(src_num)
                                            dst_num_str = format_number(dst_num)
                                            issues.append({
                                                "severity": "ERROR",
                                                "sheet": all_grades_sheet,
                                                "row": all_data["row"],
                                                "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}({src_num_str})과 '2026 전학년' 시트의 {dst_col_name}({dst_num_str})이 일치하지 않습니다."
                                            })
        
        # ===== 학생 선택 과목 정방향 검증 추가 =====
        # 각 입학생 시트의 학생 선택 영역 검증
        for year in [2026, 2025, 2024]:
            if year not in sheets_data:
                continue
            
            data = sheets_data[year]
            ws_v, ws_f, merge = data["ws_v"], data["ws_f"], data["merge"]
            sname = data["name"]
            
            # 학생 선택 과목 마커 찾기
            marker_row_student_start = find_marker_row(ws_v, ws_f, merge, "학생선택과목교과편성")
            if not marker_row_student_start:
                continue
            
            # 학생 선택 영역의 끝 찾기
            marker_row_student_real_end = ws_v.max_row + 1
            for r in range(marker_row_student_start + 1, ws_v.max_row + 1):
                course_v, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, 4 if year in [2026, 2025] else 5)
                if course_v:
                    course_str = str(course_v).strip().replace(" ", "")
                    if any(keyword in course_str for keyword in ["편성학점", "총교과", "창의적체험활동"]):
                        marker_row_student_real_end = r
                        break
            
            # 검사할 열 결정
            if year == 2026:
                # 2026년은 G열과 H열을 각각 별도로 처리
                check_cols_list = [[7], [8]]  # G열, H열을 각각 별도로
                student_course_col = 4  # D열
            elif year == 2025:
                # 2025년은 I열과 J열을 각각 별도로 처리
                check_cols_list = [[9], [10]]  # I열, J열을 각각 별도로
                student_course_col = 4  # D열
            else:  # 2024
                # 2024년은 L열과 M열을 각각 별도로 처리 (입학생 시트 기준)
                check_cols_list = [[12], [13]]  # L열, M열을 각각 별도로
                student_course_col = 5  # E열
            
            # A열(2026/2025) 또는 B열(2024) 병합 구간별로 과목 수집
            a_col = 1 if year in [2026, 2025] else 2  # A열 또는 B열
            
            # 모든 학년을 각 열별로 처리
            for col_idx, check_cols in enumerate(check_cols_list):
                check_col = check_cols[0]  # 각 학년별 열
                # col_key 결정: 전학년 시트의 열 키
                if year == 2026:
                    col_key = 'G' if check_col == 7 else 'H'  # G열(7) 또는 H열(8)
                elif year == 2025:
                    col_key = 'I' if check_col == 9 else 'J'  # I열(9) 또는 J열(10)
                else:  # 2024
                    # 입학생 시트 L(12) -> 전학년 시트 K(11), 입학생 시트 M(13) -> 전학년 시트 L(12)
                    col_key = 'K' if check_col == 12 else 'L'  # K열(전학년 11) 또는 L열(전학년 12)
                
                processed_merges_student = set()
                course_occurrence_count_student = {}
                
                for r in range(marker_row_student_start + 1, marker_row_student_real_end):
                    # A열(또는 B열) 병합 확인
                    key = (r, a_col)
                    if key in merge:
                        min_row, _, max_row, _ = merge[key]
                        a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                        if a_col_value and '증배' in str(a_col_value):
                            # '증배'가 있으면 교차 점검 제외
                            continue
                        else:
                            merge_key = (min_row, max_row)
                    else:
                        # A열(또는 B열)에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
                        a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                        if a_col_value and '증배' in str(a_col_value):
                            continue
                        merge_key = (r, r)
                    
                    if merge_key in processed_merges_student:
                        continue
                    
                    processed_merges_student.add(merge_key)
                    
                    # 해당 병합 구간에서 check_col에 숫자가 있는지 확인
                    has_number = False
                    for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student_real_end)):
                        v, _, _ = get_value_with_merge(ws_v, ws_f, merge, rr, check_col)
                        if to_number(v) is not None:
                            has_number = True
                            break
                    
                    # 해당 병합 구간의 모든 과목 수집
                    courses_in_merge_student = []  # (과목명, 행번호) 튜플 리스트
                    # 병합 구간 전체에서 숫자가 하나라도 있으면, 병합 구간의 모든 행에서 과목 수집
                    if has_number:
                        for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student_real_end)):
                            c_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge, rr, student_course_col)
                            if not c_raw or str(c_raw).strip() == "":
                                continue
                            
                            if any(keyword in str(c_raw) for keyword in ["편성학점", "총교과", "창의적체험", "편성학점수"]):
                                continue
                            
                            # 각 행의 check_col에 숫자가 있는지 확인하지 않고, 병합 구간 전체의 has_number 결과 사용
                            normalized = normalize_course_name(c_raw)
                            if normalized:
                                courses_in_merge_student.append((normalized, rr))
                    
                    # 전학년 시트에 해당 과목들이 있는지 확인
                    student_courses = student_courses_by_year.get(year, {}).get(col_key, {})
                    
                    for cn, course_row in courses_in_merge_student:
                        if cn not in student_courses:
                            # 각 열별로 적절한 열 이름 표시
                            if year == 2026:
                                col_name = "G열(1학년 1학기)" if col_key == 'G' else "H열(1학년 2학기)"
                            elif year == 2025:
                                col_name = "I열(2학년 1학기)" if col_key == 'I' else "J열(2학년 2학기)"
                            else:  # 2024
                                col_name = "K열(3학년 1학기)" if col_key == 'K' else "L열(3학년 2학기)"
                            msg_line1 = f"'{sname}' 시트의 '{cn}' 과목({course_row}행, 학생 선택)이 '2026 전학년' 시트에 없습니다."
                            msg_line2 = "선택 미달 등으로 개설되지 않은 경우도 2026 전학년 시트에 추가하고 개설여부에 X해주세요."
                            issues.append({
                                "severity": "ERROR",
                                "sheet": all_grades_sheet,
                                "row": "-",
                                "message": msg_line1 + "\n" + msg_line2
                            })
                        else:
                            # 순서 고려하여 매칭
                            all_data_list = student_courses[cn]
                            
                            if cn not in course_occurrence_count_student:
                                course_occurrence_count_student[cn] = 0
                            occurrence_index = course_occurrence_count_student[cn]
                            course_occurrence_count_student[cn] += 1
                            
                            if occurrence_index >= len(all_data_list):
                                continue
                            
                            all_data = all_data_list[occurrence_index]
                            
                            # A열에 '증배' 확인
                            a_col_for_check = 1 if year in [2026, 2025] else 2
                            a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, course_row, a_col_for_check)
                            has_jeungbae = a_col_value and '증배' in str(a_col_value)
                            
                            # 비교할 열 결정
                            if year == 2026:
                                # 2026년: G(7), H(8) 열 비교
                                src_cols = [2, 3, 5, 6, check_col, 15]
                                dst_cols = [2, 3, 5, 6, check_col, 15]
                            elif year == 2025:
                                # 2025년: I(9), J(10) 열 비교
                                src_cols = [2, 3, 5, 6, check_col, 15]
                                dst_cols = [2, 3, 5, 6, check_col, 15]
                            else:  # 2024
                                # 2024년: 입학생 시트 L(12), M(13) -> 전학년 시트 K(11), L(12)
                                dst_check_col = 11 if check_col == 12 else 12  # 전학년 시트 열
                                src_cols = [3, 4, 6, 7, check_col, 16]
                                dst_cols = [2, 3, 5, 6, dst_check_col, 15]
                            
                            # 각 열 비교
                            for i, src_col in enumerate(src_cols):
                                dst_col = dst_cols[i]
                                
                                src_val, _, _ = get_value_with_merge(ws_v, ws_f, merge, course_row, src_col)
                                dst_val = all_data.get(dst_col)
                                
                                # 문자열 비교
                                if src_col in [2, 3] or dst_col in [2, 3]:
                                    if has_jeungbae and (src_col == 2 or dst_col == 2 or (year == 2024 and (src_col == 3 or dst_col == 3))):
                                        is_gyogwa_col = (src_col == 2 or dst_col == 2) if year in [2026, 2025] else (src_col == 3 or dst_col == 3)
                                        if is_gyogwa_col:
                                            continue
                                    
                                    src_str = safe_strip(src_val)
                                    dst_str = safe_strip(dst_val) if isinstance(dst_val, str) else str(dst_val) if dst_val is not None else ""
                                    
                                    if src_str != dst_str:
                                        src_col_name = get_column_name(src_col, year)
                                        dst_col_name = get_column_name(dst_col)
                                        issues.append({
                                            "severity": "ERROR",
                                            "sheet": all_grades_sheet,
                                            "row": all_data["row"],
                                            "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}('{src_str}')과 '2026 전학년' 시트의 {dst_col_name}('{dst_str}')이 일치하지 않습니다."
                                        })
                                # 성적처리 비교
                                elif src_col == 15 or dst_col == 15 or src_col == 16 or dst_col == 16:
                                    src_str = safe_strip(src_val)
                                    dst_str = safe_strip(dst_val) if isinstance(dst_val, str) else str(dst_val) if dst_val is not None else ""
                                    
                                    if src_str != dst_str:
                                        src_col_name = get_column_name(src_col, year)
                                        dst_col_name = get_column_name(dst_col)
                                        issues.append({
                                            "severity": "ERROR",
                                            "sheet": all_grades_sheet,
                                            "row": all_data["row"],
                                            "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}('{src_str}')과 '2026 전학년' 시트의 {dst_col_name}('{dst_str}')이 일치하지 않습니다."
                                        })
                                # 숫자 비교
                                else:
                                    src_num = to_number(src_val)
                                    dst_num = dst_val if isinstance(dst_val, (int, float)) else to_number(dst_val)
                                    
                                    if src_num is not None and dst_num is not None:
                                        if abs(src_num - dst_num) > EPS:
                                            src_col_name = get_column_name(src_col, year)
                                            dst_col_name = get_column_name(dst_col)
                                            src_num_str = format_number(src_num)
                                            dst_num_str = format_number(dst_num)
                                            issues.append({
                                                "severity": "ERROR",
                                                "sheet": all_grades_sheet,
                                                "row": all_data["row"],
                                                "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}({src_num_str})과 '2026 전학년' 시트 {all_data['row']}행의 {dst_col_name}({dst_num_str})이 일치하지 않습니다."
                                            })
                                    elif src_num is not None or dst_num is not None:
                                        src_col_name = get_column_name(src_col, year)
                                        dst_col_name = get_column_name(dst_col)
                                        src_num_str = format_number(src_num)
                                        dst_num_str = format_number(dst_num)
                                        issues.append({
                                            "severity": "ERROR",
                                            "sheet": all_grades_sheet,
                                            "row": all_data["row"],
                                            "message": f"'{sname}' 시트 {course_row}행의 {src_col_name}({src_num_str})과 '2026 전학년' 시트 {all_data['row']}행의 {dst_col_name}({dst_num_str})이 일치하지 않습니다."
                                        })
        
        # 역방향 검증: '2026 전학년' 시트의 학생 선택 과목이 입학생 시트에 없는 경우
        for year in [2026, 2025, 2024]:
            # 모든 학년을 각 열별로 처리
            if year == 2026:
                col_keys = ['G', 'H']
            elif year == 2025:
                col_keys = ['I', 'J']
            else:  # 2024
                col_keys = ['K', 'L']
            
            for col_key in col_keys:
                    student_courses = student_courses_by_year.get(year, {}).get(col_key, {})
                    
                    if year not in sheets_data:
                        continue
                    
                    sheet_data = sheets_data[year]
                    ws_v, ws_f, merge = sheet_data["ws_v"], sheet_data["ws_f"], sheet_data["merge"]
                    sname = sheet_data["name"]
                    
                    # 마커 찾기
                    marker_row_src = find_marker_row(ws_v, ws_f, merge, "학교지정과목교과편성")
                    if not marker_row_src:
                        continue
                    
                    marker_row_student_end = find_marker_row(ws_v, ws_f, merge, "학생선택과목교과편성")
                    if not marker_row_student_end:
                        marker_row_student_end = ws_v.max_row + 1
                    
                    # 검사할 열 결정 (입학생 시트의 학기 열)
                    if year == 2026:
                        check_col = 7 if col_key == 'G' else 8  # G열(7) 또는 H열(8)
                        student_course_col = 4  # D열
                        a_col = 1  # A열
                    elif year == 2025:
                        check_col = 9 if col_key == 'I' else 10  # I열(9) 또는 J열(10)
                        student_course_col = 4  # D열
                        a_col = 1  # A열
                    else:  # 2024
                        # 입학생 시트 L(12) -> 전학년 시트 K(11), 입학생 시트 M(13) -> 전학년 시트 L(12)
                        check_col = 12 if col_key == 'K' else 13  # L열(12) 또는 M열(13)
                        student_course_col = 5  # E열
                        a_col = 2  # B열
                    processed_merges = set()
                    
                    # 입학생 시트에서 수집된 과목명 딕셔너리 (과목명별 출현 횟수 추적)
                    collected_courses_count = {}  # {과목명: 출현횟수}
                    
                    # 입학생 시트에서 과목 수집
                    for r in range(marker_row_src + 1, marker_row_student_end):
                        # A열 병합 확인
                        key = (r, a_col)
                        if key in merge:
                            min_row, _, max_row, _ = merge[key]
                            a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                            if a_col_value and '증배' in str(a_col_value):
                                # '증배'가 있으면 교차 점검 제외
                                continue
                            else:
                                merge_key = (min_row, max_row)
                        else:
                            # A열(또는 B열)에 '증배'가 포함되어 있는지 확인 (교차 점검 제외)
                            a_col_value, _, _ = get_value_with_merge(ws_v, ws_f, merge, r, a_col)
                            if a_col_value and '증배' in str(a_col_value):
                                continue
                            merge_key = (r, r)
                        
                        if merge_key in processed_merges:
                            continue
                        
                        processed_merges.add(merge_key)
                        
                        # 해당 병합 구간에서 check_col에 숫자가 있는지 확인
                        has_number = False
                        for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student_end)):
                            v, _, _ = get_value_with_merge(ws_v, ws_f, merge, rr, check_col)
                            if to_number(v) is not None:
                                has_number = True
                                break
                        
                        # 병합 구간 내의 모든 D열 과목명 수집 (순번 추적)
                        # 병합 구간 전체에서 숫자가 하나라도 있으면, 병합 구간의 모든 행에서 과목 수집
                        if has_number:
                            for rr in range(merge_key[0], min(merge_key[1] + 1, marker_row_student_end)):
                                course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge, rr, student_course_col)
                                if not course_raw or str(course_raw).strip() == "":
                                    continue
                                
                                # 총계 행 같은 키워드가 포함된 경우 제외
                                if any(keyword in str(course_raw) for keyword in ["편성학점", "총교과", "창의적체험", "편성학점수"]):
                                    continue
                                
                                # 각 행의 check_col에 숫자가 있는지 확인하지 않고, 병합 구간 전체의 has_number 결과 사용
                                normalized = normalize_course_name(course_raw)
                                if normalized:  # 빈 문자열이 아닌 경우만 추가
                                    # 순번 추적: 같은 과목명이 여러 번 나올 수 있으므로 출현 횟수 추적
                                    if normalized not in collected_courses_count:
                                        collected_courses_count[normalized] = 0
                                    collected_courses_count[normalized] += 1
                    
                    # 전학년 시트의 각 과목명이 입학생 시트에 있는지 확인 (순번 고려)
                    for course_norm, row_data_list in student_courses.items():
                        # 전학년 시트의 첫 번째 행에서 A열에 '증배'가 있는지 확인 (교차 점검 제외)
                        if row_data_list:
                            first_row_data = row_data_list[0]
                            a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, first_row_data["row"], 1)
                            if a_col_value and '증배' in str(a_col_value):
                                continue
                        
                        # 전학년 시트에서 해당 과목이 몇 번 나오는지 확인
                        expected_count = len(row_data_list)
                        
                        # 입학생 시트에서 해당 과목이 몇 번 나왔는지 확인
                        found_count = collected_courses_count.get(course_norm, 0)
                        
                        # 전학년 시트의 각 순번별로 입학생 시트에 있는지 확인
                        for occurrence_index, row_data in enumerate(row_data_list):
                            # 해당 순번의 과목이 입학생 시트에 있는지 확인
                            if found_count <= occurrence_index:
                                # 각 열별로 적절한 열 이름 표시
                                if year == 2026:
                                    col_name = "G열(1학년 1학기)" if col_key == 'G' else "H열(1학년 2학기)"
                                elif year == 2025:
                                    col_name = "I열(2학년 1학기)" if col_key == 'I' else "J열(2학년 2학기)"
                                else:  # 2024
                                    col_name = "K열(3학년 1학기)" if col_key == 'K' else "L열(3학년 2학기)"
                                issues.append({
                                    "severity": "ERROR",
                                    "sheet": all_grades_sheet,
                                    "row": row_data["row"],
                                    "message": f"'2026 전학년' 시트의 '{course_norm}'({row_data['row']}행)이 '{sname}' 시트의 {col_name}에 없습니다."
                                })
    
    # =========================
    # 3. '2026 전학년' 시트 총계 행 합계 검증
    # =========================
    
    # 데이터 범위 확인
    first_row = 5
    course_col = 4  # D열
    
    # 마지막 데이터 행 찾기
    last_row = None
    for rr in range(ws_all_f.max_row, first_row - 1, -1):
        v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, course_col)
        if v is not None and str(v).strip() != "":
            last_row = rr
            break
    
    if last_row is None:
        issues.append({
            "severity": "WARNING",
            "sheet": all_grades_sheet,
            "row": "-",
            "message": "D열(과목)에서 데이터 행을 찾지 못했습니다."
        })
    else:
        # 총계/합계 행 식별 (검사 제외 대상)
        exempt_rows = set()
        for rr in range(first_row, last_row + 1):
            # D열 병합 확인
            d_key = (rr, course_col)
            if d_key in merge_all:
                min_row, _, max_row, _ = merge_all[d_key]
                # 병합된 경우, 최상단 행의 값만 확인 (병합된 영역의 모든 행이 같은 값)
                if rr != min_row:
                    # 병합된 영역의 하위 행은 건너뛰기 (최상단 행에서만 확인)
                    continue
            
            course_v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, course_col)
            if course_v:
                course_str = str(course_v).strip().replace(" ", "")
                # 총계 관련 키워드가 있으면 제외
                if any(keyword in course_str for keyword in [
                    "편성학점", "총교과", "창의적체험활동", "편성학점수"
                ]):
                    # 병합된 경우, 병합된 모든 행을 exempt_rows에 추가
                    if d_key in merge_all:
                        min_row, _, max_row, _ = merge_all[d_key]
                        for exempt_rr in range(min_row, max_row + 1):
                            exempt_rows.add(exempt_rr)
                    else:
                        exempt_rows.add(rr)
        
        # A열에서 총계 행들 찾기 (필수 셀 존재 여부 확인)
        total_rows = {}  # {"학교지정": row, "학생선택": row, "총교과": row, "창의적": row, "편성학점수": row}
        
        required_cells = {
            "학교지정": "'학교 지정 과목 교과 편성 학점'",
            "학생선택": "'학생 선택 과목 교과 편성 학점'",
            "총교과": "'총 교과 편성 학점'",
            "창의적": "'창의적 체험활동 학점'",
            "편성학점수": "'편성 학점 수'"
        }
        
        for rr in range(first_row, ws_all_f.max_row + 1):
            a_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 1)  # A열
            if a_val:
                a_str = str(a_val).strip().replace(" ", "")
                
                # 안내 문구나 긴 텍스트 제외 (실제 총계 셀은 짧고 명확함)
                if len(a_str) > 30:  # 너무 긴 텍스트는 제외
                    continue
                if any(word in a_str for word in ["확인", "제대로", "다시", "주의", "주세요", "입력", "양식"]):
                    continue
                
                if ("학교지정" in a_str or "학교선택" in a_str) and "편성학점" in a_str and "과목" in a_str and "교과" in a_str:
                    total_rows["학교지정"] = rr
                elif "학생선택" in a_str and "편성학점" in a_str and "과목" in a_str and "교과" in a_str:
                    total_rows["학생선택"] = rr
                elif "총교과편성학점" in a_str or ("총교과" in a_str and "편성학점" in a_str and "과목" not in a_str):
                    total_rows["총교과"] = rr
                elif "창의적체험활동" in a_str and "학점" in a_str and "과목" not in a_str:
                    total_rows["창의적"] = rr
                elif "편성학점수" in a_str and "과목" not in a_str and "교과" not in a_str:
                    total_rows["편성학점수"] = rr
        
        # 필수 셀 존재 여부 확인
        for key, cell_name in required_cells.items():
            if key not in total_rows:
                issues.append({
                    "severity": "ERROR",
                    "sheet": all_grades_sheet,
                    "row": "-",
                    "message": f"총계 부분의 {cell_name} 셀이 존재하지 않습니다. 교육청의 양식을 확인하여 수정하고 다시 검사를 진행해주세요."
                })
        
        # G~L 열 (2026 전학년 시트의 학기별 열)
        sem_cols = list(range(7, 13))  # G~L
        sem_cols_name = "G~L"
        
        # 각 행의 G~L 합 계산 (exempt_rows 제외)
        row_total = {}
        for rr in range(first_row, last_row + 1):
            if rr in exempt_rows:
                continue
            
            course_v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, course_col)
            if course_v is None or str(course_v).strip() == "":
                continue
            
            sem_sum = 0.0
            any_num = False
            for cc in sem_cols:
                v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, cc)
                n = to_number(v)
                if n is not None:
                    sem_sum += n
                    any_num = True
            
            if any_num:
                row_total[rr] = sem_sum
        
        # (1) 학교 지정 과목 편성 학점 검증
        if "학교지정" in total_rows:
            school_row = total_rows["학교지정"]
            
            # 학교 지정 과목: 위의 행들 합계 (first_row ~ school_row-1)
            for col_idx, col_letter in enumerate(sem_cols):
                expected_sum = 0.0
                processed_merges = set()  # 이미 처리한 병합 셀 추적
                
                for rr in range(first_row, school_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_all:
                        min_row, _, max_row, _ = merge_all[key]
                        merge_key = (min_row, max_row, col_letter)
                        
                        # 이미 처리한 병합 영역이면 건너뛰기
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                
                actual_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, school_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)  # 열 번호를 문자로 변환
                    issues.append({
                        "severity": "ERROR",
                        "sheet": all_grades_sheet,
                        "row": school_row,
                        "message": f"학교 지정 과목 편성 학점 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값={expected_sum:g}"
                    })
        
        # (2) 학생 선택 과목 편성 학점 검증
        if "학생선택" in total_rows and "학교지정" in total_rows:
            student_row = total_rows["학생선택"]
            school_row = total_rows["학교지정"]
            
            # 학생 선택 과목: school_row+1 ~ student_row-1 합계 (증배 제외)
            for col_idx, col_letter in enumerate(sem_cols):
                expected_sum = 0.0
                processed_merges = set()  # 이미 처리한 병합 셀 추적
                
                for rr in range(school_row + 1, student_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 증배 확인 (A열)
                    a_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 1)
                    if a_val and "증배" in str(a_val):
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_all:
                        min_row, _, max_row, _ = merge_all[key]
                        merge_key = (min_row, max_row, col_letter)
                        
                        # 이미 처리한 병합 영역이면 건너뛰기
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                
                actual_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, student_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": all_grades_sheet,
                        "row": student_row,
                        "message": f"학생 선택 과목 편성 학점 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값={expected_sum:g} (증배 제외)"
                    })
        
        # (3) 총 교과 편성 학점 검증
        if "총교과" in total_rows and "학교지정" in total_rows and "학생선택" in total_rows:
            total_course_row = total_rows["총교과"]
            school_row = total_rows["학교지정"]
            student_row = total_rows["학생선택"]
            
            # 각 열별로 학교 지정과 학생 선택의 기댓값을 저장
            school_expected_all = {}
            student_expected_all = {}
            
            # 학교 지정 과목 기댓값 계산
            for col_letter in sem_cols:
                expected_sum = 0.0
                processed_merges = set()
                
                for rr in range(first_row, school_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_all:
                        min_row, _, max_row, _ = merge_all[key]
                        merge_key = (min_row, max_row, col_letter)
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                school_expected_all[col_letter] = expected_sum
            
            # 학생 선택 과목 기댓값 계산 (증배 제외)
            for col_letter in sem_cols:
                expected_sum = 0.0
                processed_merges = set()
                
                for rr in range(school_row + 1, student_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 증배 확인 (A열)
                    a_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 1)
                    if a_val and "증배" in str(a_val):
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_all:
                        min_row, _, max_row, _ = merge_all[key]
                        merge_key = (min_row, max_row, col_letter)
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                student_expected_all[col_letter] = expected_sum
            
            # 총 교과 = 학교 지정 기댓값 + 학생 선택 기댓값
            for col_letter in sem_cols:
                expected_sum = school_expected_all.get(col_letter, 0.0) + student_expected_all.get(col_letter, 0.0)
                
                actual_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, total_course_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": all_grades_sheet,
                        "row": total_course_row,
                        "message": f"총 교과 편성 학점 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값(학교지정+학생선택)={expected_sum:g}"
                    })
        
        # (4) 편성 학점 수 검증
        if "편성학점수" in total_rows and "학교지정" in total_rows and "학생선택" in total_rows:
            final_row = total_rows["편성학점수"]
            school_row = total_rows["학교지정"]
            student_row = total_rows["학생선택"]
            
            # 학교 지정과 학생 선택의 기댓값이 이미 위에서 계산되었는지 확인
            # 만약 총교과 검증을 거치지 않았다면 여기서 계산
            if 'school_expected_all' not in locals() or 'student_expected_all' not in locals():
                school_expected_all = {}
                student_expected_all = {}
                
                # 학교 지정 과목 기댓값 계산
                for col_letter in sem_cols:
                    expected_sum = 0.0
                    processed_merges = set()
                    
                    for rr in range(first_row, school_row):
                        if rr in exempt_rows:
                            continue
                        
                        # 병합 셀 확인
                        key = (rr, col_letter)
                        if key in merge_all:
                            min_row, _, max_row, _ = merge_all[key]
                            merge_key = (min_row, max_row, col_letter)
                            if merge_key in processed_merges:
                                continue
                            processed_merges.add(merge_key)
                        
                        val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col_letter)
                        num = to_number(val)
                        if num is not None:
                            expected_sum += num
                    school_expected_all[col_letter] = expected_sum
                
                # 학생 선택 과목 기댓값 계산 (증배 제외)
                for col_letter in sem_cols:
                    expected_sum = 0.0
                    processed_merges = set()
                    
                    for rr in range(school_row + 1, student_row):
                        if rr in exempt_rows:
                            continue
                        
                        # 증배 확인 (A열)
                        a_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 1)
                        if a_val and "증배" in str(a_val):
                            continue
                        
                        # 병합 셀 확인
                        key = (rr, col_letter)
                        if key in merge_all:
                            min_row, _, max_row, _ = merge_all[key]
                            merge_key = (min_row, max_row, col_letter)
                            if merge_key in processed_merges:
                                continue
                            processed_merges.add(merge_key)
                        
                        val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, col_letter)
                        num = to_number(val)
                        if num is not None:
                            expected_sum += num
                    student_expected_all[col_letter] = expected_sum
            
            # 편성 학점 수 = 총 교과 기댓값 + 창의적(3)
            for col_letter in sem_cols:
                total_expected = school_expected_all.get(col_letter, 0.0) + student_expected_all.get(col_letter, 0.0)
                expected_sum = total_expected + 3.0
                
                actual_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, final_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": all_grades_sheet,
                        "row": final_row,
                        "message": f"편성 학점 수 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값(총교과+창의적)={expected_sum:g}"
                    })

    # =========================
    # 4. '2026 전학년' 시트 증배 과목 검증
    # =========================
    
    # hidden 데이터가 없으면 증배 검증을 수행하지 않음
    if hidden is None or vocational_courses is None or new_courses is None or hidden_list_norm is None:
        return
    
    # 증배 관련 행 수집
    jeungbae_rows = []  # 검증 대상 행 리스트
    
    # 마지막 행 찾기 (이미 위에서 찾았지만, 다시 확인)
    if last_row is None:
        for rr in range(ws_all_f.max_row, first_row - 1, -1):
            v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, course_col)
            if v is not None and str(v).strip() != "":
                last_row = rr
                break
    
    if last_row is None:
        return  # 데이터가 없으면 검증하지 않음
    
    # 증배 관련 행 찾기
    processed_a_merges = set()  # 처리된 A열 병합 영역 추적
    
    for rr in range(first_row, last_row + 1):
        if rr in exempt_rows:
            continue  # 총계 행은 제외
        
        # A열 병합 확인
        a_key = (rr, 1)  # A열
        rows_to_check = []
        
        if a_key in merge_all:
            # 병합된 경우: 병합된 영역의 모든 행 확인
            min_row, _, max_row, _ = merge_all[a_key]
            merge_key = (min_row, max_row)
            
            # 이미 처리한 병합 영역은 건너뛰기
            if merge_key in processed_a_merges:
                continue
            processed_a_merges.add(merge_key)
            
            # 병합된 영역의 최상단 행에서 '증배' 확인
            a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, min_row, 1)
            
            if a_col_value and '증배' in str(a_col_value):
                # 병합된 모든 행을 검증 대상에 추가
                rows_to_check = list(range(min_row, max_row + 1))
        else:
            # 병합되지 않은 경우: 해당 행만 확인
            a_col_value, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, 1)
            if a_col_value and '증배' in str(a_col_value):
                rows_to_check = [rr]
        
        # 각 행마다 K, L열 확인 및 검증 대상 추가
        for check_row in rows_to_check:
            if check_row in exempt_rows:
                continue
            
            # K열(11) 또는 L열(12)에 숫자가 있는지 확인
            k_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, check_row, 11)  # K열
            l_val, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, check_row, 12)  # L열
            
            if to_number(k_val) is not None or to_number(l_val) is not None:
                continue  # K 또는 L열에 숫자가 있으면 검증 제외
            
            # 검증 대상으로 추가
            jeungbae_rows.append(check_row)
    
    # 각 증배 행에 대해 검증 수행
    for rr in jeungbae_rows:
        # 과목명 읽기
        course_raw, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, course_col)
        if course_raw is None or str(course_raw).strip() == "":
            continue
        
        if is_error_token(course_raw):
            continue
        
        course_norm = normalize_course_name(course_raw)
        if course_norm == "":
            issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": "과목명(D열)에서 괄호 제거 후 이름이 비었습니다."})
            if rr >= 109 and rr <= 116:
                print(f"DEBUG [증배검증]: 행 {rr} 과목명 정규화 후 비어있어 오류 추가")
            continue
        
        # 과목명 검증 (입학생 시트와 동일한 로직)
        parts = split_bidirectional(course_norm)
        is_bidirectional = len(parts) >= 2
        
        hidden_rec = None
        
        if is_bidirectional:
            missing = [p for p in parts if p not in hidden]
            if missing:
                # 전문교과목록에서 확인
                missing_not_in_vocational = [m for m in missing if m not in vocational_courses]
                missing_in_vocational = [m for m in missing if m in vocational_courses]
                
                # 신설교과에서 확인
                missing_not_in_new = [m for m in missing_not_in_vocational if m not in new_courses]
                missing_in_new = [m for m in missing_not_in_vocational if m in new_courses]
                
                if missing_in_vocational:
                    issues.append({"severity": "CHECK", "sheet": all_grades_sheet, "row": rr, "message": f"일반고에서 전문교과의 경우는 진로로 편성할 수 있습니다. (과목명: {', '.join(missing_in_vocational)})"})
                
                if missing_in_new:
                    for new_course in missing_in_new:
                        msg_line1 = f"'{new_course}'은(는) 교육과정에 표시되지 않은 교과목 중 신설 승인이 된 과목입니다."
                        msg_line2 = "      각 학교에서 해당 교과목을 편성하기 위해서는 교육청에 사용 승인을 받아야 합니다."
                        issues.append({
                            "severity": "CHECK", 
                            "sheet": all_grades_sheet, 
                            "row": rr, 
                            "message": msg_line1 + "\n" + msg_line2
                        })
                
                if missing_not_in_new:
                    hints = []
                    for m in missing_not_in_new:
                        close = difflib.get_close_matches(m, hidden_list_norm, n=1, cutoff=0.6)
                        if close:
                            hints.append(f"{m}→{close[0]}")
                    hint_txt = f" (유사 후보: {', '.join(hints)})" if hints else ""
                    issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"↔ 과목명 중 지침에 없는 항목: {', '.join(missing_not_in_new)}{hint_txt}"})
            
            
        else:
            if course_norm not in hidden:
                # 전문교과목록 시트에서 확인
                if course_norm in vocational_courses:
                    issues.append({"severity": "CHECK", "sheet": all_grades_sheet, "row": rr, "message": f"일반고에서 전문교과의 경우는 진로로 편성할 수 있습니다. (과목명: '{course_norm}')"})
                    hidden_rec = None
                # 신설교과 시트에서 확인
                elif course_norm in new_courses:
                    issues.append({"severity": "CHECK", "sheet": all_grades_sheet, "row": rr, "message": f"'{course_norm}'은(는) 교육과정에 표시되지 않은 교과목 중 신설 승인이 된 과목입니다. 각 학교에서 해당 교과목을 편성하기 위해서는 교육청에 사용 승인을 받아야 합니다."})
                    hidden_rec = None
                else:
                    hint = ""
                    close = difflib.get_close_matches(course_norm, hidden_list_norm, n=2, cutoff=0.6)
                    if close:
                        hint = f" (유사 과목명 후보: {', '.join(close)})"
                    issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"과목명 오류: '{course_norm}'{hint}"})
                    hidden_rec = None
            else:
                hidden_rec = hidden[course_norm]
        
        # 숨김 시트 매칭이 있을 때만 유형/기본학점/성적처리 검증
        if hidden_rec is not None:
            # 유형 검증 (C열)
            type_col = 3
            typ_v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, type_col)
            typ_s = safe_strip(typ_v)
            if typ_s == "":
                issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"유형(C{rr})이 비어 있습니다. (지침: {hidden_rec['type']})"})
            elif typ_s != hidden_rec["type"]:
                issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"유형 불일치: 시트='{typ_s}' / 지침='{hidden_rec['type']}'"})
            
            # 기본학점 검증 (E열)
            basic_col = 5
            basic_v, basic_formula, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, basic_col)
            basic_n = to_number(basic_v)
            if basic_n is None:
                if basic_formula:
                    issues.append({"severity": "WARNING", "sheet": all_grades_sheet, "row": rr, "message": f"기본학점(E{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {basic_formula})"})
                else:
                    issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"기본학점(E{rr})이 숫자가 아닙니다: {basic_v}"})
            else:
                if hidden_rec["basic"] is not None and abs(basic_n - hidden_rec["basic"]) > EPS:
                    issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"기본학점 불일치: 시트={basic_n:g} / 지침={hidden_rec['basic']:g}"})
            
            # 성적처리 확인 문구 (O열)
            grading_col = 15
            grade_v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, grading_col)
            grade_s = safe_strip(grade_v)
            if hidden_rec["grading"]:
                issues.append({
                    "severity": "CHECK",
                    "sheet": all_grades_sheet,
                    "row": rr,
                    "message": f"소인수, 공동, 증배의 경우엔 성적처리를 다르게 설정할 수 있습니다.(지침: {hidden_rec['grading']})"
                })
        
        # 운영학점 검증 (F열)
        op_col = 6  # F열
        op_v, op_formula, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, op_col)
        op_n = to_number(op_v)
        
        # 학기별 편성 학점 합계 계산 (G~L열)
        sem_cols = list(range(7, 13))  # G~L
        sem_sum = 0.0
        any_num = False
        for cc in sem_cols:
            v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, cc)
            n = to_number(v)
            if n is not None:
                sem_sum += n
                any_num = True
        
        if op_n is None:
            if op_formula:
                issues.append({"severity": "WARNING", "sheet": all_grades_sheet, "row": rr, "message": f"운영학점(F{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {op_formula})"})
            else:
                issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"운영학점(F{rr})이 숫자가 아닙니다: {op_v}"})
        else:
            # 운영학점 범위 검증
            if hidden_rec is not None and (hidden_rec["min"] is not None) and (hidden_rec["max"] is not None):
                if not (hidden_rec["min"] - EPS <= op_n <= hidden_rec["max"] + EPS):
                    issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"운영학점 범위 위반: 시트={op_n:g} / 허용범위={hidden_rec['min']:g}~{hidden_rec['max']:g}"})
            
            # 운영학점과 학기별 합계 비교
            if abs(sem_sum) <= EPS:
                # 학기별 합계가 0인 경우, 위쪽 행에서 숫자 찾기
                nearest_sum = 0.0
                found_any = False
                nearest_row = None
                
                for search_row in range(rr - 1, first_row - 1, -1):
                    row_has_number = False
                    for cc in sem_cols:
                        v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, search_row, cc)
                        n = to_number(v)
                        if n is not None and abs(n) > EPS:
                            row_has_number = True
                            break
                    
                    if row_has_number:
                        nearest_row = search_row
                        break
                
                if nearest_row is not None:
                    for cc in sem_cols:
                        v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, nearest_row, cc)
                        n = to_number(v)
                        if n is not None:
                            nearest_sum += n
                            found_any = True
                
                if found_any:
                    if abs(op_n - nearest_sum) > EPS:
                        if abs(nearest_sum - op_n * 2) <= EPS:
                            issues.append({"severity": "CHECK", "sheet": all_grades_sheet, "row": rr, "message": f"학기 편성 학점의 합({nearest_sum:g})과 운영학점({op_n:g})이 다릅니다.(학기제라면 오류가 아닐 수 있습니다.)"})
                        else:
                            issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"학기 편성 학점의 합({nearest_sum:g})과 운영학점({op_n:g})이 다릅니다."})
                else:
                    issues.append({"severity": "CHECK", "sheet": all_grades_sheet, "row": rr, "message": f"학기 편성 학점의 합이 0입니다. 선택군 과목 편성 학점을 확인해주세요."})
            else:
                if abs(op_n - sem_sum) > EPS:
                    if abs(sem_sum - op_n * 2) <= EPS:
                        issues.append({"severity": "CHECK", "sheet": all_grades_sheet, "row": rr, "message": f"학기 편성 학점의 합({sem_sum:g})과 운영학점({op_n:g})이 다릅니다.(학기제라면 오류가 아닐 수 있습니다)"})
                    else:
                        issues.append({"severity": "ERROR", "sheet": all_grades_sheet, "row": rr, "message": f"학기 편성 학점의 합({sem_sum:g})과 운영학점({op_n:g})이 다릅니다."})
        
        # 합계 열 검증 (M열 또는 N열)
        total_cols = [13, 14]  # M, N열
        for total_col in total_cols:
            total_v, _, _ = get_value_with_merge(ws_all_v, ws_all_f, merge_all, rr, total_col)
            total_n = to_number(total_v)
            
            if total_n is not None and op_n is not None:
                if abs(total_n - op_n) > EPS:
                    col_name = chr(64 + total_col)  # M 또는 N
                    issues.append({
                        "severity": "ERROR",
                        "sheet": all_grades_sheet,
                        "row": rr,
                        "message": f"{col_name}열 합계 불일치: 셀값={total_n:g}, 운영학점(F열)={op_n:g}"
                    })


def check_school_name_consistency(wb_v, wb_f, targets, issues):
    """
    모든 시트의 2행에서 학교명이 올바르게 입력되었는지 확인
    - 'OO고등학교'로 되어 있으면 오류
    - 괄호 안에 공립/국립/사립 중 하나만 있어야 함
    - 모든 시트가 동일한 학교명을 가져야 함
    """
    sheetnames = wb_v.sheetnames
    all_grades_sheet = find_all_grades_sheet(sheetnames)
    
    # 검사할 시트 목록 (입학생 시트 + 전학년 시트)
    sheets_to_check = []
    for year, sname in targets.items():
        if sname:
            sheets_to_check.append(sname)
    if all_grades_sheet:
        sheets_to_check.append(all_grades_sheet)
    
    if not sheets_to_check:
        return
    
    school_names = {}  # {시트명: 학교명}
    
    for sname in sheets_to_check:
        ws_v = wb_v[sname]
        ws_f = wb_f[sname]
        merge_lookup = build_merged_lookup(ws_f)
        
        # 2행의 학교명 찾기 (보통 병합된 셀에 있음)
        school_name = None
        for col in range(1, ws_f.max_column + 1):
            val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, 2, col)
            if val and isinstance(val, str):
                val_str = str(val).strip()
                # 학교명으로 추정되는 패턴: "고등학교" 포함
                if "고등학교" in val_str or "학년도" in val_str:
                    school_name = val_str
                    break
        
        if not school_name:
            issues.append({
                "severity": "ERROR",
                "sheet": sname,
                "row": 2,
                "message": "2행에서 학교명을 찾을 수 없습니다. 학교명이 올바르게 입력되었는지 확인해주세요."
            })
            continue
        
        school_names[sname] = school_name
        
        # 1. 'OO고등학교' 패턴 체크
        if "OO고등학교" in school_name or "○○고등학교" in school_name or "○○ 고등학교" in school_name or "OO 고등학교" in school_name:
            issues.append({
                "severity": "ERROR",
                "sheet": sname,
                "row": 2,
                "message": f"학교명이 'OO고등학교'로 되어 있습니다. 실제 학교명으로 수정해주세요. (현재: {school_name})"
            })
            continue
        
        # 2. 괄호 안 학교 유형 체크
        import re
        # 괄호 안의 내용 추출
        bracket_pattern = r'\((.*?)\)'
        brackets = re.findall(bracket_pattern, school_name)
        
        valid_types = ["공립", "국립", "사립"]
        
        if brackets:
            # 괄호 내용 확인
            bracket_content = brackets[0].strip()
            
            # 여러 단어가 있거나, 유효한 유형이 아닌 경우
            if bracket_content not in valid_types:
                # "공립/국립/사립" 같은 형태인지 확인
                if "/" in bracket_content or "," in bracket_content or " " in bracket_content:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": 2,
                        "message": f"괄호 안에 공립, 국립, 사립 중 하나만 적어야 합니다. (현재: {school_name})"
                    })
                elif bracket_content not in valid_types:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": 2,
                        "message": f"괄호 안에 '공립', '국립', '사립' 중 하나를 적어야 합니다. (현재: {school_name})"
                    })
        else:
            # 괄호가 없는 경우
            issues.append({
                "severity": "ERROR",
                "sheet": sname,
                "row": 2,
                "message": f"학교명 뒤에 괄호와 함께 공립/국립/사립을 표기해야 합니다. (현재: {school_name})"
            })
    
    # 3. 모든 시트의 학교명이 동일한지 확인
    if len(school_names) > 1:
        unique_names = set(school_names.values())
        if len(unique_names) > 1:
            # 학교명이 다른 시트들 찾기
            name_list = "\n".join([f"  • {sname}: {name}" for sname, name in school_names.items()])
            issues.append({
                "severity": "ERROR",
                "sheet": "-",
                "row": 2,
                "message": f"시트마다 학교명이 다릅니다. 모든 시트에 동일한 학교명을 입력해주세요.\n\n[각 시트의 학교명]\n{name_list}"
            })


def run_checks(xlsx_path: str):
    """
    return: (issues, summary)
      - issues: list[dict]  {severity, sheet, row, message}
      - summary: dict
    """
    if not os.path.exists(xlsx_path):
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": "파일을 찾을 수 없습니다."}], {})

    ext = os.path.splitext(xlsx_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": "지원하지 않는 확장자입니다. .xlsx 또는 .xlsm만 지원합니다."}], {})

    issues = []
    summary = {}

    try:
        wb_v = load_workbook(xlsx_path, data_only=True)
        wb_f = load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": f"엑셀 파일을 열 수 없습니다: {e}"}], {})

    sheetnames = wb_v.sheetnames

    # (1) 대상 시트 존재 확인
    targets = {}
    for y in (2026, 2025, 2024):
        sname = find_sheet_for_year(sheetnames, y)
        targets[y] = sname
        if not sname:
            issues.append({"severity": "ERROR", "sheet": "-", "row": "-", "message": f"{y} 입학생으로 시작하고 '입학생'을 포함하는 시트를 찾지 못했습니다."})

    # ========================================
    # 숨김 시트 및 전문교과목록 시트 로드
    # 우선순위: 1) 구글 스프레드시트 → 2) 엑셀 파일 내부
    # ========================================
    
    # 1단계: 구글 스프레드시트에서 가져오기 시도
    ref_wb_v, ref_wb_f, google_success, google_error = load_reference_sheets_from_google()
    
    if google_success:
        # 구글에서 성공적으로 가져옴
        ref_sheetnames = ref_wb_v.sheetnames
        hidden_name = "숨김" if "숨김" in ref_sheetnames else None
        vocational_sheet_name = "전문교과목록" if "전문교과목록" in ref_sheetnames else None
        data_source = "온라인(구글 스프레드시트)"
    else:
        # 2단계: 엑셀 파일 내부에서 찾기
        ref_wb_v = wb_v
        ref_wb_f = wb_f
        ref_sheetnames = sheetnames
        hidden_name = find_hidden_sheet_name(ref_sheetnames)
        
        # 전문교과목록 시트 찾기
        vocational_sheet_name = None
        for sname in ref_sheetnames:
            if "전문교과목록" in sname or "전문교과" in sname:
                vocational_sheet_name = sname
                break
        
        data_source = "엑셀 파일 내부"
    
    # 숨김 시트가 없으면 오류
    if not hidden_name:
        issues.append({
            "severity": "ERROR", 
            "sheet": "-", 
            "row": "-", 
            "message": "엑셀 프로그램 내에서 숨김, 전문교과목록 시트를 찾을 수 없습니다. 교육청에서 제공된 양식을 활용해 주세요."
        })
        # 구글 스프레드시트 참조 실패 시 오류 메시지 저장
        google_error_msg = google_error if not google_success and google_error else None
        return issues, {"targets": targets, "hidden_sheet": None, "data_source": None, "google_error": google_error_msg}

    ws_hidden_v = ref_wb_v[hidden_name]
    ws_hidden_f = ref_wb_f[hidden_name]
    hidden_merge = build_merged_lookup(ws_hidden_f)
    header_row = find_hidden_header_row(ws_hidden_v, ws_hidden_f, hidden_merge)
    data_start = header_row + 1

    # 숨김 과목 사전 구축
    hidden = {}
    hidden_list_norm = []
    r = data_start
    while True:
        course_raw, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 2)  # B
        if course_raw is None or str(course_raw).strip() == "":
            break
        course_norm = normalize_course_name(course_raw)

        subject_group, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 1)  # A: 교과(군)
        typ, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 3)  # C
        basic, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 4)  # D
        grade, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 5)  # E
        minc, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 6)  # F
        maxc, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 7)  # G
        special_note, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 9)  # I

        rec = {
            "course_raw": safe_strip(course_raw),
            "subject_group": safe_strip(subject_group),
            "type": safe_strip(typ),
            "basic": to_number(basic),
            "grading": safe_strip(grade),
            "min": to_number(minc),
            "max": to_number(maxc),
            "special_note": safe_strip(special_note),
            "row": r,
        }

        if course_norm in hidden:
            issues.append({
                "severity": "WARNING",
                "sheet": hidden_name,
                "row": r,
                "message": f"지침에 중복 과목명이 있습니다: '{course_norm}' (기존 {hidden[course_norm]['row']}행, 추가 {r}행). 최초 항목을 기준으로 검사합니다."
            })
        else:
            hidden[course_norm] = rec
            hidden_list_norm.append(course_norm)

        r += 1

    summary["targets"] = targets
    summary["hidden_sheet"] = hidden_name
    summary["hidden_course_count"] = len(hidden)
    summary["data_source"] = data_source
    # 구글 스프레드시트 참조 실패 시 오류 메시지 저장
    if not google_success and google_error:
        summary["google_error"] = google_error

    # 전문교과목록 시트 로드 (있으면)
    vocational_courses = set()
    
    if vocational_sheet_name:
        try:
            ws_voc_v = ref_wb_v[vocational_sheet_name]
            ws_voc_f = ref_wb_f[vocational_sheet_name]
            voc_merge = build_merged_lookup(ws_voc_f)
            
            # C열에서 과목명 읽기 (헤더 행은 1~3 사이로 가정, 데이터는 그 이후부터)
            for rr in range(2, ws_voc_f.max_row + 1):
                course_v, _, _ = get_value_with_merge(ws_voc_v, ws_voc_f, voc_merge, rr, 3)  # C열
                if course_v and str(course_v).strip():
                    course_normalized = normalize_course_name(course_v)
                    if course_normalized:
                        vocational_courses.add(course_normalized)
        except Exception:
            pass  # 전문교과목록 시트 로드 실패 시 무시
    
    summary["vocational_sheet"] = vocational_sheet_name
    summary["vocational_course_count"] = len(vocational_courses)
    
    # 신설교과 시트 로드 (있으면)
    new_courses = set()
    new_course_sheet_name = None
    
    for sname in ref_sheetnames:
        if "신설교과" in sname:
            new_course_sheet_name = sname
            break
    
    if new_course_sheet_name:
        try:
            ws_new_v = ref_wb_v[new_course_sheet_name]
            ws_new_f = ref_wb_f[new_course_sheet_name]
            new_merge = build_merged_lookup(ws_new_f)
            
            # B열에서 과목명 읽기 (헤더 행은 1~3 사이로 가정, 데이터는 그 이후부터)
            for rr in range(2, ws_new_f.max_row + 1):
                course_v, _, _ = get_value_with_merge(ws_new_v, ws_new_f, new_merge, rr, 2)  # B열
                if course_v and str(course_v).strip():
                    course_normalized = normalize_course_name(course_v)
                    if course_normalized:
                        new_courses.add(course_normalized)
        except Exception:
            pass  # 신설교과 시트 로드 실패 시 무시
    
    summary["new_course_sheet"] = new_course_sheet_name
    summary["new_course_count"] = len(new_courses)

    # 시트가 없다면 여기서 종료
    if any(v is None for v in targets.values()):
        return issues, summary

    # =========================
    # (2) 각 시트 검사
    # =========================
    for year, sname in targets.items():
        ws_v = wb_v[sname]
        ws_f = wb_f[sname]
        merge_lookup = build_merged_lookup(ws_f)

        first_row = 5
        subject_group_col = 2  # B: 교과(군)
        course_col = 4  # D
        type_col = 3    # C
        basic_col = 5   # E
        
        # 2024 시트는 열 구성이 다름
        if year == 2024:
            op_col = 7      # G (운영학점)
            sem_cols = list(range(8, 14))  # H~M
            total_cols = [14, 15]          # N, O
            grading_col = 16               # P
            op_col_name = "G"
            sem_cols_name = "H~M"
            total_cols_name = "N/O"
            compare_col = 2  # B열 (2024는 B열과 N,O열 병합 비교)
        else:  # 2025, 2026
            op_col = 6      # F (운영학점)
            sem_cols = list(range(7, 13))  # G~L
            total_cols = [13, 14]          # M, N
            grading_col = 15               # O
            op_col_name = "F"
            sem_cols_name = "G~L"
            total_cols_name = "M/N"
            compare_col = 1  # A열 (2025/2026은 A열과 M,N열 병합 비교)

        # last row 찾기: '편성 학점 수' 또는 '편성학점수'가 포함된 행
        last_row = None
        for rr in range(first_row, ws_f.max_row + 1):
            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if v is not None:
                v_str = str(v).strip().replace(" ", "")
                if "편성학점수" in v_str or "편성학점합계" in v_str:
                    last_row = rr
                    break
        
        if last_row is None:
            # '편성 학점 수'를 찾지 못한 경우
            issues.append({
                "severity": "ERROR", 
                "sheet": sname, 
                "row": "-", 
                "message": (
                    "시트의 마지막 행에서 '편성 학점 수'를 찾지 못했습니다.\n"
                    "표의 총계 부분이 양식과 같이 입력되어 있는지 확인하고 다시 실행해 주세요.\n\n"
                    "[필요한 총계 행 구성(작년 양식에서 변경되었습니다.)]\n"
                    "• 학생 지정 과목 교과 편성 학점\n"
                    "• 학생 선택 과목 교과 편성 학점\n"
                    "• 총 교과 편성 학점\n"
                    "• 창의적 체험활동 학점\n"
                    "• 편성 학점 수"
                )
            })
            continue
        
        # '편성 학점 수' 행은 검사 대상이 아니므로, 실제 검사는 그 위까지만
        check_until_row = last_row - 1

        # 증배 검사 (입학생 시트에서 제거되어야 함)
        a_col = 1 if year in [2026, 2025] else 2  # A열 또는 B열
        found_jeungbae = False
        for rr in range(first_row, check_until_row + 1):
            a_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, a_col)
            if a_val and "증배" in str(a_val):
                found_jeungbae = True
                break

        if found_jeungbae:
            issues.append({
                "severity": "WARNING",
                "sheet": sname,
                "row": "-",
                "message": (
                    "소인수, 공동, 증배 운영 과목은 전학년 시트에만 입력해주세요.\n"
                    "일과내 운영 과목이라면 교육과정 편성표 내에 '(소인수)산업수학'과 같이 입력하여 작성하면 됩니다."
                )
            })

        # 총계/합계 행은 모든 검사 제외 (D열 내용 기준)
        exempt_rows = set()
        for rr in range(first_row, check_until_row + 1):
            course_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_v:
                course_str = str(course_v).strip().replace(" ", "")
                # 총계 관련 키워드가 있으면 제외
                if any(keyword in course_str for keyword in [
                    "편성학점", "총교과", "창의적체험활동", "편성학점수"
                ]):
                    exempt_rows.add(rr)

        # row_total(각 행의 G~L 합) 계산 (총계 행만 제외, 색깔행은 포함)
        row_total = {}
        for rr in range(first_row, check_until_row + 1):
            if rr in exempt_rows:
                continue  # 총계 행만 제외

            course_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_v is None or str(course_v).strip() == "":
                continue

            if is_error_token(course_v):
                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"과목명(D{rr})에 오류값이 있습니다: {course_v}"})
                continue

            sem_sum = 0.0
            any_num = False
            for cc in sem_cols:
                v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, cc)
                n = to_number(v)
                if n is not None:
                    sem_sum += n
                    any_num = True

            row_total[rr] = sem_sum if any_num else 0.0

        # ========== 과목 단위 검사 ==========
        for rr in range(first_row, check_until_row + 1):
            # 모든 연도에서 색깔행도 기본학점/운영학점 검증에 포함
            # 총계 행만 제외 (exempt_rows에는 총계 관련 키워드가 있는 행만 포함)
            if rr in exempt_rows:
                continue  # 총계 행은 제외

            course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_raw is None or str(course_raw).strip() == "":
                continue
            if is_error_token(course_raw):
                continue

            course_norm = normalize_course_name(course_raw)
            if course_norm == "":
                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": "과목명(D열)에서 괄호 제거 후 이름이 비었습니다."})
                continue

            parts = split_bidirectional(course_norm)
            is_bidirectional = len(parts) >= 2

            # (2) 과목명 일치 여부
            hidden_rec = None
            if year == 2024:
                # 2024는 과목명 일치 검증을 하지 않음(있으면 활용, 없으면 숨김기반 검사는 생략)
                hidden_rec = hidden.get(course_norm, None)
            else:
                if is_bidirectional:
                    missing = [p for p in parts if p not in hidden]
                    if missing:
                        # 전문교과목록에서 확인
                        missing_not_in_vocational = [m for m in missing if m not in vocational_courses]
                        missing_in_vocational = [m for m in missing if m in vocational_courses]
                        
                        # 신설교과에서 확인
                        missing_not_in_new = [m for m in missing_not_in_vocational if m not in new_courses]
                        missing_in_new = [m for m in missing_not_in_vocational if m in new_courses]
                        
                        if missing_in_vocational:
                            issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"일반고에서 전문교과의 경우는 진로로 편성할 수 있습니다. (과목명: {', '.join(missing_in_vocational)})"})
                        
                        if missing_in_new:
                            for new_course in missing_in_new:
                                msg_line1 = f"'{new_course}'은(는) 교육과정에 표시되지 않은 교과목 중 신설 승인이 된 과목입니다."
                                msg_line2 = "      각 학교에서 해당 교과목을 편성하기 위해서는 교육청에 사용 승인을 받아야 합니다."
                                issues.append({
                                    "severity": "CHECK", 
                                    "sheet": sname, 
                                    "row": rr, 
                                    "message": msg_line1 + "\n" + msg_line2
                                })
                        
                        if missing_not_in_new:
                            hints = []
                            for m in missing_not_in_new:
                                close = difflib.get_close_matches(m, hidden_list_norm, n=1, cutoff=0.6)
                                if close:
                                    hints.append(f"{m}→{close[0]}")
                            hint_txt = f" (유사 후보: {', '.join(hints)})" if hints else ""
                            issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"↔ 과목명 중 지침에 없는 항목: {', '.join(missing_not_in_new)}{hint_txt}"})
                    
                    # ↔ 과목도 병합 셀 위치에 따라 좌/우 과목으로 검증
                    # 병합 셀의 top-left 행이면 왼쪽 과목, 아니면 오른쪽 과목
                    key = (rr, course_col)
                    if key in merge_lookup:
                        min_row, min_col, max_row, max_col = merge_lookup[key]
                        # 병합이 2행 이상인지 확인
                        if max_row - min_row < 1:
                            # 병합이 1행만인 경우 (실제로는 병합이 아님)
                            issues.append({
                                "severity": "ERROR",
                                "sheet": sname,
                                "row": rr,
                                "message": "교차이수과목의 경우 두 개의 행에 각 과목에 대해 작성하고 과목명만 병합해주세요"
                            })
                            hidden_rec = None
                        else:
                            if rr == min_row:
                                # 병합 영역의 첫 행 -> 왼쪽 과목
                                target_course = parts[0] if len(parts) > 0 else None
                            else:
                                # 병합 영역의 두 번째 행 이후 -> 오른쪽 과목
                                target_course = parts[1] if len(parts) > 1 else None
                            
                            if target_course and target_course in hidden:
                                hidden_rec = hidden[target_course]
                                
                                # 교과(군) 비교 (2025/2026만, '증배' 제외)
                                if hidden_rec and year in (2025, 2026):
                                    sheet_subject_group, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, subject_group_col)
                                    sheet_subject_group_str = safe_strip(sheet_subject_group)
                                    hidden_subject_group_str = hidden_rec.get("subject_group", "")
                                    
                                    # '증배'가 포함된 경우 교과(군) 체크 제외
                                    if "증배" not in sheet_subject_group_str:
                                        # 교과(군)이 비어있는 경우
                                        if not sheet_subject_group_str:
                                            issues.append({
                                                "severity": "ERROR", 
                                                "sheet": sname, 
                                                "row": rr, 
                                                "message": f"교과(군)(B열)이 비어 있습니다. (지침: '{hidden_subject_group_str}')"
                                            })
                                        elif not hidden_subject_group_str:
                                            issues.append({
                                                "severity": "WARNING", 
                                                "sheet": sname, 
                                                "row": rr, 
                                                "message": f"지침에 교과(군) 정보가 없습니다. (시트: '{sheet_subject_group_str}')"
                                            })
                                        elif sheet_subject_group_str != hidden_subject_group_str:
                                            issues.append({
                                                "severity": "ERROR", 
                                                "sheet": sname, 
                                                "row": rr, 
                                                "message": f"교과(군) 불일치: 시트(B열)='{sheet_subject_group_str}' / 지침(A열)='{hidden_subject_group_str}'"
                                            })
                            else:
                                hidden_rec = None
                    else:
                        # 병합되지 않은 경우 ERROR 발생
                        issues.append({
                            "severity": "ERROR",
                            "sheet": sname,
                            "row": rr,
                            "message": "교차이수과목의 경우 두 개의 행에 각 과목에 대해 작성하고 과목명만 병합해주세요"
                        })
                        hidden_rec = None
                else:
                    if course_norm not in hidden:
                        # 전문교과목록 시트에서 확인
                        if course_norm in vocational_courses:
                            issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"일반고에서 전문교과의 경우는 진로로 편성할 수 있습니다. (과목명: '{course_norm}')"})
                            hidden_rec = None
                        # 신설교과 시트에서 확인
                        elif course_norm in new_courses:
                            issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"'{course_norm}'은(는) 교육과정에 표시되지 않은 교과목 중 신설 승인이 된 과목입니다. 각 학교에서 해당 교과목을 편성하기 위해서는 교육청에 사용 승인을 받아야 합니다."})
                            hidden_rec = None
                        else:
                            hint = ""
                            close = difflib.get_close_matches(course_norm, hidden_list_norm, n=2, cutoff=0.6)
                            if close:
                                hint = f" (유사 과목명 후보: {', '.join(close)})"
                            issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"과목명 오류: '{course_norm}'{hint}"})
                            hidden_rec = None
                    else:
                        hidden_rec = hidden[course_norm]
                        
                        # 교과(군) 비교 (2025/2026만, '증배' 제외)
                        if hidden_rec and year in (2025, 2026):
                            sheet_subject_group, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, subject_group_col)
                            sheet_subject_group_str = safe_strip(sheet_subject_group)
                            hidden_subject_group_str = hidden_rec.get("subject_group", "")
                            
                            # '증배'가 포함된 경우 교과(군) 체크 제외
                            if "증배" not in sheet_subject_group_str:
                                # 교과(군)이 비어있는 경우
                                if not sheet_subject_group_str:
                                    issues.append({
                                        "severity": "ERROR", 
                                        "sheet": sname, 
                                        "row": rr, 
                                        "message": f"교과(군)(B열)이 비어 있습니다. (지침: '{hidden_subject_group_str}')"
                                    })
                                elif not hidden_subject_group_str:
                                    issues.append({
                                        "severity": "WARNING", 
                                        "sheet": sname, 
                                        "row": rr, 
                                        "message": f"지침에 교과(군) 정보가 없습니다. (시트: '{sheet_subject_group_str}')"
                                    })
                                elif sheet_subject_group_str != hidden_subject_group_str:
                                    issues.append({
                                        "severity": "ERROR", 
                                        "sheet": sname, 
                                        "row": rr, 
                                        "message": f"교과(군) 불일치: 시트(B열)='{sheet_subject_group_str}' / 지침(A열)='{hidden_subject_group_str}'"
                                    })

            # (3) 유형/기본학점/성적처리 (숨김 매칭이 있을 때만)
            if hidden_rec is not None:
                typ_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, type_col)
                typ_s = safe_strip(typ_v)
                if typ_s == "":
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"유형(C{rr})이 비어 있습니다. (지침: {hidden_rec['type']})"})
                elif typ_s != hidden_rec["type"]:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"유형 불일치: 시트='{typ_s}' / 지침='{hidden_rec['type']}'"})

                basic_v, basic_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, basic_col)
                basic_n = to_number(basic_v)
                if basic_n is None:
                    if basic_formula:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rr, "message": f"기본학점(E{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {basic_formula})"})
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"기본학점(E{rr})이 숫자가 아닙니다: {basic_v}"})
                else:
                    if hidden_rec["basic"] is not None and abs(basic_n - hidden_rec["basic"]) > EPS:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"기본학점 불일치: 시트={basic_n:g} / 지침={hidden_rec['basic']:g}"})

                grade_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, grading_col)
                grade_s = safe_strip(grade_v)
                if grade_s == "":
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"성적처리 유형(O{rr})이 비어 있습니다. (지침: {hidden_rec['grading']})"})
                elif grade_s != hidden_rec["grading"]:
                    # 과목명에 괄호가 있는 경우 CHECK로 처리
                    has_parenthesis = "(" in str(course_raw) or ")" in str(course_raw)
                    if has_parenthesis:
                        issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"성적처리 유형 확인. 공동교육과정 등으로 인해 이상없으면 무시 (시트='{grade_s}' / 지침='{hidden_rec['grading']}')"})
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"성적처리 유형 불일치: 시트='{grade_s}' / 지침='{hidden_rec['grading']}'"})

            # (4)(5) 운영학점 범위/합계 체크
            op_v, op_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, op_col)
            op_n = to_number(op_v)
            sem_sum = row_total.get(rr, 0.0)

            if op_n is None:
                if op_formula:
                    issues.append({"severity": "WARNING", "sheet": sname, "row": rr, "message": f"운영학점({op_col_name}{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {op_formula})"})
                else:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점({op_col_name}{rr})이 숫자가 아닙니다: {op_v}"})
            else:
                if hidden_rec is not None and (hidden_rec["min"] is not None) and (hidden_rec["max"] is not None):
                    if not (hidden_rec["min"] - EPS <= op_n <= hidden_rec["max"] + EPS):
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점 범위 위반: 시트={op_n:g} / 허용범위={hidden_rec['min']:g}~{hidden_rec['max']:g}"})

                if abs(sem_sum) <= EPS:
                    # 학기별 합계가 0인 경우, 가장 가까운 행에서 숫자를 찾아서 합산
                    nearest_sum = 0.0
                    found_any = False
                    nearest_row = None
                    
                    # 위쪽으로 검색하면서 숫자가 하나라도 있는 가장 가까운 행 찾기
                    for search_row in range(rr - 1, first_row - 1, -1):
                        row_has_number = False
                        for cc in sem_cols:
                            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, search_row, cc)
                            n = to_number(v)
                            if n is not None and abs(n) > EPS:
                                row_has_number = True
                                break
                        
                        if row_has_number:
                            nearest_row = search_row
                            break
                    
                    # 가장 가까운 행이 발견되면, 그 행의 모든 학기 열 값을 합산
                    if nearest_row is not None:
                        for cc in sem_cols:
                            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, nearest_row, cc)
                            n = to_number(v)
                            if n is not None:
                                nearest_sum += n
                                found_any = True
                    
                    if found_any:
                        # 찾은 숫자들의 합과 운영학점 비교
                        if abs(op_n - nearest_sum) > EPS:
                            # 합이 운영학점의 2배인지 확인 (학기제 가능성)
                            if abs(nearest_sum - op_n * 2) <= EPS:
                                issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"학기 편성 학점의 합({nearest_sum:g})과 운영학점({op_n:g})이 다릅니다.(학기제라면 오류가 아닐 수 있습니다.)"})
                            else:
                                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"학기 편성 학점의 합({nearest_sum:g})과 운영학점({op_n:g})이 다릅니다."})
                    else:
                        # 위쪽에서 숫자를 찾지 못한 경우
                        issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"학기 편성 학점의 합이 0입니다. 선택군 과목 편성 학점을 확인해주세요."})
                else:
                    if abs(op_n - sem_sum) > EPS:
                        # 합이 운영학점의 2배인지 확인 (학기제 가능성)
                        if abs(sem_sum - op_n * 2) <= EPS:
                            issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"학기 편성 학점의 합({sem_sum:g})과 운영학점({op_n:g})이 다릅니다.(학기제라면 오류가 아닐 수 있습니다)"})
                        else:
                            issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"학기 편성 학점의 합({sem_sum:g})과 운영학점({op_n:g})이 다릅니다."})

        # (6) 합계 열 병합 구간 합계 체크 (색깔 행은 기대값에서도 제외)
        checked_spans = set()
        for rng in ws_f.merged_cells.ranges:
            if rng.min_col in total_cols and rng.max_col == rng.min_col:
                col = rng.min_col
                if rng.max_row < first_row:
                    continue
                start = max(rng.min_row, first_row)
                end = min(rng.max_row, check_until_row)
                if start > end:
                    continue

                key = (col, rng.min_row, rng.max_row)
                if key in checked_spans:
                    continue
                checked_spans.add(key)

                total_v, total_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rng.min_row, col)
                total_n = to_number(total_v)

                expected = 0.0
                processed_merges = set()  # 이미 처리한 병합 영역 추적
                for rr in range(start, end + 1):
                    if rr in exempt_rows:
                        continue
                    cv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
                    if cv is None or str(cv).strip() == "" or is_error_token(cv):
                        continue
                    
                    # 병합 셀 확인: sem_cols(학기별 열)에 병합이 있는지 확인
                    # 병합된 셀의 경우 첫 번째 행만 합산
                    should_skip = False
                    for sem_col in sem_cols:
                        key = (rr, sem_col)
                        if key in merge_lookup:
                            min_row, _, max_row, _ = merge_lookup[key]
                            merge_key = (min_row, max_row, sem_col)
                            # 이미 처리한 병합 영역이면 건너뛰기
                            if merge_key in processed_merges:
                                should_skip = True
                                break
                            # 병합 영역의 첫 번째 행이 아니면 건너뛰기
                            if rr != min_row:
                                should_skip = True
                                break
                            # 첫 번째 행이면 병합 영역을 추적에 추가
                            processed_merges.add(merge_key)
                            break
                    
                    if should_skip:
                        continue
                    
                    expected += row_total.get(rr, 0.0)

                # 열 이름 결정
                col_name = total_cols_name.split('/')[0] if col == total_cols[0] else total_cols_name.split('/')[1]
                
                if total_n is None:
                    if total_formula:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rng.min_row, "message": f"{col_name}열 합계 셀에 수식은 있으나 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {total_formula})"})
                    else:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rng.min_row, "message": f"{col_name}열 합계 셀이 비어 있습니다. (해당 구간 {sem_cols_name} 합 기대값={expected:g})"})
                else:
                    if abs(total_n - expected) > EPS:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rng.min_row, "message": f"{col_name}열 합계 불일치: 셀값={total_n:g}, 기대값({sem_cols_name}합)={expected:g} (구간 {start}~{end}행)"})

        # 병합이 아닌 단일 셀 합계 보조 체크(색깔행 제외)
        for col in total_cols:
            for rr in range(first_row, check_until_row + 1):
                if rr in exempt_rows:
                    continue
                if (rr, col) in merge_lookup:
                    min_r, min_c, _, _ = merge_lookup[(rr, col)]
                    if not (rr == min_r and col == min_c):
                        continue
                    continue

                tv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col)
                tn = to_number(tv)
                if tn is None:
                    continue

                cv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
                if cv is None or str(cv).strip() == "" or is_error_token(cv):
                    continue

                # 열 이름 결정
                col_name = total_cols_name.split('/')[0] if col == total_cols[0] else total_cols_name.split('/')[1]
                
                expected = row_total.get(rr, 0.0)
                if abs(tn - expected) > EPS:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"{col_name}열 단일행 합계 불일치: 셀값={tn:g}, 기대값({sem_cols_name}합)={expected:g}"})

        # (7) 총계 구간 찾기: '학생 지정 과목 교과 편성 학점' ~ '학교 선택 과목 교과' 사이
        total_section_start = None
        total_section_end = None
        
        # A열에서 '학생 지정 과목 교과 편성 학점' 또는 '학교 지정 과목 교과 편성 학점' 찾기 (첫 행부터)
        # 더 넓은 검색 조건: "지정", "과목", "교과", "편성", "학점"이 모두 포함되면 찾음
        for rr in range(1, ws_f.max_row + 1):
            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, 1)  # A열
            if v is not None:
                v_str = str(v).strip().replace(" ", "")
                # "지정"과 "편성학점"이 포함되어 있으면 해당 행으로 간주
                if ("지정" in v_str or "선택" in v_str) and "편성학점" in v_str and "과목" in v_str:
                    total_section_start = rr
                    break
        
        # '학교 선택 과목 교과' 찾기 (학생 지정 이후부터)
        if total_section_start is not None:
            for rr in range(total_section_start + 1, ws_f.max_row + 1):
                v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, 1)  # A열
                if v is not None:
                    v_str = str(v).strip().replace(" ", "")
                    if "학교선택과목교과" in v_str:
                        total_section_end = rr - 1  # 그 이전 행까지
                        break
            
            # '학교 선택'을 못 찾으면 마지막 행까지
            if total_section_end is None:
                total_section_end = ws_f.max_row
        
        if total_section_start is not None and total_section_end is not None:
            # 비교 열 병합 정보 수집: 각 행이 어느 병합 범위에 속하는지 매핑
            # 2024: B열, 2025/2026: A열
            # 해당 열만 병합된 경우와 여러 열이 함께 병합된 경우 모두 수집
            a_col_merge_map = {}  # {row: (min_row, max_row)}
            for rng in ws_f.merged_cells.ranges:
                # compare_col을 포함하는 병합
                if rng.min_col <= compare_col and rng.max_col >= compare_col:
                    # 총계 구간과 겹치는 병합만 수집
                    if not (rng.max_row < total_section_start or rng.min_row > total_section_end):
                        for r in range(rng.min_row, rng.max_row + 1):
                            a_col_merge_map[r] = (rng.min_row, rng.max_row)
            
            
            # 총계 열 병합 정보 수집 및 비교 열과 비교
            # 2024: N~O 병합과 B열 비교, 2025/2026: M~N 병합과 A열 비교
            for rng in ws_f.merged_cells.ranges:
                # M열과 N열이 함께 병합된 경우 (min_col=M, max_col=N)
                if rng.min_col == total_cols[0] and rng.max_col == total_cols[1]:
                    # 총계 구간과 겹치는 병합만 검사
                    if not (rng.max_row < total_section_start or rng.min_row > total_section_end):
                        merge_start = rng.min_row
                        merge_end = rng.max_row
                        
                        # 비교 열(2024:B열, 2025/2026:A열)과 D열 값 확인 - 특정 키워드 포함 시 검사 제외
                        compare_col_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, merge_start, compare_col)
                        d_col_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, merge_start, 4)  # D열
                        skip_check = False
                        
                        # 비교 열 체크
                        if compare_col_val is not None:
                            compare_col_str = str(compare_col_val).strip()
                            compare_col_str_no_space = compare_col_str.replace(" ", "")
                            # '학교/학생 지정 과목 편성' 또는 '증배' 포함 시 제외
                            if "학교지정과목편성" in compare_col_str_no_space or "학생지정과목편성" in compare_col_str_no_space or "증배" in compare_col_str:
                                skip_check = True
                        
                        # D열 체크 - 총계 행 제외
                        if d_col_val is not None and not skip_check:
                            d_col_str = str(d_col_val).strip().replace(" ", "")
                            # 총계 관련 키워드가 있으면 제외
                            if "편성학점" in d_col_str or "총교과" in d_col_str or "창의적체험활동" in d_col_str or "편성학점수" in d_col_str:
                                skip_check = True
                        
                        if skip_check:
                            continue  # 이 병합 구간은 검사하지 않음 (병합 불일치도, 합계도 검사 안함)
                        
                        # 총계 열 병합 범위 내의 모든 행이 같은 비교 열 병합에 속하는지 확인
                        compare_merge_range = None
                        mismatch = False
                        
                        for r in range(merge_start, merge_end + 1):
                            if r in a_col_merge_map:
                                current_range = a_col_merge_map[r]
                                if compare_merge_range is None:
                                    compare_merge_range = current_range
                                elif compare_merge_range != current_range:
                                    # 다른 비교 열 병합 범위에 걸쳐있음
                                    mismatch = True
                                    break
                            else:
                                # 비교 열이 병합되지 않은 행
                                mismatch = True
                                break
                        
                        # 비교 열 병합 범위와 총계 열 병합 범위가 정확히 일치하는지 확인
                        if mismatch or compare_merge_range is None or compare_merge_range != (merge_start, merge_end):
                            compare_range_str = f"{compare_merge_range[0]}~{compare_merge_range[1]}" if compare_merge_range else "없음"
                            compare_col_name = "B" if year == 2024 else "A"
                            issues.append({
                                "severity": "ERROR",
                                "sheet": sname,
                                "row": merge_start,
                                "message": f"병합 불일치: {compare_col_name}열 병합({compare_range_str})과 {total_cols_name}열 병합({merge_start}~{merge_end}행)이 일치하지 않습니다."
                            })
                        
                        # 합계 검증 (첫 번째 총계 열 기준: 2024=N열, 2025/2026=M열)
                        # '증배'나 '학교 지정 과목 편성'은 이미 위에서 skip_check으로 제외됨
                        total_v, total_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, merge_start, total_cols[0])
                        total_n = to_number(total_v)
                        
                        if total_n is not None:
                            # 해당 구간의 기대값 계산
                            expected = 0.0
                            for rr in range(merge_start, merge_end + 1):
                                if rr in exempt_rows:
                                    continue
                                expected += row_total.get(rr, 0.0)
                            
                            if abs(total_n - expected) > EPS:
                                first_col_name = "N" if year == 2024 else "M"
                                issues.append({
                                    "severity": "ERROR",
                                    "sheet": sname,
                                    "row": merge_start,
                                    "message": f"{first_col_name}열 합계 불일치: 셀값={total_n:g}, {sem_cols_name}합={expected:g} (구간 {merge_start}~{merge_end}행)(편성학점 칸이 병합되어 있는지 확인하세요.)"
                                })
            
        # =========================
        # (8) 총계 행 합계 검증
        # =========================
        
        # A열(2025/2026) 또는 B열(2024)에서 총계 행들 찾기 (필수 셀 존재 여부 확인)
        total_rows = {}  # {"학교지정": row, "학생선택": row, "총교과": row, "창의적": row, "편성학점수": row}
        
        required_cells = {
            "학교지정": "'학교 지정 과목 교과 편성 학점'",
            "학생선택": "'학생 선택 과목 교과 편성 학점'",
            "총교과": "'총 교과 편성 학점'",
            "창의적": "'창의적 체험활동 학점'",
            "편성학점수": "'편성 학점 수'"
        }
        
        # A열(2025/2026) 또는 B열(2024) 확인
        check_col = 1 if year in [2025, 2026] else 2
        
        for rr in range(first_row, ws_f.max_row + 1):
            col_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, check_col)
            if col_val:
                col_str = str(col_val).strip().replace(" ", "")
                
                # 안내 문구나 긴 텍스트 제외 (실제 총계 셀은 짧고 명확함)
                if len(col_str) > 30:  # 너무 긴 텍스트는 제외
                    continue
                if any(word in col_str for word in ["확인", "제대로", "다시", "주의", "주세요", "입력", "양식"]):
                    continue
                
                if ("학교지정" in col_str or "학교선택" in col_str) and "편성학점" in col_str and "과목" in col_str and "교과" in col_str:
                    total_rows["학교지정"] = rr
                elif "학생선택" in col_str and "편성학점" in col_str and "과목" in col_str and "교과" in col_str:
                    total_rows["학생선택"] = rr
                elif "총교과편성학점" in col_str or ("총교과" in col_str and "편성학점" in col_str and "과목" not in col_str):
                    total_rows["총교과"] = rr
                elif "창의적체험활동" in col_str and "학점" in col_str and "과목" not in col_str:
                    total_rows["창의적"] = rr
                elif "편성학점수" in col_str and "과목" not in col_str and "교과" not in col_str:
                    total_rows["편성학점수"] = rr
        
        # 필수 셀 존재 여부 확인
        for key, cell_name in required_cells.items():
            if key not in total_rows:
                issues.append({
                    "severity": "ERROR",
                    "sheet": sname,
                    "row": "-",
                    "message": f"총계 부분의 {cell_name} 셀이 존재하지 않습니다. 교육청의 양식을 확인하여 수정하고 다시 검사를 진행해주세요."
                })
        
        # 총계 행 검증
        if "학교지정" in total_rows:
            school_row = total_rows["학교지정"]
            
            # 학교 지정 과목: 위의 행들 합계 (first_row ~ school_row-1)
            for col_idx, col_letter in enumerate(sem_cols):
                expected_sum = 0.0
                processed_merges = set()
                
                for rr in range(first_row, school_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_lookup:
                        min_row, _, max_row, _ = merge_lookup[key]
                        merge_key = (min_row, max_row, col_letter)
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                
                actual_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, school_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)  # 열 번호를 문자로 변환
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": school_row,
                        "message": f"학교 지정 과목 편성 학점 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값={expected_sum:g}"
                    })
            
            # M/N열 (또는 N/O열) 합계 = G~L (또는 H~M) 합
            total_col = total_cols[0]  # M열 또는 N열
            sem_sum = 0.0
            for col_letter in sem_cols:
                val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, school_row, col_letter)
                num = to_number(val)
                if num is not None:
                    sem_sum += num
            
            actual_total, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, school_row, total_col)
            actual_total_num = to_number(actual_total)
            
            if actual_total_num is not None and abs(actual_total_num - sem_sum) > EPS:
                total_col_name = chr(64 + total_col)
                issues.append({
                    "severity": "ERROR",
                    "sheet": sname,
                    "row": school_row,
                    "message": f"학교 지정 과목 편성 학점 {total_col_name}열 합계 오류: 셀값={actual_total_num:g}, 기대값({sem_cols_name}합)={sem_sum:g}"
                })
        
        # 학생 선택 과목 검증
        if "학생선택" in total_rows and "학교지정" in total_rows:
            student_row = total_rows["학생선택"]
            school_row = total_rows["학교지정"]
            
            # 학생 선택 과목: school_row+1 ~ student_row-1 합계 (증배 제외)
            for col_idx, col_letter in enumerate(sem_cols):
                expected_sum = 0.0
                processed_merges = set()
                
                for rr in range(school_row + 1, student_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 증배 확인
                    a_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, compare_col)
                    if a_val and "증배" in str(a_val):
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_lookup:
                        min_row, _, max_row, _ = merge_lookup[key]
                        merge_key = (min_row, max_row, col_letter)
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                
                actual_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, student_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": student_row,
                        "message": f"학생 선택 과목 편성 학점 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값={expected_sum:g} (증배 제외)"
                    })
            
            # M/N열 합계
            sem_sum = 0.0
            for col_letter in sem_cols:
                val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, student_row, col_letter)
                num = to_number(val)
                if num is not None:
                    sem_sum += num
            
            total_col = total_cols[0]
            actual_total, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, student_row, total_col)
            actual_total_num = to_number(actual_total)
            
            if actual_total_num is not None and abs(actual_total_num - sem_sum) > EPS:
                total_col_name = chr(64 + total_col)
                issues.append({
                    "severity": "ERROR",
                    "sheet": sname,
                    "row": student_row,
                    "message": f"학생 선택 과목 편성 학점 {total_col_name}열 합계 오류: 셀값={actual_total_num:g}, 기대값({sem_cols_name}합)={sem_sum:g}"
                })
        
        # 총 교과 편성 학점 검증
        if "총교과" in total_rows and "학교지정" in total_rows and "학생선택" in total_rows:
            total_subject_row = total_rows["총교과"]
            school_row = total_rows["학교지정"]
            student_row = total_rows["학생선택"]
            
            # 각 열별로 학교 지정과 학생 선택의 기댓값을 저장
            school_expected = {}
            student_expected = {}
            
            # 학교 지정 과목 기댓값 계산
            for col_letter in sem_cols:
                expected_sum = 0.0
                processed_merges = set()
                
                for rr in range(first_row, school_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_lookup:
                        min_row, _, max_row, _ = merge_lookup[key]
                        merge_key = (min_row, max_row, col_letter)
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                school_expected[col_letter] = expected_sum
            
            # 학생 선택 과목 기댓값 계산
            for col_letter in sem_cols:
                expected_sum = 0.0
                processed_merges = set()
                
                for rr in range(school_row + 1, student_row):
                    if rr in exempt_rows:
                        continue
                    
                    # 증배 확인
                    a_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, compare_col)
                    if a_val and "증배" in str(a_val):
                        continue
                    
                    # 병합 셀 확인
                    key = (rr, col_letter)
                    if key in merge_lookup:
                        min_row, _, max_row, _ = merge_lookup[key]
                        merge_key = (min_row, max_row, col_letter)
                        if merge_key in processed_merges:
                            continue
                        processed_merges.add(merge_key)
                    
                    val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col_letter)
                    num = to_number(val)
                    if num is not None:
                        expected_sum += num
                student_expected[col_letter] = expected_sum
            
            # 총 교과 = 학교 지정 기댓값 + 학생 선택 기댓값
            for col_letter in sem_cols:
                expected_sum = school_expected.get(col_letter, 0.0) + student_expected.get(col_letter, 0.0)
                
                actual_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, total_subject_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": total_subject_row,
                        "message": f"총 교과 편성 학점 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값(학교지정+학생선택)={expected_sum:g}"
                    })
            
            # M/N열 합계 = 각 열의 총 교과 기댓값 합
            expected_total = sum(school_expected.get(col, 0.0) + student_expected.get(col, 0.0) for col in sem_cols)
            
            total_col = total_cols[0]
            actual_total, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, total_subject_row, total_col)
            actual_total_num = to_number(actual_total)
            
            if actual_total_num is not None and abs(actual_total_num - expected_total) > EPS:
                total_col_name = chr(64 + total_col)
                issues.append({
                    "severity": "ERROR",
                    "sheet": sname,
                    "row": total_subject_row,
                    "message": f"총 교과 편성 학점 {total_col_name}열 합계 오류: 셀값={actual_total_num:g}, 기대값={expected_total:g}"
                })
        
        # 창의적 체험활동 검증
        if "창의적" in total_rows:
            creative_row = total_rows["창의적"]
            
            # G~L (또는 H~M) 각 학기 열 검증: 값 존재 여부 및 1~5 범위 확인
            semester_values = []
            for col_letter in sem_cols:
                val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, creative_row, col_letter)
                num = to_number(val)
                col_name = chr(64 + col_letter)
                
                # 값이 없으면 오류
                if num is None:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": creative_row,
                        "message": f"창의적 체험활동 학점 {col_name}열에 값이 없습니다."
                    })
                # 값이 있지만 1~5 범위가 아니면 오류 + 경고
                elif num < 1.0 - EPS or num > 5.0 + EPS:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": creative_row,
                        "message": f"창의적 체험활동 학점 {col_name}열 오류: 셀값={num:g}, 허용 범위=1~5"
                    })
                    issues.append({
                        "severity": "WARNING",
                        "sheet": sname,
                        "row": creative_row,
                        "message": f"창의적 체험활동 학점 {col_name}열은 1~5 범위 내로 설정해야 합니다."
                    })
                else:
                    # 값이 있고 1~5 범위 내면 정상 (합계 계산을 위해 저장)
                    semester_values.append(num)
            
            # 합계 검증: 각 학기 열의 합과 총합 열 비교
            # total_cols 리스트의 모든 열을 확인하여 값이 있는 열을 찾음
            actual_total_num = None
            total_col = None
            total_col_name = None
            
            for col in total_cols:
                val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, creative_row, col)
                num = to_number(val)
                if num is not None:
                    actual_total_num = num
                    total_col = col
                    total_col_name = chr(64 + col)
                    break
            
            # 각 학기 열의 합 계산
            expected_total = sum(semester_values)
            
            if actual_total_num is None:
                # 모든 total_cols 열에 값이 없는 경우
                col_names = "/".join([chr(64 + col) for col in total_cols])
                issues.append({
                    "severity": "ERROR",
                    "sheet": sname,
                    "row": creative_row,
                    "message": f"창의적 체험활동 학점 총합 열({col_names})에 값이 없습니다."
                })
            elif abs(actual_total_num - expected_total) > EPS:
                issues.append({
                    "severity": "ERROR",
                    "sheet": sname,
                    "row": creative_row,
                    "message": f"창의적 체험활동 학점 {total_col_name}열 합계 오류: 셀값={actual_total_num:g}, 기대값(각 학기 열 합)={expected_total:g}"
                })
        
        # 편성 학점 수 검증
        if "편성학점수" in total_rows and "학교지정" in total_rows and "학생선택" in total_rows:
            final_row = total_rows["편성학점수"]
            school_row = total_rows["학교지정"]
            student_row = total_rows["학생선택"]
            
            # 학교 지정과 학생 선택의 기댓값이 이미 위에서 계산되었는지 확인
            # 만약 총교과 검증을 거치지 않았다면 여기서 계산
            if 'school_expected' not in locals() or 'student_expected' not in locals():
                school_expected = {}
                student_expected = {}
                
                # 학교 지정 과목 기댓값 계산
                for col_letter in sem_cols:
                    expected_sum = 0.0
                    processed_merges = set()
                    
                    for rr in range(first_row, school_row):
                        if rr in exempt_rows:
                            continue
                        
                        # 병합 셀 확인
                        key = (rr, col_letter)
                        if key in merge_lookup:
                            min_row, _, max_row, _ = merge_lookup[key]
                            merge_key = (min_row, max_row, col_letter)
                            if merge_key in processed_merges:
                                continue
                            processed_merges.add(merge_key)
                        
                        val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col_letter)
                        num = to_number(val)
                        if num is not None:
                            expected_sum += num
                    school_expected[col_letter] = expected_sum
                
                # 학생 선택 과목 기댓값 계산
                for col_letter in sem_cols:
                    expected_sum = 0.0
                    processed_merges = set()
                    
                    for rr in range(school_row + 1, student_row):
                        if rr in exempt_rows:
                            continue
                        
                        # 증배 확인
                        a_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, compare_col)
                        if a_val and "증배" in str(a_val):
                            continue
                        
                        # 병합 셀 확인
                        key = (rr, col_letter)
                        if key in merge_lookup:
                            min_row, _, max_row, _ = merge_lookup[key]
                            merge_key = (min_row, max_row, col_letter)
                            if merge_key in processed_merges:
                                continue
                            processed_merges.add(merge_key)
                        
                        val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col_letter)
                        num = to_number(val)
                        if num is not None:
                            expected_sum += num
                    student_expected[col_letter] = expected_sum
            
            for col_idx, col_letter in enumerate(sem_cols):
                # 총교과 기댓값 + 창의적(3)
                total_expected = school_expected.get(col_letter, 0.0) + student_expected.get(col_letter, 0.0)
                expected_sum = total_expected + 3.0
                
                actual_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, final_row, col_letter)
                actual_num = to_number(actual_val)
                
                if actual_num is not None and abs(actual_num - expected_sum) > EPS:
                    col_name = chr(64 + col_letter)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": final_row,
                        "message": f"편성 학점 수 {col_name}열 합계 오류: 셀값={actual_num:g}, 기대값(총교과+창의적)={expected_sum:g}"
                    })
            
            # M/N열 합계 체크 = 총 교과 기댓값 합 + 창의적(18)
            total_col = total_cols[0]
            actual_final, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, final_row, total_col)
            actual_final_num = to_number(actual_final)
            
            # 기댓값 = 각 열의 (학교지정 + 학생선택) 합 + 18
            expected_final_total = sum(school_expected.get(col, 0.0) + student_expected.get(col, 0.0) for col in sem_cols) + 18.0
            
            if actual_final_num is not None:
                # 합계 체크
                if abs(actual_final_num - expected_final_total) > EPS:
                    total_col_name = chr(64 + total_col)
                    issues.append({
                        "severity": "ERROR",
                        "sheet": sname,
                        "row": final_row,
                        "message": f"편성 학점 수 {total_col_name}열 합계 오류: 셀값={actual_final_num:g}, 기대값(총교과+창의적)={expected_final_total:g}"
                    })

    # =========================
    # (9) 2026 전학년 시트 검증
    # =========================
    check_all_grades_sheet(wb_v, wb_f, targets, issues, hidden, vocational_courses, new_courses, hidden_list_norm)

    # =========================
    # (10) 학교명 일관성 검증
    # =========================
    check_school_name_consistency(wb_v, wb_f, targets, issues)

    # =========================
    # (11) 2025, 2026 입학생 시트 최신버전 확인
    # =========================
    show_version_warning = False
    if data_source == "엑셀 파일 내부":
        for year in [2025, 2026]:
            if year not in targets:
                continue
            sname = targets[year]
            
            # 해당 시트의 error + warning 개수 세기
            error_warning_count = sum(
                1 for issue in issues 
                if issue.get("sheet") == sname 
                and issue.get("severity") in ["ERROR", "WARNING"]
            )
            
            if error_warning_count >= 50:
                show_version_warning = True
                break
    
    summary["show_version_warning"] = show_version_warning

    return issues, summary


# =========================
# GUI
# =========================

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("교육과정 편성표 확인 프로그램")
        self.root.minsize(1020, 700)

        self.colors = {
            "bg": "#F6F7FF",
            "card": "#FFFFFF",
            "text": "#1F2937",
            "muted": "#6B7280",
            "accent": "#7C6CF6",
            "danger": "#EF4444",
            "warn": "#F59E0B",
            "check": "#3B82F6",
        }
        self.root.configure(bg=self.colors["bg"])

        self.style = ttk.Style(self.root)
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        base_font = ("Malgun Gothic", 11)
        title_font = ("Malgun Gothic", 16, "bold")

        self.style.configure("TFrame", background=self.colors["bg"])
        self.style.configure("Card.TFrame", background=self.colors["card"])
        self.style.configure("TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=base_font)
        self.style.configure("Title.TLabel", font=title_font, foreground=self.colors["text"])
        self.style.configure("Muted.TLabel", foreground=self.colors["muted"], font=("Malgun Gothic", 10))

        self._build_ui()
        self.xlsx_path = None

    def _build_ui(self):
        header = ttk.Frame(self.root, padding=(18, 18, 18, 10))
        header.pack(fill="x")

        ttk.Label(header, text="교육과정 편성표 확인 프로그램", style="Title.TLabel").pack(anchor="w")
        
        text_frame = ttk.Frame(header)
        text_frame.pack(fill="x", pady=(6, 0))
        
        # 강조 부분 1: "교육청에서 제공한 엑셀 파일"
        tk.Label(
            text_frame,
            text="교육청에서 제공한 엑셀 파일",
            bg=self.colors["bg"],
            fg="#EF4444",
            font=("Malgun Gothic", 10, "bold")
        ).pack(side="left")
        
        # 일반 텍스트 부분
        ttk.Label(
            text_frame,
            text="에 작성된 편성표를 점검합니다. 파일을 ",
            style="Muted.TLabel"
        ).pack(side="left")
        
        # 강조 부분 2: "저장하고 닫은 후에"
        tk.Label(
            text_frame,
            text="저장하고 닫은 후에",
            bg=self.colors["bg"],
            fg="#EF4444",
            font=("Malgun Gothic", 10, "bold")
        ).pack(side="left")
        
        # 일반 텍스트 부분
        ttk.Label(
            text_frame,
            text=" 업로드하세요.",
            style="Muted.TLabel"
        ).pack(side="left")
        
        # 두 번째 줄 텍스트
        text_frame2 = ttk.Frame(header)
        text_frame2.pack(fill="x", pady=(2, 0))
        
        ttk.Label(
            text_frame2,
            text="프로그램이 잘못 판단할 수 있으니, 확인했을 때 이상이 없다면 오류를 무시하셔도 됩니다.",
            style="Muted.TLabel"
        ).pack(anchor="w")

        download_btn = tk.Button(
            text_frame,
            text="편성표 양식 다운로드",
            command=lambda: webbrowser.open("https://drive.google.com/drive/folders/1wvdV4VQD7kUD7eVEvypPf39LDWZLxfze?usp=sharing"),
            bg=self.colors["accent"],
            fg="white",
            bd=0,
            activebackground=self.colors["accent"],
            activeforeground="white",
            padx=12,
            pady=6,
            font=("Malgun Gothic", 10),
            cursor="hand2"
        )
        download_btn.pack(side="right", padx=(10, 0))

        body = ttk.Frame(self.root, padding=(18, 8, 18, 18))
        body.pack(fill="both", expand=True)

        card = ttk.Frame(body, style="Card.TFrame", padding=(16, 16))
        card.pack(fill="x")

        row1 = ttk.Frame(card, style="Card.TFrame")
        row1.pack(fill="x")

        self.path_var = tk.StringVar(value="선택된 파일 없음")
        ttk.Label(row1, text="엑셀 파일:", style="Muted.TLabel").pack(side="left")
        ttk.Label(row1, textvariable=self.path_var).pack(side="left", padx=(8, 0))

        btn_frame = ttk.Frame(card, style="Card.TFrame")
        btn_frame.pack(fill="x", pady=(12, 0))

        self.btn_pick = tk.Button(
            btn_frame,
            text="파일 선택",
            command=self.pick_file,
            bg=self.colors["accent"],
            fg="white",
            bd=0,
            activebackground=self.colors["accent"],
            activeforeground="white",
            padx=16,
            pady=10,
            font=("Malgun Gothic", 11, "bold"),
            cursor="hand2"
        )
        self.btn_pick.pack(side="left")

        self.btn_run = tk.Button(
            btn_frame,
            text="검사 실행",
            command=self.run,
            bg="#C7C9D9",
            fg="white",
            bd=0,
            activebackground="#C7C9D9",
            activeforeground="white",
            padx=16,
            pady=10,
            font=("Malgun Gothic", 11, "bold"),
            cursor="hand2",
            state="disabled"
        )
        self.btn_run.pack(side="left", padx=(10, 0))

        status_frame = ttk.Frame(card, style="Card.TFrame")
        status_frame.pack(fill="x", pady=(12, 0))

        self.status_var = tk.StringVar(value="대기 중")
        ttk.Label(status_frame, textvariable=self.status_var, style="Muted.TLabel").pack(side="left")

        self.progress = ttk.Progressbar(status_frame, mode="indeterminate", length=220)
        self.progress.pack(side="right")

        # 결과: Notebook(탭)으로 시트별 출력
        out_card = ttk.Frame(body, style="Card.TFrame", padding=(16, 16))
        out_card.pack(fill="both", expand=True, pady=(14, 0))

        ttk.Label(out_card, text="문제상황(시트별)", style="Muted.TLabel").pack(anchor="w")

        self.nb = ttk.Notebook(out_card)
        self.nb.pack(fill="both", expand=True, pady=(8, 0))

        self.text_widgets = {}  # tab_name -> ScrolledText

        # 기본 탭: 전체/기타 (실행 시 대상 시트 탭은 동적으로 재구성)
        self._reset_tabs(["전체", "기타"])

    def _reset_tabs(self, tab_names):
        # 기존 탭 제거
        for tab_id in self.nb.tabs():
            self.nb.forget(tab_id)
        self.text_widgets.clear()

        for name in tab_names:
            frame = ttk.Frame(self.nb, padding=(8, 8))
            self.nb.add(frame, text=name)

            txt = ScrolledText(
                frame,
                wrap="word",
                height=18,
                font=("Malgun Gothic", 10),
                bg="#FBFBFE",
                fg=self.colors["text"],
                relief="solid",
                bd=1,
                padx=10,
                pady=10
            )
            txt.pack(fill="both", expand=True)
            txt.tag_configure("ERROR", foreground=self.colors["danger"], font=("Malgun Gothic", 10))
            txt.tag_configure("WARNING", foreground=self.colors["warn"], font=("Malgun Gothic", 10))
            txt.tag_configure("CHECK", foreground=self.colors["check"], font=("Malgun Gothic", 10))
            txt.tag_configure("INFO", foreground=self.colors["muted"], font=("Malgun Gothic", 9))
            txt.tag_configure("HEADER", font=("Malgun Gothic", 11, "bold"), foreground="#5B21B6")
            txt.tag_configure("COURSE", font=("Malgun Gothic", 10, "bold"), foreground="#7C3AED")
            txt.tag_configure("RIGHT_ALIGN", justify="right", font=("Malgun Gothic", 9), foreground=self.colors["muted"])
            self.text_widgets[name] = txt

    def pick_file(self):
        path = filedialog.askopenfilename(
            title="교육과정 편성표 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return
        self.xlsx_path = path
        self.path_var.set(path)
        self.btn_run.configure(state="normal", bg=self.colors["accent"], activebackground=self.colors["accent"])

    def run(self):
        if not self.xlsx_path:
            messagebox.showwarning("안내", "먼저 엑셀 파일을 선택하세요.")
            return

        # 초기화(탭은 summary 읽고 나서 구성)
        self.status_var.set("검사 중...")
        self.progress.start(12)
        self.root.update_idletasks()

        try:
            issues, summary = run_checks(self.xlsx_path)
        except Exception as e:
            import traceback
            self.progress.stop()
            self.status_var.set("오류 발생")
            messagebox.showerror("오류", f"검사 중 예외가 발생했습니다:\n{e}\n\n{traceback.format_exc()}")
            return

        self.progress.stop()

        # 결과 타입 검증
        if not isinstance(issues, list):
            messagebox.showerror("오류", f"검사 결과 타입 오류: issues가 list가 아닙니다. (타입: {type(issues)})")
            return
        if not isinstance(summary, dict):
            messagebox.showerror("오류", f"검사 결과 타입 오류: summary가 dict가 아닙니다. (타입: {type(summary)})")
            return

        try:
            # 탭 구성: 전체 + (대상 시트) + 2026 전학년 + 기타
            targets = (summary.get("targets") or {})
            tab_names = ["전체"]
            # 연도 시트(존재하는 것만)
            for y in (2026, 2025, 2024):
                s = targets.get(y)
                if s and s not in tab_names:
                    tab_names.append(s)
            
            # 2026 전학년 시트 추가
            try:
                wb_temp = load_workbook(self.xlsx_path, data_only=True)
                all_grades_sheet = find_all_grades_sheet(wb_temp.sheetnames)
                if all_grades_sheet and all_grades_sheet not in tab_names:
                    tab_names.append(all_grades_sheet)
            except Exception:
                pass
            
            tab_names.append("기타")
            self._reset_tabs(tab_names)

            # 출력
            self._print_summary(summary, issues)
            self._print_issues_per_sheet(issues, summary)
        except Exception as e:
            import traceback
            self.progress.stop()
            self.status_var.set("오류 발생")
            messagebox.showerror("오류", f"결과 출력 중 예외가 발생했습니다:\n{e}\n\n상세:\n{traceback.format_exc()}")
            return

        err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
        warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")
        check_cnt = sum(1 for x in issues if x.get("severity") == "CHECK")
        if err_cnt == 0:
            self.status_var.set(f"검사 완료: 오류 없음 (경고 {warn_cnt}건, 확인 {check_cnt}건)")
        else:
            self.status_var.set(f"검사 완료: 오류 {err_cnt}건 / 경고 {warn_cnt}건 / 확인 {check_cnt}건")

        # 기본으로 "전체" 탭 보여주기
        try:
            self.nb.select(0)
        except Exception:
            pass

    def _w(self, tab, text, tag="INFO"):
        txt = self.text_widgets.get(tab)
        if not txt:
            txt = self.text_widgets.get("기타")
        txt.insert("end", text, tag)

    def _print_summary(self, summary, issues):
        tab = "전체"
        txt = self.text_widgets[tab]
        txt.delete("1.0", "end")

        self._w(tab, "[검사 개요]\n", "HEADER")
        self._w(tab, f"- 파일: {self.xlsx_path}\n", "INFO")

        # 안내 메시지 (최신버전 확인)
        if summary.get("show_version_warning", False):
            self._w(tab, "\n[안내]\n", "HEADER")
            self._w(tab, "참조되는 시트가 최신버전이 아닙니다. 교육청 제공 양식에 다시 작성해주세요.\n\n", "WARNING")
        
        # 구글 스프레드시트 참조 실패 안내
        google_error = summary.get("google_error")
        if google_error:
            self._w(tab, "\n[온라인 데이터 참조 안내]\n", "HEADER")
            # requests 라이브러리 관련 메시지는 간단하게 표시
            if "온라인 데이터 참조 기능을 사용할 수 없습니다" in google_error:
                self._w(tab, "온라인 데이터를 참조할 수 없어 엑셀 파일 내부의 '숨김' 및 '전문교과목록' 시트를 사용합니다.\n\n", "INFO")
            else:
                self._w(tab, f"구글 스프레드시트를 참조하지 못했습니다: {google_error}\n", "WARNING")
                self._w(tab, "엑셀 파일 내부의 '숨김' 및 '전문교과목록' 시트를 사용합니다.\n\n", "INFO")

        targets = summary.get("targets") or {}
        self._w(tab, "- 시트 확인:\n", "INFO")
        for y in (2026, 2025, 2024):
            s = targets.get(y)
            if s:
                self._w(tab, f"  · {y}: {s}\n", "INFO")
            else:
                self._w(tab, f"  · {y}: (없음)\n", "WARNING")

        hidden = summary.get("hidden_sheet")
        cnt = summary.get("hidden_course_count", 0)
        data_source = summary.get("data_source", "알 수 없음")
        vocational_cnt = summary.get("vocational_course_count", 0)
        
        if hidden:
            self._w(tab, f"- 지침 시트: {hidden} (과목 {cnt}개)\n", "INFO")
            self._w(tab, f"- 전문교과목록: {vocational_cnt}개 과목\n", "INFO")
            self._w(tab, f"- 데이터 출처: {data_source}\n", "INFO")
        else:
            self._w(tab, "- 지침 시트: (없음)\n", "ERROR")

        err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
        warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")
        check_cnt = sum(1 for x in issues if x.get("severity") == "CHECK")
        self._w(tab, f"- 총계: 오류 {err_cnt}건 / 경고 {warn_cnt}건 / 확인 {check_cnt}건\n\n", "INFO")

        self._w(tab, "[시트별 안내]\n", "HEADER")
        self._w(tab, "- 각 탭에서 해당 시트의 문제상황만 확인할 수 있습니다.\n", "INFO")
        self._w(tab, "- '기타' 탭에는 파일/시트 누락 등 특정 시트에 귀속되지 않는 오류가 표시됩니다.\n\n", "INFO")

    def _print_issues_per_sheet(self, issues, summary):
        # 다른 탭들은 내용만 초기화(전체는 summary가 있으므로 유지)
        for name, txt in self.text_widgets.items():
            if name == "전체":
                continue
            txt.delete("1.0", "end")

        if not issues:
            self._w("전체", "문제 없음.\n", "INFO")
            return

        # summary가 딕셔너리가 아닌 경우 처리
        if not isinstance(summary, dict):
            summary = {}

        # 그룹핑
        groups = {}
        for it in issues:
            # it이 딕셔너리가 아닌 경우 처리
            if not isinstance(it, dict):
                continue
            sheet = it.get("sheet", "-") or "-"
            groups.setdefault(sheet, []).append(it)

        sev_rank = {"ERROR": 0, "WARNING": 1, "CHECK": 2, "INFO": 3}

        def sort_key(x):
            row = x.get("row", "-")
            try:
                row_n = int(row)
            except Exception:
                row_n = 10**9
            return (sev_rank.get(x.get("severity", "INFO"), 9), row_n)

        # 2024, 2025, 2026 시트 및 2026 전학년 시트 확인
        targets = summary.get("targets") or {}
        sheet_2024 = targets.get(2024)
        sheet_2025 = targets.get(2025)
        sheet_2026 = targets.get(2026)
        # 모든 시트 이름 목록 생성 (groups에 있는 시트 + text_widgets에 있는 시트)
        all_sheet_names = set(groups.keys()) | set(self.text_widgets.keys())
        all_grades_sheet = find_all_grades_sheet(list(all_sheet_names))
        
        # 먼저 모든 시트에 안내 메시지 출력 (오류가 없어도 출력)
        for tab_name in self.text_widgets.keys():
            if tab_name == "전체" or tab_name == "기타":
                continue
            
            # 2026 시트인 경우 안내 메시지 출력
            if tab_name == sheet_2026 and sheet_2026:
                self._w(tab_name, "[안내]\n", "HEADER")
                self._w(tab_name, "교차이수과목의 경우 ↔ 왼쪽 과목을 윗줄, 오른쪽 과목을 아랫줄으로 판단합니다.\n\n", "INFO")
            
            # 2025 시트인 경우 안내 메시지 출력
            if tab_name == sheet_2025 and sheet_2025:
                self._w(tab_name, "[안내]\n", "HEADER")
                self._w(tab_name, "교차이수과목의 경우 ↔ 왼쪽 과목을 윗줄, 오른쪽 과목을 아랫줄으로 판단합니다.\n\n", "INFO")
            
            # 2024 시트인 경우 안내 메시지 출력
            if tab_name == sheet_2024 and sheet_2024:
                self._w(tab_name, "[안내]\n", "HEADER")
                self._w(tab_name, "2015개정 교육과정의 과목명의 경우는 일치 여부를 확인하지 않습니다.\n", "INFO")
                self._w(tab_name, "지침의 표를 확인하고 정확하게 입력해주세요.\n\n", "INFO")
            
            # 2026 전학년 시트인 경우 안내 메시지 출력
            if tab_name == all_grades_sheet and all_grades_sheet:
                self._w(tab_name, "[안내]\n", "HEADER")
                self._w(tab_name, "개설 여부는 프로그램 상 확인 절차가 따로 없습니다. 선택군은 학년별로 다르게 정리하여 병합해해주세요.\n\n", "INFO")

        # 각 시트 탭에 출력
        for sheet, items in groups.items():
            tab = sheet if sheet in self.text_widgets else "기타"
            
            # 해당 시트의 오류 개수 확인
            error_count = sum(1 for it in items if it.get("severity") == "ERROR")
            
            # 오류가 50개 이상이면 경고 메시지 출력
            if error_count >= 50:
                self._w(tab, "[경고] ", "ERROR")
                self._w(tab, f"오류가 50개 이상 발견됩니다. 양식이 올바른지 확인해주세요.(교육청 양식 참고)\n\n", "WARNING")
            
            # 행 번호별로 그룹핑
            row_groups = {}
            
            for it in sorted(items, key=sort_key):
                row = it.get("row", "-")
                row_groups.setdefault(row, []).append(it)
            
            self._w(tab, "[문제 목록]\n", "HEADER")
            
            # 엑셀 파일에서 행 정보를 읽어오기 위한 준비
            try:
                from openpyxl import load_workbook
                wb_temp = load_workbook(self.xlsx_path, data_only=True)
            except Exception:
                wb_temp = None
            
            # 행 번호별로 출력
            for row_num, row_items in sorted(row_groups.items(), key=lambda x: (x[0] == "-", int(x[0]) if str(x[0]).isdigit() else 10**9, x[0])):
                # 행 정보 추출
                row_label = None
                
                if row_num != "-" and str(row_num).isdigit() and wb_temp and sheet in wb_temp.sheetnames:
                    try:
                        ws = wb_temp[sheet]
                        row_int = int(row_num)
                        
                        # 과목명 열 결정: 2024 시트는 E열(5), 나머지는 D열(4)
                        course_col = 5 if sheet == sheet_2024 else 4
                        
                        # 과목명
                        course_cell = ws.cell(row_int, course_col).value
                        if course_cell and str(course_cell).strip():
                            course_name = normalize_course_name(course_cell)
                            if course_name:
                                row_label = course_name
                        
                        # 2024 시트에서 E열에 과목명이 없으면 D열도 확인
                        if not row_label and sheet == sheet_2024:
                            d_cell = ws.cell(row_int, 4).value
                            if d_cell and str(d_cell).strip():
                                d_name = normalize_course_name(d_cell)
                                if d_name:
                                    row_label = d_name
                        
                        # 과목명이 없으면 A열(1) 또는 B열(2) 확인
                        if not row_label:
                            a_cell = ws.cell(row_int, 1).value
                            if a_cell and str(a_cell).strip():
                                a_text = str(a_cell).strip()
                                # 너무 긴 텍스트는 잘라냄
                                if len(a_text) > 30:
                                    a_text = a_text[:27] + "..."
                                row_label = a_text
                            else:
                                # B열도 확인
                                b_cell = ws.cell(row_int, 2).value
                                if b_cell and str(b_cell).strip():
                                    b_text = str(b_cell).strip()
                                    if len(b_text) > 30:
                                        b_text = b_text[:27] + "..."
                                    row_label = b_text
                    except Exception:
                        pass
                
                # 메시지에서 과목명 추출 (파일을 읽을 수 없는 경우 대비)
                if not row_label:
                    import re
                    for it in row_items:
                        msg = it.get("message", "")
                        matches = re.findall(r"'([^']+)'", msg)
                        if matches:
                            # 첫 번째 작은따옴표 안의 텍스트가 과목명일 가능성이 높음
                            potential_name = matches[0]
                            if len(potential_name) < 30:
                                row_label = potential_name
                                break
                
                # 행 헤더
                if row_num == "-":
                    # '기타' 섹션은 특별 처리: '2026 전학년' 시트 관련 오류를 시트별로 그룹핑
                    import re
                    missing_course_pattern = r"'([^']+)'\s*시트.*?'([^']+)'\s*과목이\s*'2026\s*전학년'\s*시트에\s*없습니다"
                    missing_with_row_pattern = r"'([^']+)'\s*시트\s*(\d+)행의\s*'([^']+)'\s*과목이\s*'2026\s*전학년'\s*시트에\s*없습니다"
                    # 학생 선택 과목 패턴: '{시트명}' 시트의 '{과목명}' 과목({행번호}행[,...])이 '2026 전학년' 시트의 {열명}에 없습니다.
                    student_missing_pattern = r"'([^']+)'\s*시트의\s*'([^']+)'\s*과목\((\d+)행[^)]*\)이\s*'2026\s*전학년'\s*시트의\s*[^에]+에\s*없습니다"
                    
                    # 시트별로 그룹핑 (학교 지정/학생 선택 구분)
                    sheet_groups = {}
                    other_items = []
                    
                    for it in row_items:
                        msg = it.get("message", "")
                        
                        # 학생 선택 과목 패턴 (학교 지정과 학생 선택 사이 영역 또는 학생 선택 영역)
                        match = re.search(student_missing_pattern, msg)
                        if match:
                            source_sheet = match.group(1)
                            course = match.group(2)
                            row_no = match.group(3)
                            if source_sheet not in sheet_groups:
                                sheet_groups[source_sheet] = {
                                    "school": {"with_row": [], "without_row": []},
                                    "student": {"with_row": [], "without_row": []}
                                }
                            sheet_groups[source_sheet]["student"]["with_row"].append((course, row_no, it))
                            continue
                        
                        # 행 번호 있는 패턴 (학교 지정 과목)
                        match = re.search(missing_with_row_pattern, msg)
                        if match:
                            source_sheet = match.group(1)
                            row_no = match.group(2)
                            course = match.group(3)
                            if source_sheet not in sheet_groups:
                                sheet_groups[source_sheet] = {
                                    "school": {"with_row": [], "without_row": []},
                                    "student": {"with_row": [], "without_row": []}
                                }
                            sheet_groups[source_sheet]["school"]["with_row"].append((course, row_no, it))
                            continue
                        
                        # 행 번호 없는 패턴 (학교 지정 과목)
                        match = re.search(missing_course_pattern, msg)
                        if match:
                            source_sheet = match.group(1)
                            course = match.group(2)
                            if source_sheet not in sheet_groups:
                                sheet_groups[source_sheet] = {
                                    "school": {"with_row": [], "without_row": []},
                                    "student": {"with_row": [], "without_row": []}
                                }
                            sheet_groups[source_sheet]["school"]["without_row"].append((course, it))
                            continue
                        
                        # 패턴에 맞지 않는 기타 오류
                        other_items.append(it)
                    
                    # 시트별로 출력
                    for source_sheet in sorted(sheet_groups.keys()):
                        data = sheet_groups[source_sheet]
                        
                        has_school = data["school"]["with_row"] or data["school"]["without_row"]
                        has_student = data["student"]["with_row"] or data["student"]["without_row"]
                        
                        # 학교 지정 과목 또는 학생 선택 과목이 있으면 제목 출력 (한 번만)
                        if has_school or has_student:
                            self._w(tab, f"\n▶ '{source_sheet}'에 있지만, '2026 전학년' 시트에 없는 과목\n", "COURSE")
                            self._w(tab, "─" * 80 + "\n", "INFO")
                        
                        # 학교 지정 과목 출력
                        if has_school:
                            # 행 번호 있는 것들
                            for course, row_no, it in data["school"]["with_row"]:
                                sev = it.get("severity", "INFO")
                                self._w(tab, f"  [{sev}] {course} ({row_no}행)\n", 
                                       sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                            
                            # 행 번호 없는 것들
                            for course, it in data["school"]["without_row"]:
                                sev = it.get("severity", "INFO")
                                self._w(tab, f"  [{sev}] {course}\n", 
                                       sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                        
                        # 학생 선택 과목 출력
                        if has_student:
                            # 행 번호 있는 것들
                            for course, row_no, it in data["student"]["with_row"]:
                                sev = it.get("severity", "INFO")
                                self._w(tab, f"  [{sev}] {course} ({row_no}행)\n", 
                                       sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                            
                            # 행 번호 없는 것들
                            for course, it in data["student"]["without_row"]:
                                sev = it.get("severity", "INFO")
                                self._w(tab, f"  [{sev}] {course}\n", 
                                       sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                            
                            # 학생 선택 과목 목록 마지막에 안내 문구 한 번만 추가
                            self._w(tab, "      선택 미달 등으로 개설되지 않은 경우도 2026 전학년 시트에 추가하고 개설여부에 X해주세요.\n", "INFO")
                    
                    # 기타 오류들
                    if other_items:
                        self._w(tab, f"\n▶ 기타\n", "COURSE")
                        self._w(tab, "─" * 80 + "\n", "INFO")
                        for it in other_items:
                            sev = it.get("severity", "INFO")
                            msg = it.get("message", "")
                            
                            # 메시지에 줄바꿈이 있으면 첫 줄만 severity 표시, 나머지는 들여쓰기
                            lines = msg.split('\n')
                            if len(lines) > 1:
                                self._w(tab, f"  [{sev}] {lines[0]}\n", 
                                       sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                                for line in lines[1:]:
                                    if line.strip():  # 빈 줄이 아닌 경우만 출력
                                        self._w(tab, f"      {line}\n", 
                                               sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                            else:
                                self._w(tab, f"  [{sev}] {msg}\n", 
                                       sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                else:
                    if row_label:
                        self._w(tab, f"\n▶ {row_num}행 - {row_label}\n", "COURSE")
                    else:
                        self._w(tab, f"\n▶ {row_num}행\n", "COURSE")
                    self._w(tab, "─" * 80 + "\n", "INFO")
                    
                    for it in row_items:
                        sev = it.get("severity", "INFO")
                        msg = it.get("message", "")
                        
                        # 메시지에 줄바꿈이 있으면 첫 줄만 severity 표시, 나머지는 들여쓰기
                        lines = msg.split('\n')
                        if len(lines) > 1:
                            self._w(tab, f"  [{sev}] {lines[0]}\n", 
                                   sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                            for line in lines[1:]:
                                if line.strip():  # 빈 줄이 아닌 경우만 출력
                                    self._w(tab, f"      {line}\n", 
                                           sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")
                        else:
                            self._w(tab, f"  [{sev}] {msg}\n", 
                                   sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")

            err_cnt = sum(1 for x in items if x.get("severity") == "ERROR")
            warn_cnt = sum(1 for x in items if x.get("severity") == "WARNING")
            check_cnt = sum(1 for x in items if x.get("severity") == "CHECK")
            self._w(tab, "\n" + "=" * 80 + "\n", "INFO")
            self._w(tab, f"[전체 요약] 오류 {err_cnt}건, 경고 {warn_cnt}건, 확인 {check_cnt}건\n", "HEADER")

        # 오류가 없는 시트에 메시지 출력
        for tab_name in self.text_widgets.keys():
            if tab_name == "전체":
                continue
            # 해당 시트에 이슈가 없으면 메시지 출력
            if tab_name not in groups:
                self._w(tab_name, "발견된 오류가 없습니다.\n", "HEADER")
        
        # 전체 탭에도 전체 지침 간단 요약(원하면 제거 가능)
        self._w("전체", "[전체 문제 요약(시트별)]\n", "HEADER")
        for sheet, items in sorted(groups.items(), key=lambda kv: kv[0]):
            err_cnt = sum(1 for x in items if x.get("severity") == "ERROR")
            warn_cnt = sum(1 for x in items if x.get("severity") == "WARNING")
            check_cnt = sum(1 for x in items if x.get("severity") == "CHECK")
            label = sheet if sheet != "-" else "기타"
            self._w("전체", f"- {label}: 오류 {err_cnt} / 경고 {warn_cnt} / 확인 {check_cnt}\n", "INFO")
        
        # 제작자 정보 (오른쪽 정렬)
        self._w("전체", "\n제작: 창원여고 과정부\n", "RIGHT_ALIGN")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
