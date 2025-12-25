import os
import threading
import traceback
from copy import copy
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side


# =========================
# Excel Helpers
# =========================
def norm_text(v) -> str:
    """공백/개행/탭 제거 후 비교용 문자열로 정규화"""
    if v is None:
        return ""
    s = str(v)
    return "".join(s.split())


def argb_from_rgb(r: int, g: int, b: int) -> str:
    """openpyxl PatternFill용 ARGB(FFRRGGBB)"""
    return f"FF{r:02X}{g:02X}{b:02X}"


def solid_fill_rgb(r: int, g: int, b: int) -> PatternFill:
    c = argb_from_rgb(r, g, b)
    return PatternFill(fill_type="solid", start_color=c, end_color=c)


def cell_has_any_border(cell) -> bool:
    b = cell.border
    return any(
        getattr(b, side).style is not None
        for side in ("left", "right", "top", "bottom")
    )


def find_last_bordered_row(ws, start_row: int, last_col: int, max_scan: int = 2000) -> int:
    """
    표의 '아래 끝'을: start_row부터 아래로 내려가며
    A~last_col 범위에서 '어떤 셀이든 테두리가 존재하는' 마지막 행으로 판단.
    """
    max_r = min(ws.max_row, max_scan)
    last = start_row
    for r in range(start_row, max_r + 1):
        has_border = False
        for c in range(1, last_col + 1):
            if cell_has_any_border(ws.cell(r, c)):
                has_border = True
                break
        if has_border:
            last = r
    return last


def copy_border_with(orig: Border, **changes) -> Border:
    """
    Border는 불변 객체이므로, 기존 border를 기반으로 일부 side만 변경.
    """
    data = {
        "left": orig.left,
        "right": orig.right,
        "top": orig.top,
        "bottom": orig.bottom,
        "diagonal": orig.diagonal,
        "diagonal_direction": orig.diagonal_direction,
        "outline": orig.outline,
        "vertical": orig.vertical,
        "horizontal": orig.horizontal,
    }
    data.update(changes)
    return Border(**data)


def set_cell_alignment_center(cell):
    """
    기존 alignment를 최대한 유지하면서 가로/세로 가운데 + 자동줄바꿈만 강제.
    (세로 텍스트/회전 등의 속성을 가능하면 보존)
    """
    a = cell.alignment
    na = copy(a)
    na.horizontal = "center"
    na.vertical = "center"
    na.wrap_text = True
    cell.alignment = na


def apply_thin_all_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin")
    thin_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = thin_all


def apply_outer_medium_border(ws, min_row, max_row, min_col, max_col):
    medium = Side(style="medium")

    # Top edge
    for c in range(min_col, max_col + 1):
        cell = ws.cell(min_row, c)
        cell.border = copy_border_with(cell.border, top=medium)

    # Bottom edge
    for c in range(min_col, max_col + 1):
        cell = ws.cell(max_row, c)
        cell.border = copy_border_with(cell.border, bottom=medium)

    # Left edge
    for r in range(min_row, max_row + 1):
        cell = ws.cell(r, min_col)
        cell.border = copy_border_with(cell.border, left=medium)

    # Right edge
    for r in range(min_row, max_row + 1):
        cell = ws.cell(r, max_col)
        cell.border = copy_border_with(cell.border, right=medium)


def apply_row_outer_medium_border(ws, row, min_col, max_col):
    """특정 행(가로 띠)의 바깥 테두리를 medium으로"""
    medium = Side(style="medium")

    # top & bottom for whole row segment
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row, c)
        cell.border = copy_border_with(cell.border, top=medium, bottom=medium)

    # left edge at first col
    left_cell = ws.cell(row, min_col)
    left_cell.border = copy_border_with(left_cell.border, left=medium)

    # right edge at last col
    right_cell = ws.cell(row, max_col)
    right_cell.border = copy_border_with(right_cell.border, right=medium)


def apply_row_bottom_medium(ws, row, min_col, max_col):
    """특정 행의 '아래쪽' 테두리를 medium으로"""
    medium = Side(style="medium")
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row, c)
        cell.border = copy_border_with(cell.border, bottom=medium)


def fill_entire_row(ws, row, min_col, max_col, fill: PatternFill):
    for c in range(min_col, max_col + 1):
        ws.cell(row, c).fill = fill


def find_merged_range_for_value_in_col(ws, col: int, target_norm: str):
    """
    지정 열(col)에서, 병합된 영역의 좌상단 값이 target_norm과 일치하는 병합 범위를 찾음.
    """
    for mr in ws.merged_cells.ranges:
        if mr.min_col == col:
            v = ws.cell(mr.min_row, mr.min_col).value
            if norm_text(v) == target_norm:
                return mr
    return None


def remove_inner_borders_for_selection_groups(ws, start_row, end_row, search_col, check_cols, log):
    """
    특정 열(search_col)에서 '선택군'으로 시작하는 병합 셀을 찾고,
    병합된 첫 행의 check_cols 범위에 숫자가 있으면 
    그 열의 병합 영역 내부 수평 테두리를 제거
    """
    no_border = Side(style=None)
    found_count = 0
    
    for mr in ws.merged_cells.ranges:
        # 해당 열의 병합 셀인지 확인 (min_col과 max_col이 모두 search_col인 경우)
        if mr.min_col == search_col and mr.max_col == search_col:
            if mr.min_row >= start_row and mr.min_row <= end_row:
                # "선택군"으로 시작하는지 확인
                cell_value = ws.cell(mr.min_row, mr.min_col).value
                if cell_value and norm_text(str(cell_value)).startswith("선택군"):
                    found_count += 1
                    log(f"  - '선택군' 발견: {chr(64+search_col)}{mr.min_row}:{chr(64+search_col)}{mr.max_row}")
                    
                    # 병합된 첫 행의 check_cols 범위에서 숫자가 있는 열 찾기
                    first_row = mr.min_row
                    cols_with_numbers = []
                    
                    for col_idx in check_cols:
                        cell_val = ws.cell(first_row, col_idx).value
                        # 숫자인지 확인 (int, float 또는 숫자로 변환 가능한 문자열)
                        if cell_val is not None:
                            try:
                                float(str(cell_val))
                                cols_with_numbers.append(col_idx)
                            except (ValueError, TypeError):
                                pass
                    
                    if cols_with_numbers:
                        log(f"    숫자가 있는 열: {', '.join([chr(64+c) for c in cols_with_numbers])}")
                        
                        # 해당 열들의 병합 영역 내부 테두리 제거 (마지막 행 제외)
                        for col_idx in cols_with_numbers:
                            for row_idx in range(mr.min_row, mr.max_row):  # max_row는 포함하지 않음
                                cell = ws.cell(row_idx, col_idx)
                                cell.border = copy_border_with(cell.border, bottom=no_border)
                        
                        log(f"    병합 영역 내부 테두리 제거 완료")
    
    if found_count > 0:
        log(f"  - 총 {found_count}개의 '선택군' 병합 셀 처리 완료")
    
    return found_count


# =========================
# Core Processing
# =========================
def get_target_sheets(wb):
    """
    '로 시작' 조건 대신 실제 현장에서 흔한 명명(예: 2026학년도 입학생...)까지 고려하여
    연도/키워드 포함 여부로 매칭. (예시는 '예시' 포함 시 제외)
    """
    sheetnames = wb.sheetnames

    def pick(year: str, keyword: str, exclude_example=True):
        out = []
        for n in sheetnames:
            if exclude_example and ("예시" in n):
                continue
            if (year in n) and (keyword in n):
                out.append(n)
        return out

    # 요구 그룹
    g_2026_all = pick("2026", "전학년")  # 예시 제외
    g_2026_ent = pick("2026", "입학생")
    g_2025_ent = pick("2025", "입학생")
    g_2024_ent = pick("2024", "입학생")

    return {
        "2026 전학년": g_2026_all,
        "2026 입학생 ~": g_2026_ent,
        "2025 입학생 ~": g_2025_ent,
        "2024 입학생 ~": g_2024_ent,
    }


def process_one_sheet(ws, last_col: int, sheet_name: str, log):
    """
    - 표는 3행부터 시작한다고 가정
    - 표 아래 끝은 'A~last_col 영역에서 테두리가 존재하는 마지막 행'으로 판단
    """
    start_row = 3
    end_row = find_last_bordered_row(ws, start_row=start_row, last_col=last_col)

    min_col = 1
    max_col = last_col

    log(f"  - 표 범위 추정: A{start_row}:{chr(64+max_col)}{end_row}")

    # 1) 가운데 맞춤
    for r in range(start_row, end_row + 1):
        for c in range(min_col, max_col + 1):
            set_cell_alignment_center(ws.cell(r, c))

    # 2) '학교지정과목'(개행 포함 가능) 병합셀 종료 다음 행이 '학교 지정 과목 교과~'면 노란색
    yellow_row = None
    mr = find_merged_range_for_value_in_col(ws, col=1, target_norm="학교지정과목")
    if mr:
        candidate_row = mr.max_row + 1
        a_val_norm = norm_text(ws.cell(candidate_row, 1).value)
        if a_val_norm.startswith("학교지정과목교과"):
            yellow_row = candidate_row
            fill_entire_row(ws, yellow_row, min_col, max_col, solid_fill_rgb(255, 255, 0))
            log(f"  - 노란색 행 적용(학교 지정 과목 교과~): {yellow_row}행")
        else:
            log("  - '학교 지정 과목 교과~' 행을 찾지 못함(조건 미충족)")
    else:
        log("  - '학교지정과목' 병합셀을 찾지 못함")

    # 3) A열 표의 가장 아래 셀이 '총계'로 끝나면 색 적용
    bottom_a = ws.cell(end_row, 1).value
    if norm_text(bottom_a).endswith("총계"):
        fill_entire_row(ws, end_row, min_col, max_col, solid_fill_rgb(146, 208, 80))      # 총계 행
        if end_row - 1 >= start_row:
            fill_entire_row(ws, end_row - 1, min_col, max_col, solid_fill_rgb(255, 192, 0))
        if end_row - 2 >= start_row:
            fill_entire_row(ws, end_row - 2, min_col, max_col, solid_fill_rgb(255, 192, 0))
        if end_row - 3 >= start_row:
            fill_entire_row(ws, end_row - 3, min_col, max_col, solid_fill_rgb(255, 255, 0))
        log(f"  - 총계 하이라이트 적용: {end_row}행 기준")
    else:
        log("  - 표 하단 A열이 '총계'로 끝나지 않아(조건 미충족) 총계 색상 규칙 미적용")

    # 4) 3행부터 표 전체: 모든 테두리 얇게 + 바깥쪽 medium
    apply_thin_all_borders(ws, start_row, end_row, min_col, max_col)
    apply_outer_medium_border(ws, start_row, end_row, min_col, max_col)
    log("  - 전체 테두리: 내부 thin / 외곽 medium 적용")

    # 5) 4행 아래부분(= 4행의 하단선) medium
    if 4 <= end_row:
        apply_row_bottom_medium(ws, row=4, min_col=min_col, max_col=max_col)
        log("  - 4행 하단 테두리 medium 적용")

    # 6) 노란색 행(학교 지정 과목 교과~)의 바깥 테두리 medium
    if yellow_row:
        apply_row_outer_medium_border(ws, row=yellow_row, min_col=min_col, max_col=max_col)
        log(f"  - 노란색 행 바깥 테두리 medium 적용: {yellow_row}행")
    
    # 7) 선택군 병합 셀의 내부 테두리 제거
    if ("2024" in sheet_name) and ("입학생" in sheet_name):
        # 2024 입학생: B열에서 선택군 찾고, H~M열(8~13) 체크
        remove_inner_borders_for_selection_groups(
            ws, start_row, end_row, 
            search_col=2,  # B열
            check_cols=range(8, 14),  # H~M열
            log=log
        )
    else:
        # 2026 전학년, 2026 입학생, 2025 입학생: A열에서 선택군 찾고, G~L열(7~12) 체크
        remove_inner_borders_for_selection_groups(
            ws, start_row, end_row,
            search_col=1,  # A열
            check_cols=range(7, 13),  # G~L열
            log=log
        )


def adjust_workbook(input_path: str, output_path: str, log):
    ext = os.path.splitext(input_path)[1].lower()
    keep_vba = (ext == ".xlsm")

    wb = load_workbook(input_path, keep_vba=keep_vba)

    groups = get_target_sheets(wb)

    # 1) 시트 존재 여부 체크(예시 제외)
    missing = []
    for label, names in groups.items():
        if not names:
            missing.append(label)

    if missing:
        # 요구 문구 형태로 출력하고 종료
        for m in missing:
            log(f"{m} 으로 시작하는 시트가 없습니다")
        raise ValueError("필수 시트 누락")

    log("필수 시트 확인 완료. 서식 수정 진행...")

    # 2) 서식 수정
    processed = 0
    for label, names in groups.items():
        for sn in names:
            ws = wb[sn]
            # 열 범위 규칙
            if ("2024" in sn) and ("입학생" in sn):
                last_col = 16  # A~P
            else:
                last_col = 15  # A~O

            log(f"[시트] {sn} (범위 A~{chr(64+last_col)})")
            process_one_sheet(ws, last_col=last_col, sheet_name=sn, log=log)
            processed += 1

    wb.save(output_path)
    log(f"완료: {processed}개 시트 처리 후 저장됨 -> {output_path}")


# =========================
# Tkinter GUI
# =========================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("교육과정 편성표 양식 조정 프로그램")
        self.root.geometry("980x680")
        self.root.minsize(920, 620)

        self.colors = {
            "bg": "#F6F2FF",        # 연보라
            "panel": "#FFFFFF",
            "accent": "#BFA8FF",    # 파스텔 퍼플
            "accent2": "#AEE6D8",   # 파스텔 민트
            "text": "#2D2D2D",
            "muted": "#6B6B6B",
        }

        self.root.configure(bg=self.colors["bg"])

        self._build_style()
        self._build_ui()

        self.file_path = None
        self.running = False

    def _build_style(self):
        style = ttk.Style()
        # OS 기본 테마 유지 + 색만 조정
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("TFrame", background=self.colors["bg"])
        style.configure("Panel.TFrame", background=self.colors["panel"])
        style.configure("TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=("맑은 고딕", 10))
        style.configure("Title.TLabel", font=("맑은 고딕", 16, "bold"), foreground=self.colors["text"])
        style.configure("Sub.TLabel", font=("맑은 고딕", 10), foreground=self.colors["muted"])

        style.configure("Accent.TButton", font=("맑은 고딕", 10, "bold"))
        style.map("Accent.TButton",
                  background=[("active", self.colors["accent"])])
        style.configure("TEntry", font=("맑은 고딕", 10))

    def _build_ui(self):
        outer = ttk.Frame(self.root)
        outer.pack(fill="both", expand=True, padx=18, pady=18)

        header = ttk.Frame(outer)
        header.pack(fill="x", pady=(0, 12))

        ttk.Label(header, text="교육과정 편성표 양식 조정", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header, text="파일 업로드 → 시트 점검 → 표 서식 자동 수정 → 새 파일로 저장", style="Sub.TLabel").pack(anchor="w", pady=(4, 0))

        panel = ttk.Frame(outer, style="Panel.TFrame")
        panel.pack(fill="x", pady=(0, 12), ipadx=12, ipady=12)

        # 파일 선택
        row1 = ttk.Frame(panel, style="Panel.TFrame")
        row1.pack(fill="x", padx=12, pady=(10, 6))

        ttk.Label(row1, text="교육과정 엑셀 파일").pack(side="left")
        self.path_var = tk.StringVar(value="선택된 파일 없음")
        self.path_entry = ttk.Entry(row1, textvariable=self.path_var, state="readonly")
        self.path_entry.pack(side="left", fill="x", expand=True, padx=10)

        ttk.Button(row1, text="파일 선택", style="Accent.TButton", command=self.pick_file).pack(side="left")

        # 실행 버튼
        row2 = ttk.Frame(panel, style="Panel.TFrame")
        row2.pack(fill="x", padx=12, pady=(6, 10))

        self.run_btn = ttk.Button(row2, text="양식 조정 실행", style="Accent.TButton", command=self.run)
        self.run_btn.pack(side="left")

        self.status_var = tk.StringVar(value="대기 중")
        ttk.Label(row2, textvariable=self.status_var, style="Sub.TLabel").pack(side="left", padx=12)

        # 로그(문제상황)
        log_frame = ttk.Frame(outer, style="Panel.TFrame")
        log_frame.pack(fill="both", expand=True, ipadx=12, ipady=12)

        ttk.Label(log_frame, text="문제상황 / 처리 로그", style="TLabel").pack(anchor="w", padx=12, pady=(10, 6))

        self.log_text = tk.Text(log_frame, height=20, wrap="word", font=("맑은 고딕", 10))
        self.log_text.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        scroll = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        scroll.place(relx=1.0, rely=0.0, relheight=1.0, anchor="ne")

        self._log("프로그램이 준비되었습니다.\n- '파일 선택' 후 '양식 조정 실행'을 누르세요.\n")

    def _log(self, msg: str):
        def _append():
            self.log_text.insert("end", msg + ("\n" if not msg.endswith("\n") else ""))
            self.log_text.see("end")
        self.root.after(0, _append)

    def pick_file(self):
        fp = filedialog.askopenfilename(
            title="교육과정 편성표 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")]
        )
        if not fp:
            return
        self.file_path = fp
        self.path_var.set(fp)
        self._log(f"[선택됨] {fp}")
        self.status_var.set("파일 선택 완료")

    def run(self):
        if self.running:
            return
        if not self.file_path:
            messagebox.showwarning("안내", "먼저 엑셀 파일을 선택하세요.")
            return

        # 저장 경로 선택
        base, ext = os.path.splitext(self.file_path)
        default_name = os.path.basename(base) + "_양식조정" + ext
        output_path = filedialog.asksaveasfilename(
            title="저장할 파일 이름 지정",
            defaultextension=ext,
            initialfile=default_name,
            filetypes=[("Excel 파일", f"*{ext}"), ("모든 파일", "*.*")]
        )
        if not output_path:
            return

        self.running = True
        self.run_btn.configure(state="disabled")
        self.status_var.set("처리 중...")

        self._log("\n========================================")
        self._log("실행 시작")
        self._log("========================================")

        t = threading.Thread(target=self._worker, args=(self.file_path, output_path), daemon=True)
        t.start()

    def _worker(self, input_path, output_path):
        try:
            adjust_workbook(input_path, output_path, log=self._log)
            self.root.after(0, lambda: messagebox.showinfo("완료", "양식 조정이 완료되었습니다."))
            self.status_var.set("완료")
        except ValueError as ve:
            # 필수 시트 누락 등: 요구사항대로 메시지 출력 후 종료
            self._log(f"[중단] {ve}")
            self.status_var.set("중단(필수 조건 미충족)")
            self.root.after(0, lambda: messagebox.showwarning("중단", "필수 시트가 없어 작업을 중단했습니다.\n로그를 확인하세요."))
        except Exception:
            self._log("[오류] 예기치 못한 오류가 발생했습니다.")
            self._log(traceback.format_exc())
            self.status_var.set("오류")
            self.root.after(0, lambda: messagebox.showerror("오류", "오류가 발생했습니다.\n로그를 확인하세요."))
        finally:
            def _end():
                self.running = False
                self.run_btn.configure(state="normal")
            self.root.after(0, _end)


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
