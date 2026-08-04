# -*- coding: utf-8 -*-
"""Inspect A-column merges and labels on the test workbook."""
import importlib.util
from pathlib import Path
from openpyxl import load_workbook

PROG = Path(r"C:\Users\jack1\Documents\GitHub\cursor\교육과정 편성표 점검 프로그램.py")
XLSX = Path(
    r"C:\Users\jack1\OneDrive - 창원여자고등학교\강의\2026.08.14. 교육과정 연수"
    r"\2027학년도 학교교육과정 편성표_테스트용1.xlsx"
)
spec = importlib.util.spec_from_file_location("cc", PROG)
cc = importlib.util.module_from_spec(spec)
spec.loader.exec_module(cc)

wb_v = load_workbook(XLSX, data_only=True)
wb_f = load_workbook(XLSX, data_only=False)

for year in (2027, 2026, 2025):
    sname = cc.find_sheet_for_year(wb_v.sheetnames, year)
    ws_v, ws_f = wb_v[sname], wb_f[sname]
    merge = cc.build_merged_lookup(ws_f)
    print("=" * 60, sname)
    seen = set()
    for rng in ws_f.merged_cells.ranges:
        if rng.min_col != 1 or rng.max_col != 1:
            continue
        if rng.max_row - rng.min_row < 1:
            continue
        key = (rng.min_row, rng.max_row)
        if key in seen:
            continue
        seen.add(key)
        a_val, _, _ = cc.get_value_with_merge(ws_v, ws_f, merge, rng.min_row, 1)
        # semester sum on top row
        top_sum = 0.0
        for c in range(7, 13):
            n = cc.to_number(cc.get_value_with_merge(ws_v, ws_f, merge, rng.min_row, c)[0])
            if n is not None:
                top_sum += n
        courses = []
        for r in range(rng.min_row, min(rng.max_row, rng.min_row + 4) + 1):
            d, _, _ = cc.get_value_with_merge(ws_v, ws_f, merge, r, 4)
            if d:
                courses.append(str(d)[:20])
        print(
            f"A-merge {rng.min_row}-{rng.max_row} rows={rng.max_row-rng.min_row+1} "
            f"A={a_val!r} top_sem_sum={top_sum:g} courses~{courses}"
        )
