# -*- coding: utf-8 -*-
"""
교육과정 편성표 확인 프로그램 (Tkinter + openpyxl)
- 단일 파일(모듈 분리 없음)
- .xlsx / .xlsm 지원 (openpyxl)

요구사항 반영(버전 v3)
1) 운영학점(F) vs G~L 합:
   - G~L 합이 0이면 '바로 위 행의 운영학점(F)'과 같으면 통과
2) 2024 입학생 시트:
   - 과목명(D) "숨김 시트와의 일치 여부"는 확인하지 않음(불일치 오류 미출력)
3) 2025/2026 입학생 시트:
   - 과목명(D) 셀에 채우기 색(흰색 제외)이 있으면 그 행은 "모든 검사" 제외
4) 과목명에 ↔ 가 있으면:
   - 좌/우 과목명을 각각 숨김 시트에 존재하는지 확인(없으면 오류)
   - ↔ 행은 숨김 기반(유형/기본학점/성적처리/범위) 비교는 생략
   - 단, 내부 일관성(운영학점 vs G~L 합, M/N 합계 계산)은 계속 점검(색깔 행은 전부 제외)

사용 방법
1) pip install openpyxl
2) python curriculum_checker_v3.py
"""

import os
import re
import difflib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

from openpyxl import load_workbook


# =========================
# 유틸
# =========================

EPS = 1e-9


def normalize_course_name(name: str) -> str:
    """괄호( ) 안 내용을 제거하고, 양끝 공백만 제거(내부 공백은 유지)."""
    if name is None:
        return ""
    s = str(name)
    s = re.sub(r"\([^)]*\)", "", s)  # ( ... ) 제거
    return s.strip()


def split_bidirectional(name: str):
    """'음악↔미술' 같은 문자열을 ['음악','미술']로 분해(양쪽 공백 제거)."""
    if name is None:
        return []
    s = str(name)
    if "↔" not in s:
        return []
    parts = [p.strip() for p in s.split("↔")]
    parts = [p for p in parts if p != ""]
    return parts


def is_error_token(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().upper()
    return s in ("#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def to_number(value):
    """숫자 변환(정수/실수). 실패 시 None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # NaN 체크
        if isinstance(value, float) and (value != value):
            return None
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def find_sheet_for_year(sheetnames, year: int):
    """
    '2026 입학생...' 또는 '2026학년도 입학생...' 등:
    - 시트명이 year로 시작하고 '입학생'을 포함하면 매칭
    """
    y = str(year)
    candidates = [n for n in sheetnames if n.startswith(y) and ("입학생" in n)]
    if candidates:
        return candidates[0]
    # 보조: 공백 제거 후 '2026입학생' 시작
    for n in sheetnames:
        n2 = n.replace(" ", "")
        if n2.startswith(f"{y}입학생"):
            return n
    return None


def find_hidden_sheet_name(sheetnames):
    """'숨김' 시트를 찾음(정확 일치 우선, 포함은 차선)."""
    if "숨김" in sheetnames:
        return "숨김"
    for n in sheetnames:
        if "숨김" in n:
            return n
    return None


def build_merged_lookup(ws):
    """
    셀 좌표(행,열) -> (min_row, min_col, max_row, max_col) 매핑을 만든다.
    """
    lookup = {}
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = (min_row, min_col, max_row, max_col)
    return lookup


def get_value_with_merge(ws_values, ws_formula, merged_lookup, row, col):
    """
    data_only 값(ws_values) 기준으로:
    - 해당 셀이 병합 영역이면 top-left 값으로 보정
    - 동시에 수식(ws_formula)의 존재도 반환
    return: (value, formula_str_or_none, (used_row, used_col))
    """
    used_row, used_col = row, col
    key = (row, col)

    if key in merged_lookup:
        min_row, min_col, _, _ = merged_lookup[key]
        used_row, used_col = min_row, min_col

    v = ws_values.cell(used_row, used_col).value
    f = ws_formula.cell(used_row, used_col).value
    formula = f if isinstance(f, str) and f.startswith("=") else None
    return v, formula, (used_row, used_col)


def find_hidden_header_row(ws_values, ws_formula, merged_lookup):
    """
    숨김 시트에서 '과목명' 헤더가 있는 행을 찾음(기본 2행).
    """
    for r in range(1, 21):
        v, _, _ = get_value_with_merge(ws_values, ws_formula, merged_lookup, r, 2)  # B열
        if v is not None and str(v).strip() == "과목명":
            return r
    return 2


def safe_strip(v):
    if v is None:
        return ""
    return str(v).strip()


def is_colored_fill(cell) -> bool:
    """
    '색깔이 있는 경우' 판단:
    - 패턴이 있고(대개 solid), 흰색이 아닌 채우기면 True
    주의: 테마색/인덱스색은 정확한 RGB 환산이 어려워도 '색이 있다'로 간주(보수적으로 제외).
    """
    try:
        fill = cell.fill
    except Exception:
        return False

    if fill is None:
        return False

    # 기본(무채움)은 patternType None/None 또는 'none'일 가능성이 높음
    pt = getattr(fill, "patternType", None)
    if pt is None or str(pt).lower() in ("none", "null"):
        return False

    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return True  # 패턴이 있는데 색 정보가 없으면 일단 색 있는 것으로 처리

    ctype = getattr(fg, "type", None)
    val = getattr(fg, "value", None) or getattr(fg, "rgb", None)

    # theme/indexed는 정확 RGB가 없을 수 있으므로 색으로 간주(제외)
    if ctype in ("theme", "indexed"):
        return True

    if not val:
        return True

    s = str(val).upper()
    # ARGB(예: FF112233) 또는 RGB(112233) 형태가 들어올 수 있음
    # 흰색 계열이면 제외(= 색 없는 것으로 처리)
    # - 끝 6자리가 FFFFFF면 흰색으로 간주
    if len(s) >= 6 and s[-6:] == "FFFFFF":
        return False
    # openpyxl이 "00000000" 등을 줄 수도 있는데 이는 무채움/자동에 가까움
    if s in ("00000000", "000000", "FFFFFFFF", "00FFFFFF", "FFFFFF"):
        return False

    return True


# =========================
# 핵심 검사 로직
# =========================

def run_checks(xlsx_path: str):
    """
    요구사항 기반 검사 수행.
    return: (issues, summary)
      - issues: list[dict]  {severity, sheet, row, message}
      - summary: dict
    """
    if not os.path.exists(xlsx_path):
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": "파일을 찾을 수 없습니다."}], {})

    ext = os.path.splitext(xlsx_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": "지원하지 않는 확장자입니다. .xlsx 또는 .xlsm만 지원합니다."}], {})

    issues = []
    summary = {}

    try:
        wb_v = load_workbook(xlsx_path, data_only=True)
        wb_f = load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": f"엑셀 파일을 열 수 없습니다: {e}"}], {})

    sheetnames = wb_v.sheetnames

    # (1) 대상 시트 존재 확인
    targets = {}
    for y in (2026, 2025, 2024):
        sname = find_sheet_for_year(sheetnames, y)
        targets[y] = sname
        if not sname:
            issues.append({"severity": "ERROR", "sheet": "-", "row": "-", "message": f"{y} 입학생으로 시작하고 '입학생'을 포함하는 시트를 찾지 못했습니다."})

    # 숨김 시트 로드
    hidden_name = find_hidden_sheet_name(sheetnames)
    if not hidden_name:
        issues.append({"severity": "ERROR", "sheet": "-", "row": "-", "message": "숨김 시트를 찾지 못했습니다(시트명에 '숨김' 포함 필요)."})
        return issues, {"targets": targets, "hidden_sheet": None}

    ws_hidden_v = wb_v[hidden_name]
    ws_hidden_f = wb_f[hidden_name]
    hidden_merge = build_merged_lookup(ws_hidden_f)
    header_row = find_hidden_header_row(ws_hidden_v, ws_hidden_f, hidden_merge)
    data_start = header_row + 1

    # 숨김 과목 사전 구축
    # B:과목명, C:유형, D:기본학점, E:성적처리, F:최소, G:최대
    hidden = {}
    hidden_list_norm = []
    r = data_start
    while True:
        course_raw, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 2)  # B
        if course_raw is None or str(course_raw).strip() == "":
            break
        course_norm = normalize_course_name(course_raw)

        typ, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 3)  # C
        basic, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 4)  # D
        grade, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 5)  # E
        minc, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 6)  # F
        maxc, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 7)  # G

        rec = {
            "course_raw": safe_strip(course_raw),
            "type": safe_strip(typ),
            "basic": to_number(basic),
            "grading": safe_strip(grade),
            "min": to_number(minc),
            "max": to_number(maxc),
            "row": r,
        }

        if course_norm in hidden:
            issues.append({
                "severity": "WARNING",
                "sheet": hidden_name,
                "row": r,
                "message": f"숨김 시트에 중복 과목명이 있습니다: '{course_norm}' (기존 {hidden[course_norm]['row']}행, 추가 {r}행). 최초 항목을 기준으로 검사합니다."
            })
        else:
            hidden[course_norm] = rec
            hidden_list_norm.append(course_norm)

        r += 1

    summary["targets"] = targets
    summary["hidden_sheet"] = hidden_name
    summary["hidden_course_count"] = len(hidden)

    # 시트가 없다면 여기서 종료(숨김은 읽었으니 보고는 가능)
    if any(v is None for v in targets.values()):
        return issues, summary

    # =========================
    # (2) 각 시트 검사
    # =========================
    for year, sname in targets.items():
        ws_v = wb_v[sname]
        ws_f = wb_f[sname]
        merge_lookup = build_merged_lookup(ws_f)

        first_row = 5
        course_col = 4  # D
        type_col = 3    # C
        basic_col = 5   # E
        op_col = 6      # F
        sem_cols = list(range(7, 13))  # G~L
        total_cols = [13, 14]          # M, N
        grading_col = 15               # O

        # last row (D 기준)
        last_row = None
        for rr in range(ws_f.max_row, first_row - 1, -1):
            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if v is None or str(v).strip() == "":
                continue
            last_row = rr
            break
        if last_row is None:
            issues.append({"severity": "WARNING", "sheet": sname, "row": "-", "message": "D열(과목)에서 데이터 행을 찾지 못했습니다."})
            continue

        # 2025/2026: 색깔 행(과목명 D 셀 기준)은 모든 검사 제외
        exempt_rows = set()
        if year in (2025, 2026):
            for rr in range(first_row, last_row + 1):
                cell = ws_f.cell(rr, course_col)
                if is_colored_fill(cell):
                    exempt_rows.add(rr)

        # row_total(각 행의 G~L 합) 계산 (색깔 행은 계산에서도 제외)
        row_total = {}
        for rr in range(first_row, last_row + 1):
            if rr in exempt_rows:
                continue

            course_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_v is None or str(course_v).strip() == "":
                continue

            if is_error_token(course_v):
                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"과목명(D{rr})에 오류값이 있습니다: {course_v}"})
                continue

            sem_sum = 0.0
            any_num = False
            for cc in sem_cols:
                v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, cc)
                n = to_number(v)
                if n is not None:
                    sem_sum += n
                    any_num = True

            row_total[rr] = sem_sum if any_num else 0.0

        # ========== 과목 단위 검사 ==========
        for rr in range(first_row, last_row + 1):
            if rr in exempt_rows:
                # 색깔 행은 모든 검사 제외
                continue

            course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_raw is None or str(course_raw).strip() == "":
                continue
            if is_error_token(course_raw):
                continue

            course_norm = normalize_course_name(course_raw)
            if course_norm == "":
                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": "과목명(D열)에서 괄호 제거 후 이름이 비었습니다."})
                continue

            parts = split_bidirectional(course_norm)
            is_bidirectional = len(parts) >= 2

            # (2) 과목명 일치 여부:
            # - 2024는 검증하지 않음(단, 숨김 기반 검사는 '매칭될 때만' 수행)
            hidden_rec = None
            if year == 2024:
                hidden_rec = hidden.get(course_norm, None)  # 있으면 활용, 없으면 숨김 기반 검사는 생략
            else:
                # 2025/2026 (색깔 행 제외된 상태)
                if is_bidirectional:
                    missing = []
                    for p in parts:
                        if p not in hidden:
                            missing.append(p)
                    if missing:
                        # 유사 후보 힌트(각 missing마다 1개)
                        hints = []
                        for m in missing:
                            close = difflib.get_close_matches(m, hidden_list_norm, n=1, cutoff=0.6)
                            if close:
                                hints.append(f"{m}→{close[0]}")
                        hint_txt = f" (유사 후보: {', '.join(hints)})" if hints else ""
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"↔ 과목명 중 숨김 시트에 없는 항목: {', '.join(missing)}{hint_txt}"})
                    # ↔ 행은 숨김 기반(유형/기본학점/성적처리/범위) 비교 생략
                    hidden_rec = None
                else:
                    if course_norm not in hidden:
                        hint = ""
                        close = difflib.get_close_matches(course_norm, hidden_list_norm, n=2, cutoff=0.6)
                        if close:
                            hint = f" (유사 과목명 후보: {', '.join(close)})"
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"숨김 시트 과목명과 불일치: '{course_norm}'{hint}"})
                        hidden_rec = None
                    else:
                        hidden_rec = hidden[course_norm]

            # (3) 유형/기본학점/성적처리 (숨김 매칭이 있을 때만)
            if hidden_rec is not None:
                # 유형(C)
                typ_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, type_col)
                typ_s = safe_strip(typ_v)
                if typ_s == "":
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"유형(C{rr})이 비어 있습니다. (숨김: {hidden_rec['type']})"})
                elif typ_s != hidden_rec["type"]:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"유형 불일치: 시트='{typ_s}' / 숨김='{hidden_rec['type']}'"})

                # 기본학점(E)
                basic_v, basic_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, basic_col)
                basic_n = to_number(basic_v)
                if basic_n is None:
                    if basic_formula:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rr, "message": f"기본학점(E{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {basic_formula})"})
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"기본학점(E{rr})이 숫자가 아닙니다: {basic_v}"})
                else:
                    if hidden_rec["basic"] is not None and abs(basic_n - hidden_rec["basic"]) > EPS:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"기본학점 불일치: 시트={basic_n:g} / 숨김={hidden_rec['basic']:g}"})

                # 성적처리(O)
                grade_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, grading_col)
                grade_s = safe_strip(grade_v)
                if grade_s == "":
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"성적처리 유형(O{rr})이 비어 있습니다. (숨김: {hidden_rec['grading']})"})
                elif grade_s != hidden_rec["grading"]:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"성적처리 유형 불일치: 시트='{grade_s}' / 숨김='{hidden_rec['grading']}'"})

            # (4)(5) 운영학점 범위/합계 체크
            # - 범위(최소~최대)는 숨김 매칭이 있을 때만
            # - 합계(운영학점 vs G~L합)는 모든 행(색깔 행 제외)에 대해 수행
            op_v, op_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, op_col)
            op_n = to_number(op_v)
            sem_sum = row_total.get(rr, 0.0)

            if op_n is None:
                if op_formula:
                    issues.append({"severity": "WARNING", "sheet": sname, "row": rr, "message": f"운영학점(F{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {op_formula})"})
                else:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점(F{rr})이 숫자가 아닙니다: {op_v}"})
            else:
                # (4) 범위 체크(숨김 매칭이 있을 때만, ↔ 및 2024-미매칭은 생략)
                if hidden_rec is not None and (hidden_rec["min"] is not None) and (hidden_rec["max"] is not None):
                    if not (hidden_rec["min"] - EPS <= op_n <= hidden_rec["max"] + EPS):
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점 범위 위반: 시트={op_n:g} / 허용범위={hidden_rec['min']:g}~{hidden_rec['max']:g}"})

                # (5) 운영학점 vs G~L합
                if abs(sem_sum) <= EPS:
                    # 합이 0이면 바로 위 행 운영학점과 비교(같으면 OK)
                    prev = None
                    if rr > first_row:
                        pv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr - 1, op_col)
                        prev = to_number(pv)
                    if prev is not None and abs(op_n - prev) <= EPS:
                        pass  # OK
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"G~L 합이 0이므로 위 행 운영학점과 비교해야 합니다: 현재={op_n:g}, 위행={(prev if prev is not None else '없음')}"})
                else:
                    if abs(op_n - sem_sum) > EPS:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점(F)과 G~L 합 불일치: 운영학점={op_n:g}, G~L합={sem_sum:g}"})

        # (6) M/N 병합 구간 합계 체크 (색깔 행은 기대값에서도 제외)
        checked_spans = set()
        for rng in ws_f.merged_cells.ranges:
            if rng.min_col in total_cols and rng.max_col == rng.min_col:
                col = rng.min_col
                if rng.max_row < first_row:
                    continue

                start = max(rng.min_row, first_row)
                end = min(rng.max_row, last_row)
                if start > end:
                    continue

                key = (col, rng.min_row, rng.max_row)
                if key in checked_spans:
                    continue
                checked_spans.add(key)

                total_v, total_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rng.min_row, col)
                total_n = to_number(total_v)

                expected = 0.0
                for rr in range(start, end + 1):
                    if rr in exempt_rows:
                        continue
                    cv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
                    if cv is None or str(cv).strip() == "" or is_error_token(cv):
                        continue
                    expected += row_total.get(rr, 0.0)

                if total_n is None:
                    if total_formula:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rng.min_row, "message": f"{'M' if col==13 else 'N'}열 합계 셀에 수식은 있으나 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {total_formula})"})
                    else:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rng.min_row, "message": f"{'M' if col==13 else 'N'}열 합계 셀이 비어 있습니다. (해당 구간 G~L 합 기대값={expected:g})"})
                else:
                    if abs(total_n - expected) > EPS:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rng.min_row, "message": f"{'M' if col==13 else 'N'}열 병합구간 합계 불일치: 셀값={total_n:g}, 기대값(G~L합)={expected:g} (구간 {start}~{end}행, 색깔행 제외)"})

        # 병합이 아닌 단일 셀 합계 보조 체크(색깔행 제외)
        for col in total_cols:
            for rr in range(first_row, last_row + 1):
                if rr in exempt_rows:
                    continue

                if (rr, col) in merge_lookup:
                    min_r, min_c, _, _ = merge_lookup[(rr, col)]
                    if not (rr == min_r and col == min_c):
                        continue
                    continue

                tv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col)
                tn = to_number(tv)
                if tn is None:
                    continue

                cv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
                if cv is None or str(cv).strip() == "" or is_error_token(cv):
                    continue

                expected = row_total.get(rr, 0.0)
                if abs(tn - expected) > EPS:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"{'M' if col==13 else 'N'}열 단일행 합계 불일치: 셀값={tn:g}, 기대값(G~L합)={expected:g}"})

    return issues, summary


# =========================
# GUI
# =========================

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("교육과정 편성표 확인 프로그램")
        self.root.minsize(980, 650)

        self.colors = {
            "bg": "#F6F7FF",
            "card": "#FFFFFF",
            "text": "#1F2937",
            "muted": "#6B7280",
            "accent": "#7C6CF6",
            "danger": "#EF4444",
            "warn": "#F59E0B",
        }
        self.root.configure(bg=self.colors["bg"])

        self.style = ttk.Style(self.root)
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        base_font = ("Malgun Gothic", 11)
        title_font = ("Malgun Gothic", 16, "bold")

        self.style.configure("TFrame", background=self.colors["bg"])
        self.style.configure("Card.TFrame", background=self.colors["card"])
        self.style.configure("TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=base_font)
        self.style.configure("Title.TLabel", font=title_font, foreground=self.colors["text"])
        self.style.configure("Muted.TLabel", foreground=self.colors["muted"], font=("Malgun Gothic", 10))

        self._build_ui()
        self.xlsx_path = None

    def _build_ui(self):
        header = ttk.Frame(self.root, padding=(18, 18, 18, 10))
        header.pack(fill="x")

        ttk.Label(header, text="교육과정 편성표 확인 프로그램", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            header,
            text="엑셀(.xlsx/.xlsm) 업로드 후, 시트/과목/학점/합계를 자동 점검합니다.",
            style="Muted.TLabel"
        ).pack(anchor="w", pady=(6, 0))

        body = ttk.Frame(self.root, padding=(18, 8, 18, 18))
        body.pack(fill="both", expand=True)

        card = ttk.Frame(body, style="Card.TFrame", padding=(16, 16))
        card.pack(fill="x")

        row1 = ttk.Frame(card, style="Card.TFrame")
        row1.pack(fill="x")

        self.path_var = tk.StringVar(value="선택된 파일 없음")
        ttk.Label(row1, text="엑셀 파일:", style="Muted.TLabel").pack(side="left")
        ttk.Label(row1, textvariable=self.path_var).pack(side="left", padx=(8, 0))

        btn_frame = ttk.Frame(card, style="Card.TFrame")
        btn_frame.pack(fill="x", pady=(12, 0))

        self.btn_pick = tk.Button(
            btn_frame,
            text="파일 선택",
            command=self.pick_file,
            bg=self.colors["accent"],
            fg="white",
            bd=0,
            activebackground=self.colors["accent"],
            activeforeground="white",
            padx=16,
            pady=10,
            font=("Malgun Gothic", 11, "bold"),
            cursor="hand2"
        )
        self.btn_pick.pack(side="left")

        self.btn_run = tk.Button(
            btn_frame,
            text="검사 실행",
            command=self.run,
            bg="#C7C9D9",
            fg="white",
            bd=0,
            activebackground="#C7C9D9",
            activeforeground="white",
            padx=16,
            pady=10,
            font=("Malgun Gothic", 11, "bold"),
            cursor="hand2",
            state="disabled"
        )
        self.btn_run.pack(side="left", padx=(10, 0))

        status_frame = ttk.Frame(card, style="Card.TFrame")
        status_frame.pack(fill="x", pady=(12, 0))

        self.status_var = tk.StringVar(value="대기 중")
        ttk.Label(status_frame, textvariable=self.status_var, style="Muted.TLabel").pack(side="left")

        self.progress = ttk.Progressbar(status_frame, mode="indeterminate", length=220)
        self.progress.pack(side="right")

        out_card = ttk.Frame(body, style="Card.TFrame", padding=(16, 16))
        out_card.pack(fill="both", expand=True, pady=(14, 0))

        ttk.Label(out_card, text="문제상황", style="Muted.TLabel").pack(anchor="w")
        self.out = ScrolledText(
            out_card,
            wrap="word",
            height=18,
            font=("Consolas", 10),
            bg="#FBFBFE",
            fg=self.colors["text"],
            relief="solid",
            bd=1,
            padx=10,
            pady=10
        )
        self.out.pack(fill="both", expand=True, pady=(8, 0))

        self.out.tag_configure("ERROR", foreground=self.colors["danger"])
        self.out.tag_configure("WARNING", foreground=self.colors["warn"])
        self.out.tag_configure("INFO", foreground=self.colors["muted"])
        self.out.tag_configure("HEADER", font=("Malgun Gothic", 11, "bold"))

    def pick_file(self):
        path = filedialog.askopenfilename(
            title="교육과정 편성표 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return
        self.xlsx_path = path
        self.path_var.set(path)
        self.btn_run.configure(state="normal", bg=self.colors["accent"], activebackground=self.colors["accent"])

    def run(self):
        if not self.xlsx_path:
            messagebox.showwarning("안내", "먼저 엑셀 파일을 선택하세요.")
            return

        self.out.delete("1.0", "end")
        self.status_var.set("검사 중...")
        self.progress.start(12)
        self.root.update_idletasks()

        try:
            issues, summary = run_checks(self.xlsx_path)
        except Exception as e:
            self.progress.stop()
            self.status_var.set("오류 발생")
            messagebox.showerror("오류", f"검사 중 예외가 발생했습니다:\n{e}")
            return

        self.progress.stop()

        self._print_summary(summary)
        self._print_issues(issues)

        err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
        warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")

        if err_cnt == 0:
            self.status_var.set(f"검사 완료: 오류 없음 (경고 {warn_cnt}건)")
        else:
            self.status_var.set(f"검사 완료: 오류 {err_cnt}건 / 경고 {warn_cnt}건")

    def _print_summary(self, summary):
        self.out.insert("end", "[검사 개요]\n", "HEADER")
        self.out.insert("end", f"- 파일: {self.xlsx_path}\n", "INFO")

        targets = summary.get("targets") or {}
        self.out.insert("end", "- 시트 확인:\n", "INFO")
        for y in (2026, 2025, 2024):
            s = targets.get(y)
            if s:
                self.out.insert("end", f"  · {y}: {s}\n", "INFO")
            else:
                self.out.insert("end", f"  · {y}: (없음)\n", "WARNING")

        hidden = summary.get("hidden_sheet")
        cnt = summary.get("hidden_course_count", 0)
        if hidden:
            self.out.insert("end", f"- 숨김 시트: {hidden} (과목 {cnt}개)\n\n", "INFO")
        else:
            self.out.insert("end", "- 숨김 시트: (없음)\n\n", "ERROR")

    def _print_issues(self, issues):
        if not issues:
            self.out.insert("end", "문제 없음.\n", "INFO")
            return

        sev_rank = {"ERROR": 0, "WARNING": 1, "INFO": 2}

        def key(x):
            sheet = x.get("sheet", "")
            row = x.get("row", "-")
            try:
                row_n = int(row)
            except Exception:
                row_n = 10**9
            return (sev_rank.get(x.get("severity", "INFO"), 9), sheet, row_n)

        issues_sorted = sorted(issues, key=key)

        self.out.insert("end", "[문제 목록]\n", "HEADER")

        for it in issues_sorted:
            sev = it.get("severity", "INFO")
            sheet = it.get("sheet", "-")
            row = it.get("row", "-")
            msg = it.get("message", "")
            line = f"- [{sev}] {sheet} / 행 {row}: {msg}\n"
            self.out.insert("end", line, sev if sev in ("ERROR", "WARNING") else "INFO")

        err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
        warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")
        self.out.insert("end", "\n", "INFO")
        self.out.insert("end", f"[요약] 오류 {err_cnt}건, 경고 {warn_cnt}건\n", "HEADER")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
