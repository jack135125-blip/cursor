# -*- coding: utf-8 -*-
"""
교육과정 편성표 확인 프로그램 — Flask 웹앱
기존 curriculum_checker 검사 로직(run_checks)을 재사용합니다.
"""

from __future__ import annotations

import importlib.util
import os
import re
import sys
import tempfile
import uuid
from pathlib import Path

from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
CHECKER_PATH = BASE_DIR.parent / "curriculum_checker_12.26_테스트용 완성본.py"
UPLOAD_DIR = BASE_DIR / "uploads"
ALLOWED_EXT = {".xlsx", ".xlsm"}
TEMPLATE_DOWNLOAD_URL = (
    "https://drive.google.com/drive/folders/1wvdV4VQD7kUD7eVEvypPf39LDWZLxfze?usp=sharing"
)

UPLOAD_DIR.mkdir(exist_ok=True)


def _load_checker():
    if not CHECKER_PATH.exists():
        raise FileNotFoundError(f"검사 모듈을 찾을 수 없습니다: {CHECKER_PATH}")
    spec = importlib.util.spec_from_file_location("curriculum_checker", CHECKER_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules["curriculum_checker"] = module
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


checker = _load_checker()
run_checks = checker.run_checks
find_all_grades_sheet = checker.find_all_grades_sheet
normalize_course_name = checker.normalize_course_name

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024  # 40MB


# ---------------------------------------------------------------------------
# 결과 포맷 (데스크톱 App._print_* 와 동일 구조)
# ---------------------------------------------------------------------------

def _line(text: str, tag: str = "INFO") -> dict:
    return {"text": text, "tag": tag}


def _append_issue_lines(lines: list, sev: str, msg: str) -> None:
    tag = sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO"
    parts = msg.split("\n")
    if len(parts) > 1:
        lines.append(_line(f"  [{sev}] {parts[0]}", tag))
        for part in parts[1:]:
            if part.strip():
                lines.append(_line(f"      {part}", tag))
    else:
        lines.append(_line(f"  [{sev}] {msg}", tag))


def format_results(xlsx_path: str, filename: str, issues: list, summary: dict) -> dict:
    """시트별 탭 텍스트 라인을 구성해 프론트엔드로 전달."""
    if not isinstance(issues, list):
        issues = []
    if not isinstance(summary, dict):
        summary = {}

    targets = summary.get("targets") or {}
    tab_names = ["전체"]
    for y in (2026, 2025, 2024):
        s = targets.get(y)
        if s and s not in tab_names:
            tab_names.append(s)

    all_grades_sheet = None
    try:
        wb_temp = load_workbook(xlsx_path, data_only=True)
        all_grades_sheet = find_all_grades_sheet(wb_temp.sheetnames)
        if all_grades_sheet and all_grades_sheet not in tab_names:
            tab_names.append(all_grades_sheet)
    except Exception:
        wb_temp = None

    tab_names.append("기타")
    tabs: dict[str, list] = {name: [] for name in tab_names}

    # --- 전체 탭: 검사 개요 ---
    overview = tabs["전체"]
    overview.append(_line("[검사 개요]", "HEADER"))
    overview.append(_line(f"- 파일: {filename}", "INFO"))
    overview.append(_line("- 시트 확인:", "INFO"))
    for y in (2026, 2025, 2024):
        s = targets.get(y)
        if s:
            overview.append(_line(f"  · {y}: {s}", "INFO"))
        else:
            overview.append(_line(f"  · {y}: (없음)", "WARNING"))

    hidden = summary.get("hidden_sheet")
    cnt = summary.get("hidden_course_count", 0)
    data_source = summary.get("data_source", "알 수 없음")
    vocational_cnt = summary.get("vocational_course_count", 0)

    if hidden:
        overview.append(_line(f"- 지침 시트: {hidden} (과목 {cnt}개)", "INFO"))
        overview.append(_line(f"- 전문교과목록: {vocational_cnt}개 과목", "INFO"))
        overview.append(_line(f"- 데이터 출처: {data_source}", "INFO"))
    else:
        overview.append(_line("- 지침 시트: (없음)", "ERROR"))

    err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
    warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")
    check_cnt = sum(1 for x in issues if x.get("severity") == "CHECK")
    overview.append(
        _line(f"- 총계: 오류 {err_cnt}건 / 경고 {warn_cnt}건 / 확인 {check_cnt}건", "INFO")
    )
    overview.append(_line("", "INFO"))
    overview.append(_line("[시트별 안내]", "HEADER"))
    overview.append(_line("- 각 탭에서 해당 시트의 문제상황만 확인할 수 있습니다.", "INFO"))
    overview.append(
        _line("- '기타' 탭에는 파일/시트 누락 등 특정 시트에 귀속되지 않는 오류가 표시됩니다.", "INFO")
    )
    overview.append(_line("", "INFO"))

    if not issues:
        overview.append(_line("문제 없음.", "INFO"))
        status = f"검사 완료: 오류 없음 (경고 {warn_cnt}건, 확인 {check_cnt}건)"
        return {
            "ok": True,
            "status": status,
            "counts": {"error": err_cnt, "warning": warn_cnt, "check": check_cnt},
            "tab_names": tab_names,
            "tabs": tabs,
        }

    groups: dict[str, list] = {}
    for it in issues:
        if not isinstance(it, dict):
            continue
        sheet = it.get("sheet", "-") or "-"
        groups.setdefault(sheet, []).append(it)

    sev_rank = {"ERROR": 0, "WARNING": 1, "CHECK": 2, "INFO": 3}

    def sort_key(x):
        row = x.get("row", "-")
        try:
            row_n = int(row)
        except Exception:
            row_n = 10**9
        return (sev_rank.get(x.get("severity", "INFO"), 9), row_n)

    sheet_2024 = targets.get(2024)
    sheet_2026 = targets.get(2026)
    if all_grades_sheet is None:
        all_sheet_names = set(groups.keys()) | set(tab_names)
        all_grades_sheet = find_all_grades_sheet(list(all_sheet_names))

    # 시트별 안내
    for tab_name in tab_names:
        if tab_name in ("전체", "기타"):
            continue
        lines = tabs[tab_name]
        if tab_name == sheet_2026 and sheet_2026:
            lines.append(_line("[안내]", "HEADER"))
            lines.append(
                _line(
                    "교차이수과목의 경우 ↔ 왼쪽 과목을 윗줄, 오른쪽 과목을 아랫줄으로 판단합니다.",
                    "INFO",
                )
            )
            lines.append(_line("", "INFO"))
        if tab_name == sheet_2024 and sheet_2024:
            lines.append(_line("[안내]", "HEADER"))
            lines.append(
                _line("2015개정 교육과정의 과목명의 경우는 일치 여부를 확인하지 않습니다.", "INFO")
            )
            lines.append(_line("지침의 표를 확인하고 정확하게 입력해주세요.", "INFO"))
            lines.append(_line("", "INFO"))
        if tab_name == all_grades_sheet and all_grades_sheet:
            lines.append(_line("[안내]", "HEADER"))
            lines.append(
                _line(
                    "개설 여부는 프로그램 상 확인 절차가 따로 없습니다. 선택군은 학년별로 다르게 정리해주세요.",
                    "INFO",
                )
            )
            lines.append(_line("", "INFO"))

    if wb_temp is None:
        try:
            wb_temp = load_workbook(xlsx_path, data_only=True)
        except Exception:
            wb_temp = None

    for sheet, items in groups.items():
        tab = sheet if sheet in tabs else "기타"
        lines = tabs[tab]

        error_count = sum(1 for it in items if it.get("severity") == "ERROR")
        if error_count >= 50:
            lines.append(_line("[경고] 오류가 50개 이상 발견됩니다. 양식이 올바른지 확인해주세요.(교육청 양식 참고)", "WARNING"))
            lines.append(_line("", "INFO"))

        row_groups: dict = {}
        for it in sorted(items, key=sort_key):
            row = it.get("row", "-")
            row_groups.setdefault(row, []).append(it)

        lines.append(_line("[문제 목록]", "HEADER"))

        def row_sort_key(item):
            key = item[0]
            return (key == "-", int(key) if str(key).isdigit() else 10**9, str(key))

        for row_num, row_items in sorted(row_groups.items(), key=row_sort_key):
            row_label = None

            if row_num != "-" and str(row_num).isdigit() and wb_temp and sheet in wb_temp.sheetnames:
                try:
                    ws = wb_temp[sheet]
                    row_int = int(row_num)
                    course_col = 5 if sheet == sheet_2024 else 4
                    course_cell = ws.cell(row_int, course_col).value
                    if course_cell and str(course_cell).strip():
                        course_name = normalize_course_name(course_cell)
                        if course_name:
                            row_label = course_name
                    if not row_label and sheet == sheet_2024:
                        d_cell = ws.cell(row_int, 4).value
                        if d_cell and str(d_cell).strip():
                            d_name = normalize_course_name(d_cell)
                            if d_name:
                                row_label = d_name
                    if not row_label:
                        a_cell = ws.cell(row_int, 1).value
                        if a_cell and str(a_cell).strip():
                            a_text = str(a_cell).strip()
                            row_label = a_text[:27] + "..." if len(a_text) > 30 else a_text
                        else:
                            b_cell = ws.cell(row_int, 2).value
                            if b_cell and str(b_cell).strip():
                                b_text = str(b_cell).strip()
                                row_label = b_text[:27] + "..." if len(b_text) > 30 else b_text
                except Exception:
                    pass

            if not row_label:
                for it in row_items:
                    matches = re.findall(r"'([^']+)'", it.get("message", ""))
                    if matches and len(matches[0]) < 30:
                        row_label = matches[0]
                        break

            if row_num == "-":
                missing_course_pattern = (
                    r"'([^']+)'\s*시트.*?'([^']+)'\s*과목이\s*'2026\s*전학년'\s*시트에\s*없습니다"
                )
                missing_with_row_pattern = (
                    r"'([^']+)'\s*시트\s*(\d+)행의\s*'([^']+)'\s*과목이\s*'2026\s*전학년'\s*시트에\s*없습니다"
                )
                sheet_groups: dict = {}
                other_items = []

                for it in row_items:
                    msg = it.get("message", "")
                    match = re.search(missing_with_row_pattern, msg)
                    if match:
                        source_sheet = match.group(1)
                        sheet_groups.setdefault(source_sheet, {"with_row": [], "without_row": []})
                        sheet_groups[source_sheet]["with_row"].append(
                            (match.group(3), match.group(2), it)
                        )
                        continue
                    match = re.search(missing_course_pattern, msg)
                    if match:
                        source_sheet = match.group(1)
                        sheet_groups.setdefault(source_sheet, {"with_row": [], "without_row": []})
                        sheet_groups[source_sheet]["without_row"].append((match.group(2), it))
                        continue
                    other_items.append(it)

                for source_sheet in sorted(sheet_groups.keys()):
                    data = sheet_groups[source_sheet]
                    lines.append(
                        _line(
                            f"▶ '{source_sheet}'에 있지만, '2026 전학년' 시트에 없는 과목",
                            "COURSE",
                        )
                    )
                    lines.append(_line("─" * 60, "INFO"))
                    for course, row_no, it in data["with_row"]:
                        sev = it.get("severity", "INFO")
                        tag = sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO"
                        lines.append(_line(f"  [{sev}] {course} ({row_no}행)", tag))
                    for course, it in data["without_row"]:
                        sev = it.get("severity", "INFO")
                        tag = sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO"
                        lines.append(_line(f"  [{sev}] {course}", tag))

                if other_items:
                    lines.append(_line("▶ 기타", "COURSE"))
                    lines.append(_line("─" * 60, "INFO"))
                    for it in other_items:
                        _append_issue_lines(lines, it.get("severity", "INFO"), it.get("message", ""))
            else:
                if row_label:
                    lines.append(_line(f"▶ {row_num}행 - {row_label}", "COURSE"))
                else:
                    lines.append(_line(f"▶ {row_num}행", "COURSE"))
                lines.append(_line("─" * 60, "INFO"))
                for it in row_items:
                    _append_issue_lines(lines, it.get("severity", "INFO"), it.get("message", ""))

        e = sum(1 for x in items if x.get("severity") == "ERROR")
        w = sum(1 for x in items if x.get("severity") == "WARNING")
        c = sum(1 for x in items if x.get("severity") == "CHECK")
        lines.append(_line("=" * 60, "INFO"))
        lines.append(_line(f"[전체 요약] 오류 {e}건, 경고 {w}건, 확인 {c}건", "HEADER"))

    for tab_name in tab_names:
        if tab_name == "전체":
            continue
        # 시트명 키 또는 '-'→기타 로 매핑된 이슈가 있으면 비어 있지 않음
        has_issues = tab_name in groups or (tab_name == "기타" and "-" in groups)
        if not has_issues:
            has_problem_header = any(
                ln.get("text") == "[문제 목록]" for ln in tabs[tab_name]
            )
            if not has_problem_header:
                tabs[tab_name].append(_line("발견된 오류가 없습니다.", "HEADER"))

    overview.append(_line("[전체 문제 요약(시트별)]", "HEADER"))
    for sheet, items in sorted(groups.items(), key=lambda kv: kv[0]):
        e = sum(1 for x in items if x.get("severity") == "ERROR")
        w = sum(1 for x in items if x.get("severity") == "WARNING")
        c = sum(1 for x in items if x.get("severity") == "CHECK")
        label = sheet if sheet != "-" else "기타"
        overview.append(_line(f"- {label}: 오류 {e} / 경고 {w} / 확인 {c}", "INFO"))

    if err_cnt == 0:
        status = f"검사 완료: 오류 없음 (경고 {warn_cnt}건, 확인 {check_cnt}건)"
    else:
        status = f"검사 완료: 오류 {err_cnt}건 / 경고 {warn_cnt}건 / 확인 {check_cnt}건"

    return {
        "ok": True,
        "status": status,
        "counts": {"error": err_cnt, "warning": warn_cnt, "check": check_cnt},
        "tab_names": tab_names,
        "tabs": tabs,
    }


# ---------------------------------------------------------------------------
# 라우트
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html", template_url=TEMPLATE_DOWNLOAD_URL)


@app.route("/api/check", methods=["POST"])
def api_check():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일이 전송되지 않았습니다."}), 400

    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "파일을 선택하세요."}), 400

    original_name = file.filename
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_EXT:
        return jsonify(
            {"ok": False, "error": "지원하지 않는 확장자입니다. .xlsx 또는 .xlsm만 지원합니다."}
        ), 400

    safe_stem = secure_filename(Path(original_name).stem) or "upload"
    save_name = f"{safe_stem}_{uuid.uuid4().hex[:8]}{ext}"
    save_path = UPLOAD_DIR / save_name

    try:
        file.save(save_path)
        issues, summary = run_checks(str(save_path))
        result = format_results(str(save_path), original_name, issues, summary)
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": f"검사 중 오류가 발생했습니다: {e}"}), 500
    finally:
        try:
            if save_path.exists():
                os.remove(save_path)
        except OSError:
            pass


@app.route("/health")
def health():
    return jsonify({"ok": True, "checker": CHECKER_PATH.name})


if __name__ == "__main__":
    # 개발용: http://127.0.0.1:5000
    app.run(host="127.0.0.1", port=5000, debug=True)
