# -*- coding: utf-8 -*-
"""
교육과정 편성표 확인 프로그램 (Tkinter + openpyxl)
- 단일 파일(모듈 분리 없음)
- .xlsx / .xlsm 지원 (openpyxl)

요구사항 반영(버전 v4)
1) 문제상황 출력:
   - 시트별로 별도 "칸"에 표시(Notebook 탭: 전체/각 시트/기타)
2) 2025/2026 입학생 시트:
   - 과목명(D열) 셀에 채우기 색(흰색 제외)이 있으면 그 행은 "모든 검사" 제외
   - 병합 셀인 경우, 해당 D셀의 병합 top-left 셀의 색도 함께 판단(엑셀에서 보이는 색을 더 정확히 반영)
3) 운영학점(F) vs G~L 합:
   - G~L 합이 0이면 '바로 위 행의 운영학점(F)'과 같으면 통과
4) 2024 입학생 시트:
   - 과목명(D) 숨김 시트 일치 여부는 확인하지 않음(불일치 오류 미출력)
5) 과목명에 ↔ 가 있으면:
   - 좌/우 과목명을 각각 숨김 시트에 존재하는지 확인(없으면 오류)
   - ↔ 행은 숨김 기반(유형/기본학점/성적처리/범위) 비교는 생략
   - 단, 내부 일관성(운영학점 vs G~L 합, M/N 합계 계산)은 계속 점검(색깔 행은 전부 제외)

사용 방법
1) pip install openpyxl
2) python curriculum_checker_v4.py
"""

import os
import re
import difflib
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

from openpyxl import load_workbook


# =========================
# 유틸
# =========================

EPS = 1e-9


def normalize_course_name(name: str) -> str:
    """괄호( ) 안 내용을 제거하고, 양끝 공백만 제거(내부 공백은 유지)."""
    if name is None:
        return ""
    s = str(name)
    s = re.sub(r"\([^)]*\)", "", s)  # ( ... ) 제거
    return s.strip()


def split_bidirectional(name: str):
    """'음악↔미술' 같은 문자열을 ['음악','미술']로 분해(양쪽 공백 제거)."""
    if name is None:
        return []
    s = str(name)
    if "↔" not in s:
        return []
    parts = [p.strip() for p in s.split("↔")]
    return [p for p in parts if p != ""]


def is_error_token(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().upper()
    return s in ("#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def to_number(value):
    """숫자 변환(정수/실수). 실패 시 None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value):  # NaN
            return None
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def find_sheet_for_year(sheetnames, year: int):
    """
    '2026 입학생...' 또는 '2026학년도 입학생...' 등:
    - 시트명이 year로 시작하고 '입학생'을 포함하면 매칭
    """
    y = str(year)
    candidates = [n for n in sheetnames if n.startswith(y) and ("입학생" in n)]
    if candidates:
        return candidates[0]
    for n in sheetnames:
        n2 = n.replace(" ", "")
        if n2.startswith(f"{y}입학생"):
            return n
    return None


def find_hidden_sheet_name(sheetnames):
    """'숨김' 시트를 찾음(정확 일치 우선, 포함은 차선)."""
    if "숨김" in sheetnames:
        return "숨김"
    for n in sheetnames:
        if "숨김" in n:
            return n
    return None


def build_merged_lookup(ws):
    """
    셀 좌표(행,열) -> (min_row, min_col, max_row, max_col) 매핑을 만든다.
    """
    lookup = {}
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = (min_row, min_col, max_row, max_col)
    return lookup


def get_value_with_merge(ws_values, ws_formula, merged_lookup, row, col):
    """
    data_only 값(ws_values) 기준으로:
    - 해당 셀이 병합 영역이면 top-left 값으로 보정
    - 동시에 수식(ws_formula)의 존재도 반환
    return: (value, formula_str_or_none, (used_row, used_col))
    """
    used_row, used_col = row, col
    key = (row, col)

    if key in merged_lookup:
        min_row, min_col, _, _ = merged_lookup[key]
        used_row, used_col = min_row, min_col

    v = ws_values.cell(used_row, used_col).value
    f = ws_formula.cell(used_row, used_col).value
    formula = f if isinstance(f, str) and f.startswith("=") else None
    return v, formula, (used_row, used_col)


def find_hidden_header_row(ws_values, ws_formula, merged_lookup):
    """숨김 시트에서 '과목명' 헤더 행 찾기(기본 2행)."""
    for r in range(1, 21):
        v, _, _ = get_value_with_merge(ws_values, ws_formula, merged_lookup, r, 2)  # B열
        if v is not None and str(v).strip() == "과목명":
            return r
    return 2


def safe_strip(v):
    if v is None:
        return ""
    return str(v).strip()


def is_colored_fill(cell) -> bool:
    """
    '색깔이 있는 경우' 판단:
    - 패턴이 있고(대개 solid), 흰색이 아닌 채우기면 True
    - theme/indexed 등 RGB가 애매한 경우도 색으로 간주(보수적으로 제외)
    """
    try:
        fill = cell.fill
    except Exception:
        return False

    if fill is None:
        return False

    pt = getattr(fill, "patternType", None)
    if pt is None or str(pt).lower() in ("none", "null"):
        return False

    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return True

    ctype = getattr(fg, "type", None)
    val = getattr(fg, "value", None) or getattr(fg, "rgb", None)

    if ctype in ("theme", "indexed"):
        return True

    if not val:
        return True

    s = str(val).upper()
    if len(s) >= 6 and s[-6:] == "FFFFFF":
        return False
    if s in ("00000000", "000000", "FFFFFFFF", "00FFFFFF", "FFFFFF"):
        return False

    return True


def is_course_cell_colored(ws_f, merge_lookup, row, col) -> bool:
    """
    과목명 셀(D)에 색이 있는지 판단:
    - (row, col)이 병합 셀인 경우: 병합 top-left 셀의 fill을 함께 확인
    - (row, col) 자체 fill도 확인
    """
    colored = False
    # D 셀 자체
    try:
        colored = colored or is_colored_fill(ws_f.cell(row, col))
    except Exception:
        pass

    # 병합 top-left
    key = (row, col)
    if key in merge_lookup:
        min_row, min_col, _, _ = merge_lookup[key]
        try:
            colored = colored or is_colored_fill(ws_f.cell(min_row, min_col))
        except Exception:
            pass

    return colored


# =========================
# 핵심 검사 로직
# =========================

def run_checks(xlsx_path: str):
    """
    return: (issues, summary)
      - issues: list[dict]  {severity, sheet, row, message}
      - summary: dict
    """
    if not os.path.exists(xlsx_path):
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": "파일을 찾을 수 없습니다."}], {})

    ext = os.path.splitext(xlsx_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": "지원하지 않는 확장자입니다. .xlsx 또는 .xlsm만 지원합니다."}], {})

    issues = []
    summary = {}

    try:
        wb_v = load_workbook(xlsx_path, data_only=True)
        wb_f = load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return ([{"severity": "ERROR", "sheet": "-", "row": "-", "message": f"엑셀 파일을 열 수 없습니다: {e}"}], {})

    sheetnames = wb_v.sheetnames

    # (1) 대상 시트 존재 확인
    targets = {}
    for y in (2026, 2025, 2024):
        sname = find_sheet_for_year(sheetnames, y)
        targets[y] = sname
        if not sname:
            issues.append({"severity": "ERROR", "sheet": "-", "row": "-", "message": f"{y} 입학생으로 시작하고 '입학생'을 포함하는 시트를 찾지 못했습니다."})

    # 숨김 시트 로드
    hidden_name = find_hidden_sheet_name(sheetnames)
    if not hidden_name:
        issues.append({"severity": "ERROR", "sheet": "-", "row": "-", "message": "지침 시트를 찾지 못했습니다(시트명에 '숨김' 포함 필요)."})
        return issues, {"targets": targets, "hidden_sheet": None}

    ws_hidden_v = wb_v[hidden_name]
    ws_hidden_f = wb_f[hidden_name]
    hidden_merge = build_merged_lookup(ws_hidden_f)
    header_row = find_hidden_header_row(ws_hidden_v, ws_hidden_f, hidden_merge)
    data_start = header_row + 1

    # 숨김 과목 사전 구축
    hidden = {}
    hidden_list_norm = []
    r = data_start
    while True:
        course_raw, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 2)  # B
        if course_raw is None or str(course_raw).strip() == "":
            break
        course_norm = normalize_course_name(course_raw)

        typ, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 3)  # C
        basic, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 4)  # D
        grade, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 5)  # E
        minc, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 6)  # F
        maxc, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 7)  # G
        special_note, _, _ = get_value_with_merge(ws_hidden_v, ws_hidden_f, hidden_merge, r, 9)  # I

        rec = {
            "course_raw": safe_strip(course_raw),
            "type": safe_strip(typ),
            "basic": to_number(basic),
            "grading": safe_strip(grade),
            "min": to_number(minc),
            "max": to_number(maxc),
            "special_note": safe_strip(special_note),
            "row": r,
        }

        if course_norm in hidden:
            issues.append({
                "severity": "WARNING",
                "sheet": hidden_name,
                "row": r,
                "message": f"지침에 중복 과목명이 있습니다: '{course_norm}' (기존 {hidden[course_norm]['row']}행, 추가 {r}행). 최초 항목을 기준으로 검사합니다."
            })
        else:
            hidden[course_norm] = rec
            hidden_list_norm.append(course_norm)

        r += 1

    summary["targets"] = targets
    summary["hidden_sheet"] = hidden_name
    summary["hidden_course_count"] = len(hidden)

    # 시트가 없다면 여기서 종료
    if any(v is None for v in targets.values()):
        return issues, summary

    # =========================
    # (2) 각 시트 검사
    # =========================
    for year, sname in targets.items():
        ws_v = wb_v[sname]
        ws_f = wb_f[sname]
        merge_lookup = build_merged_lookup(ws_f)

        first_row = 5
        course_col = 4  # D
        type_col = 3    # C
        basic_col = 5   # E
        
        # 2024 시트는 열 구성이 다름
        if year == 2024:
            op_col = 7      # G (운영학점)
            sem_cols = list(range(8, 14))  # H~M
            total_cols = [14, 15]          # N, O
            grading_col = 16               # P
            op_col_name = "G"
            sem_cols_name = "H~M"
            total_cols_name = "N/O"
            compare_col = 2  # B열 (2024는 B열과 N,O열 병합 비교)
        else:  # 2025, 2026
            op_col = 6      # F (운영학점)
            sem_cols = list(range(7, 13))  # G~L
            total_cols = [13, 14]          # M, N
            grading_col = 15               # O
            op_col_name = "F"
            sem_cols_name = "G~L"
            total_cols_name = "M/N"
            compare_col = 1  # A열 (2025/2026은 A열과 M,N열 병합 비교)

        # last row 찾기: '편성 학점 수' 또는 '편성학점수'가 포함된 행
        last_row = None
        for rr in range(first_row, ws_f.max_row + 1):
            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if v is not None:
                v_str = str(v).strip().replace(" ", "")
                if "편성학점수" in v_str or "편성학점합계" in v_str:
                    last_row = rr
                    break
        
        if last_row is None:
            # '편성 학점 수'를 찾지 못한 경우
            issues.append({
                "severity": "ERROR", 
                "sheet": sname, 
                "row": "-", 
                "message": (
                    "시트의 마지막 행에서 '편성 학점 수'를 찾지 못했습니다.\n"
                    "표의 총계 부분이 양식과 같이 입력되어 있는지 확인하고 다시 실행해 주세요.\n\n"
                    "[필요한 총계 행 구성(작년 양식에서 변경되었습니다.)]\n"
                    "• 학생 지정 과목 교과 편성 학점\n"
                    "• 학생 선택 과목 교과 편성 학점\n"
                    "• 총 교과 편성 학점\n"
                    "• 창의적 체험활동 학점\n"
                    "• 편성 학점 수"
                )
            })
            continue
        
        # '편성 학점 수' 행은 검사 대상이 아니므로, 실제 검사는 그 위까지만
        check_until_row = last_row - 1

        # 2025/2026: 과목명(D) '색깔' 행은 모든 검사 제외
        exempt_rows = set()
        if year in (2025, 2026, 2024):
            for rr in range(first_row, check_until_row + 1):
                if is_course_cell_colored(ws_f, merge_lookup, rr, course_col):
                    exempt_rows.add(rr)

        # row_total(각 행의 G~L 합) 계산 (색깔 행은 계산에서도 제외)
        row_total = {}
        for rr in range(first_row, check_until_row + 1):
            if rr in exempt_rows:
                continue

            course_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_v is None or str(course_v).strip() == "":
                continue

            if is_error_token(course_v):
                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"과목명(D{rr})에 오류값이 있습니다: {course_v}"})
                continue

            sem_sum = 0.0
            any_num = False
            for cc in sem_cols:
                v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, cc)
                n = to_number(v)
                if n is not None:
                    sem_sum += n
                    any_num = True

            row_total[rr] = sem_sum if any_num else 0.0

        # ========== 과목 단위 검사 ==========
        for rr in range(first_row, check_until_row + 1):
            if rr in exempt_rows:
                continue  # 색깔 행은 전부 제외

            course_raw, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
            if course_raw is None or str(course_raw).strip() == "":
                continue
            if is_error_token(course_raw):
                continue

            course_norm = normalize_course_name(course_raw)
            if course_norm == "":
                issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": "과목명(D열)에서 괄호 제거 후 이름이 비었습니다."})
                continue

            parts = split_bidirectional(course_norm)
            is_bidirectional = len(parts) >= 2

            # (2) 과목명 일치 여부
            hidden_rec = None
            if year == 2024:
                # 2024는 과목명 일치 검증을 하지 않음(있으면 활용, 없으면 숨김기반 검사는 생략)
                hidden_rec = hidden.get(course_norm, None)
            else:
                if is_bidirectional:
                    missing = [p for p in parts if p not in hidden]
                    if missing:
                        hints = []
                        for m in missing:
                            close = difflib.get_close_matches(m, hidden_list_norm, n=1, cutoff=0.6)
                            if close:
                                hints.append(f"{m}→{close[0]}")
                        hint_txt = f" (유사 후보: {', '.join(hints)})" if hints else ""
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"↔ 과목명 중 지침에 없는 항목: {', '.join(missing)}{hint_txt}"})
                    
                    # ↔ 과목도 병합 셀 위치에 따라 좌/우 과목으로 검증
                    # 병합 셀의 top-left 행이면 왼쪽 과목, 아니면 오른쪽 과목
                    key = (rr, course_col)
                    if key in merge_lookup:
                        min_row, min_col, _, _ = merge_lookup[key]
                        if rr == min_row:
                            # 병합 영역의 첫 행 -> 왼쪽 과목
                            target_course = parts[0] if len(parts) > 0 else None
                        else:
                            # 병합 영역의 두 번째 행 이후 -> 오른쪽 과목
                            target_course = parts[1] if len(parts) > 1 else None
                        
                        if target_course and target_course in hidden:
                            hidden_rec = hidden[target_course]
                        else:
                            hidden_rec = None
                    else:
                        # 병합되지 않은 경우는 검증 생략
                        hidden_rec = None
                else:
                    if course_norm not in hidden:
                        hint = ""
                        close = difflib.get_close_matches(course_norm, hidden_list_norm, n=2, cutoff=0.6)
                        if close:
                            hint = f" (유사 과목명 후보: {', '.join(close)})"
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"과목명 오류: '{course_norm}'{hint}"})
                        hidden_rec = None
                    else:
                        hidden_rec = hidden[course_norm]

            # (3) 유형/기본학점/성적처리 (숨김 매칭이 있을 때만)
            if hidden_rec is not None:
                typ_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, type_col)
                typ_s = safe_strip(typ_v)
                if typ_s == "":
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"유형(C{rr})이 비어 있습니다. (지침: {hidden_rec['type']})"})
                elif typ_s != hidden_rec["type"]:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"유형 불일치: 시트='{typ_s}' / 지침='{hidden_rec['type']}'"})

                basic_v, basic_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, basic_col)
                basic_n = to_number(basic_v)
                if basic_n is None:
                    if basic_formula:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rr, "message": f"기본학점(E{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {basic_formula})"})
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"기본학점(E{rr})이 숫자가 아닙니다: {basic_v}"})
                else:
                    if hidden_rec["basic"] is not None and abs(basic_n - hidden_rec["basic"]) > EPS:
                        # 숨김 시트 I열에 특수 사항이 있는지 확인
                        if hidden_rec.get("special_note") and hidden_rec["special_note"] != "":
                            issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"기본학점 불일치: 시트={basic_n:g} / 지침={hidden_rec['basic']:g}. 일반고에서 진로 선택으로 개설할 수 있는 과목. 이상없으면 무시"})
                        else:
                            issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"기본학점 불일치: 시트={basic_n:g} / 지침={hidden_rec['basic']:g}"})

                grade_v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, grading_col)
                grade_s = safe_strip(grade_v)
                if grade_s == "":
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"성적처리 유형(O{rr})이 비어 있습니다. (지침: {hidden_rec['grading']})"})
                elif grade_s != hidden_rec["grading"]:
                    # 과목명에 괄호가 있는 경우 CHECK로 처리
                    has_parenthesis = "(" in str(course_raw) or ")" in str(course_raw)
                    if has_parenthesis:
                        issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"성적처리 유형 확인. 공동교육과정 등으로 인해 이상없으면 무시 (시트='{grade_s}' / 지침='{hidden_rec['grading']}')"})
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"성적처리 유형 불일치: 시트='{grade_s}' / 지침='{hidden_rec['grading']}'"})

            # (4)(5) 운영학점 범위/합계 체크
            op_v, op_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, op_col)
            op_n = to_number(op_v)
            sem_sum = row_total.get(rr, 0.0)

            if op_n is None:
                if op_formula:
                    issues.append({"severity": "WARNING", "sheet": sname, "row": rr, "message": f"운영학점({op_col_name}{rr})이 수식이지만 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {op_formula})"})
                else:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점({op_col_name}{rr})이 숫자가 아닙니다: {op_v}"})
            else:
                if hidden_rec is not None and (hidden_rec["min"] is not None) and (hidden_rec["max"] is not None):
                    if not (hidden_rec["min"] - EPS <= op_n <= hidden_rec["max"] + EPS):
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점 범위 위반: 시트={op_n:g} / 허용범위={hidden_rec['min']:g}~{hidden_rec['max']:g}"})

                if abs(sem_sum) <= EPS:
                    prev = None
                    if rr > first_row:
                        pv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr - 1, op_col)
                        prev = to_number(pv)
                    if prev is not None and abs(op_n - prev) <= EPS:
                        pass
                    else:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"{sem_cols_name} 합이 0이므로 위 행 운영학점과 비교해야 합니다: 현재={op_n:g}, 위행={(prev if prev is not None else '없음')}"})
                else:
                    if abs(op_n - sem_sum) > EPS:
                        # 합이 운영학점의 2배인지 확인 (학기제 가능성)
                        if abs(sem_sum - op_n * 2) <= EPS:
                            issues.append({"severity": "CHECK", "sheet": sname, "row": rr, "message": f"운영학점과 편성된 학점({sem_cols_name}열) 불일치: 운영학점={op_n:g}, {sem_cols_name}합={sem_sum:g} (학기제라면 오류가 아닙니다)"})
                        else:
                            issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"운영학점과 편성된 학점({sem_cols_name}열) 불일치: 운영학점={op_n:g}, {sem_cols_name}합={sem_sum:g}"})

        # (6) 합계 열 병합 구간 합계 체크 (색깔 행은 기대값에서도 제외)
        checked_spans = set()
        for rng in ws_f.merged_cells.ranges:
            if rng.min_col in total_cols and rng.max_col == rng.min_col:
                col = rng.min_col
                if rng.max_row < first_row:
                    continue
                start = max(rng.min_row, first_row)
                end = min(rng.max_row, check_until_row)
                if start > end:
                    continue

                key = (col, rng.min_row, rng.max_row)
                if key in checked_spans:
                    continue
                checked_spans.add(key)

                total_v, total_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rng.min_row, col)
                total_n = to_number(total_v)

                expected = 0.0
                for rr in range(start, end + 1):
                    if rr in exempt_rows:
                        continue
                    cv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
                    if cv is None or str(cv).strip() == "" or is_error_token(cv):
                        continue
                    expected += row_total.get(rr, 0.0)

                # 열 이름 결정
                col_name = total_cols_name.split('/')[0] if col == total_cols[0] else total_cols_name.split('/')[1]
                
                if total_n is None:
                    if total_formula:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rng.min_row, "message": f"{col_name}열 합계 셀에 수식은 있으나 결과값이 없습니다(엑셀 재계산/저장 필요). (수식: {total_formula})"})
                    else:
                        issues.append({"severity": "WARNING", "sheet": sname, "row": rng.min_row, "message": f"{col_name}열 합계 셀이 비어 있습니다. (해당 구간 {sem_cols_name} 합 기대값={expected:g})"})
                else:
                    if abs(total_n - expected) > EPS:
                        issues.append({"severity": "ERROR", "sheet": sname, "row": rng.min_row, "message": f"{col_name}열 합계 불일치: 셀값={total_n:g}, 기대값({sem_cols_name}합)={expected:g} (구간 {start}~{end}행, 색깔행 제외)"})

        # 병합이 아닌 단일 셀 합계 보조 체크(색깔행 제외)
        for col in total_cols:
            for rr in range(first_row, check_until_row + 1):
                if rr in exempt_rows:
                    continue
                if (rr, col) in merge_lookup:
                    min_r, min_c, _, _ = merge_lookup[(rr, col)]
                    if not (rr == min_r and col == min_c):
                        continue
                    continue

                tv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, col)
                tn = to_number(tv)
                if tn is None:
                    continue

                cv, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, course_col)
                if cv is None or str(cv).strip() == "" or is_error_token(cv):
                    continue

                # 열 이름 결정
                col_name = total_cols_name.split('/')[0] if col == total_cols[0] else total_cols_name.split('/')[1]
                
                expected = row_total.get(rr, 0.0)
                if abs(tn - expected) > EPS:
                    issues.append({"severity": "ERROR", "sheet": sname, "row": rr, "message": f"{col_name}열 단일행 합계 불일치: 셀값={tn:g}, 기대값({sem_cols_name}합)={expected:g}"})

        # (7) 총계 구간 찾기: '학생 지정 과목 교과 편성 학점' ~ '학교 선택 과목 교과' 사이
        total_section_start = None
        total_section_end = None
        
        # A열에서 '학생 지정 과목 교과 편성 학점' 또는 '학교 지정 과목 교과 편성 학점' 찾기 (첫 행부터)
        # 더 넓은 검색 조건: "지정", "과목", "교과", "편성", "학점"이 모두 포함되면 찾음
        for rr in range(1, ws_f.max_row + 1):
            v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, 1)  # A열
            if v is not None:
                v_str = str(v).strip().replace(" ", "")
                # "지정"과 "편성학점"이 포함되어 있으면 해당 행으로 간주
                if ("지정" in v_str or "선택" in v_str) and "편성학점" in v_str and "과목" in v_str:
                    total_section_start = rr
                    break
        
        # '학교 선택 과목 교과' 찾기 (학생 지정 이후부터)
        if total_section_start is not None:
            for rr in range(total_section_start + 1, ws_f.max_row + 1):
                v, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, rr, 1)  # A열
                if v is not None:
                    v_str = str(v).strip().replace(" ", "")
                    if "학교선택과목교과" in v_str:
                        total_section_end = rr - 1  # 그 이전 행까지
                        break
            
            # '학교 선택'을 못 찾으면 마지막 행까지
            if total_section_end is None:
                total_section_end = ws_f.max_row
        
        if total_section_start is not None and total_section_end is not None:
            # 비교 열 병합 정보 수집: 각 행이 어느 병합 범위에 속하는지 매핑
            # 2024: B열, 2025/2026: A열
            # 해당 열만 병합된 경우와 여러 열이 함께 병합된 경우 모두 수집
            a_col_merge_map = {}  # {row: (min_row, max_row)}
            for rng in ws_f.merged_cells.ranges:
                # compare_col을 포함하는 병합
                if rng.min_col <= compare_col and rng.max_col >= compare_col:
                    # 총계 구간과 겹치는 병합만 수집
                    if not (rng.max_row < total_section_start or rng.min_row > total_section_end):
                        for r in range(rng.min_row, rng.max_row + 1):
                            a_col_merge_map[r] = (rng.min_row, rng.max_row)
            
            
            # 총계 열 병합 정보 수집 및 비교 열과 비교
            # 2024: N~O 병합과 B열 비교, 2025/2026: M~N 병합과 A열 비교
            for rng in ws_f.merged_cells.ranges:
                # M열과 N열이 함께 병합된 경우 (min_col=M, max_col=N)
                if rng.min_col == total_cols[0] and rng.max_col == total_cols[1]:
                    # 총계 구간과 겹치는 병합만 검사
                    if not (rng.max_row < total_section_start or rng.min_row > total_section_end):
                        merge_start = rng.min_row
                        merge_end = rng.max_row
                        
                        # 비교 열(2024:B열, 2025/2026:A열)과 D열 값 확인 - 특정 키워드 포함 시 검사 제외
                        compare_col_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, merge_start, compare_col)
                        d_col_val, _, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, merge_start, 4)  # D열
                        skip_check = False
                        
                        # 비교 열 체크
                        if compare_col_val is not None:
                            compare_col_str = str(compare_col_val).strip()
                            compare_col_str_no_space = compare_col_str.replace(" ", "")
                            # '학교/학생 지정 과목 편성' 또는 '증배' 포함 시 제외
                            if "학교지정과목편성" in compare_col_str_no_space or "학생지정과목편성" in compare_col_str_no_space or "증배" in compare_col_str:
                                skip_check = True
                        
                        # D열 체크 - 총계 행 제외
                        if d_col_val is not None and not skip_check:
                            d_col_str = str(d_col_val).strip().replace(" ", "")
                            # 총계 관련 키워드가 있으면 제외
                            if "편성학점" in d_col_str or "총교과" in d_col_str or "창의적체험활동" in d_col_str or "편성학점수" in d_col_str:
                                skip_check = True
                        
                        if skip_check:
                            continue  # 이 병합 구간은 검사하지 않음 (병합 불일치도, 합계도 검사 안함)
                        
                        # 총계 열 병합 범위 내의 모든 행이 같은 비교 열 병합에 속하는지 확인
                        compare_merge_range = None
                        mismatch = False
                        
                        for r in range(merge_start, merge_end + 1):
                            if r in a_col_merge_map:
                                current_range = a_col_merge_map[r]
                                if compare_merge_range is None:
                                    compare_merge_range = current_range
                                elif compare_merge_range != current_range:
                                    # 다른 비교 열 병합 범위에 걸쳐있음
                                    mismatch = True
                                    break
                            else:
                                # 비교 열이 병합되지 않은 행
                                mismatch = True
                                break
                        
                        # 비교 열 병합 범위와 총계 열 병합 범위가 정확히 일치하는지 확인
                        if mismatch or compare_merge_range is None or compare_merge_range != (merge_start, merge_end):
                            compare_range_str = f"{compare_merge_range[0]}~{compare_merge_range[1]}" if compare_merge_range else "없음"
                            compare_col_name = "B" if year == 2024 else "A"
                            issues.append({
                                "severity": "ERROR",
                                "sheet": sname,
                                "row": merge_start,
                                "message": f"병합 불일치: {compare_col_name}열 병합({compare_range_str})과 {total_cols_name}열 병합({merge_start}~{merge_end}행)이 일치하지 않습니다."
                            })
                        
                        # 합계 검증 (첫 번째 총계 열 기준: 2024=N열, 2025/2026=M열)
                        # '증배'나 '학교 지정 과목 편성'은 이미 위에서 skip_check으로 제외됨
                        total_v, total_formula, _ = get_value_with_merge(ws_v, ws_f, merge_lookup, merge_start, total_cols[0])
                        total_n = to_number(total_v)
                        
                        if total_n is not None:
                            # 해당 구간의 기대값 계산
                            expected = 0.0
                            for rr in range(merge_start, merge_end + 1):
                                if rr in exempt_rows:
                                    continue
                                expected += row_total.get(rr, 0.0)
                            
                            if abs(total_n - expected) > EPS:
                                first_col_name = "N" if year == 2024 else "M"
                                issues.append({
                                    "severity": "ERROR",
                                    "sheet": sname,
                                    "row": merge_start,
                                    "message": f"{first_col_name}열 합계 불일치: 셀값={total_n:g}, 기대값({sem_cols_name}합)={expected:g} (구간 {merge_start}~{merge_end}행)"
                                })
            

    return issues, summary


# =========================
# GUI
# =========================

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("교육과정 편성표 확인 프로그램")
        self.root.minsize(1020, 700)

        self.colors = {
            "bg": "#F6F7FF",
            "card": "#FFFFFF",
            "text": "#1F2937",
            "muted": "#6B7280",
            "accent": "#7C6CF6",
            "danger": "#EF4444",
            "warn": "#F59E0B",
            "check": "#3B82F6",
        }
        self.root.configure(bg=self.colors["bg"])

        self.style = ttk.Style(self.root)
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        base_font = ("Malgun Gothic", 11)
        title_font = ("Malgun Gothic", 16, "bold")

        self.style.configure("TFrame", background=self.colors["bg"])
        self.style.configure("Card.TFrame", background=self.colors["card"])
        self.style.configure("TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=base_font)
        self.style.configure("Title.TLabel", font=title_font, foreground=self.colors["text"])
        self.style.configure("Muted.TLabel", foreground=self.colors["muted"], font=("Malgun Gothic", 10))

        self._build_ui()
        self.xlsx_path = None

    def _build_ui(self):
        header = ttk.Frame(self.root, padding=(18, 18, 18, 10))
        header.pack(fill="x")

        ttk.Label(header, text="교육과정 편성표 확인 프로그램", style="Title.TLabel").pack(anchor="w")
        
        text_frame = ttk.Frame(header)
        text_frame.pack(fill="x", pady=(6, 0))
        
        # 강조 부분 1: "교육청에서 제공한 엑셀 파일"
        tk.Label(
            text_frame,
            text="교육청에서 제공한 엑셀 파일",
            bg=self.colors["bg"],
            fg="#EF4444",
            font=("Malgun Gothic", 10, "bold")
        ).pack(side="left")
        
        # 일반 텍스트 부분
        ttk.Label(
            text_frame,
            text="에 작성된 편성표를 점검합니다. 파일을 ",
            style="Muted.TLabel"
        ).pack(side="left")
        
        # 강조 부분 2: "저장하고 닫은 후에"
        tk.Label(
            text_frame,
            text="저장하고 닫은 후에",
            bg=self.colors["bg"],
            fg="#EF4444",
            font=("Malgun Gothic", 10, "bold")
        ).pack(side="left")
        
        # 일반 텍스트 부분
        ttk.Label(
            text_frame,
            text=" 업로드하세요.",
            style="Muted.TLabel"
        ).pack(side="left")
        
        download_btn = tk.Button(
            text_frame,
            text="편성표 양식 다운로드",
            command=lambda: webbrowser.open("https://drive.google.com/drive/folders/1wvdV4VQD7kUD7eVEvypPf39LDWZLxfze?usp=sharing"),
            bg=self.colors["accent"],
            fg="white",
            bd=0,
            activebackground=self.colors["accent"],
            activeforeground="white",
            padx=12,
            pady=6,
            font=("Malgun Gothic", 10),
            cursor="hand2"
        )
        download_btn.pack(side="right", padx=(10, 0))

        body = ttk.Frame(self.root, padding=(18, 8, 18, 18))
        body.pack(fill="both", expand=True)

        card = ttk.Frame(body, style="Card.TFrame", padding=(16, 16))
        card.pack(fill="x")

        row1 = ttk.Frame(card, style="Card.TFrame")
        row1.pack(fill="x")

        self.path_var = tk.StringVar(value="선택된 파일 없음")
        ttk.Label(row1, text="엑셀 파일:", style="Muted.TLabel").pack(side="left")
        ttk.Label(row1, textvariable=self.path_var).pack(side="left", padx=(8, 0))

        btn_frame = ttk.Frame(card, style="Card.TFrame")
        btn_frame.pack(fill="x", pady=(12, 0))

        self.btn_pick = tk.Button(
            btn_frame,
            text="파일 선택",
            command=self.pick_file,
            bg=self.colors["accent"],
            fg="white",
            bd=0,
            activebackground=self.colors["accent"],
            activeforeground="white",
            padx=16,
            pady=10,
            font=("Malgun Gothic", 11, "bold"),
            cursor="hand2"
        )
        self.btn_pick.pack(side="left")

        self.btn_run = tk.Button(
            btn_frame,
            text="검사 실행",
            command=self.run,
            bg="#C7C9D9",
            fg="white",
            bd=0,
            activebackground="#C7C9D9",
            activeforeground="white",
            padx=16,
            pady=10,
            font=("Malgun Gothic", 11, "bold"),
            cursor="hand2",
            state="disabled"
        )
        self.btn_run.pack(side="left", padx=(10, 0))

        status_frame = ttk.Frame(card, style="Card.TFrame")
        status_frame.pack(fill="x", pady=(12, 0))

        self.status_var = tk.StringVar(value="대기 중")
        ttk.Label(status_frame, textvariable=self.status_var, style="Muted.TLabel").pack(side="left")

        self.progress = ttk.Progressbar(status_frame, mode="indeterminate", length=220)
        self.progress.pack(side="right")

        # 결과: Notebook(탭)으로 시트별 출력
        out_card = ttk.Frame(body, style="Card.TFrame", padding=(16, 16))
        out_card.pack(fill="both", expand=True, pady=(14, 0))

        ttk.Label(out_card, text="문제상황(시트별)", style="Muted.TLabel").pack(anchor="w")

        self.nb = ttk.Notebook(out_card)
        self.nb.pack(fill="both", expand=True, pady=(8, 0))

        self.text_widgets = {}  # tab_name -> ScrolledText

        # 기본 탭: 전체/기타 (실행 시 대상 시트 탭은 동적으로 재구성)
        self._reset_tabs(["전체", "기타"])

    def _reset_tabs(self, tab_names):
        # 기존 탭 제거
        for tab_id in self.nb.tabs():
            self.nb.forget(tab_id)
        self.text_widgets.clear()

        for name in tab_names:
            frame = ttk.Frame(self.nb, padding=(8, 8))
            self.nb.add(frame, text=name)

            txt = ScrolledText(
                frame,
                wrap="word",
                height=18,
                font=("Consolas", 10),
                bg="#FBFBFE",
                fg=self.colors["text"],
                relief="solid",
                bd=1,
                padx=10,
                pady=10
            )
            txt.pack(fill="both", expand=True)
            txt.tag_configure("ERROR", foreground=self.colors["danger"])
            txt.tag_configure("WARNING", foreground=self.colors["warn"])
            txt.tag_configure("CHECK", foreground=self.colors["check"])
            txt.tag_configure("INFO", foreground=self.colors["muted"])
            txt.tag_configure("HEADER", font=("Malgun Gothic", 11, "bold"))
            self.text_widgets[name] = txt

    def pick_file(self):
        path = filedialog.askopenfilename(
            title="교육과정 편성표 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return
        self.xlsx_path = path
        self.path_var.set(path)
        self.btn_run.configure(state="normal", bg=self.colors["accent"], activebackground=self.colors["accent"])

    def run(self):
        if not self.xlsx_path:
            messagebox.showwarning("안내", "먼저 엑셀 파일을 선택하세요.")
            return

        # 초기화(탭은 summary 읽고 나서 구성)
        self.status_var.set("검사 중...")
        self.progress.start(12)
        self.root.update_idletasks()

        try:
            issues, summary = run_checks(self.xlsx_path)
        except Exception as e:
            self.progress.stop()
            self.status_var.set("오류 발생")
            messagebox.showerror("오류", f"검사 중 예외가 발생했습니다:\n{e}")
            return

        self.progress.stop()

        # 탭 구성: 전체 + (대상 시트) + 기타
        targets = (summary.get("targets") or {})
        tab_names = ["전체"]
        # 연도 시트(존재하는 것만)
        for y in (2026, 2025, 2024):
            s = targets.get(y)
            if s and s not in tab_names:
                tab_names.append(s)
        tab_names.append("기타")
        self._reset_tabs(tab_names)

        # 출력
        self._print_summary(summary, issues)
        self._print_issues_per_sheet(issues)

        err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
        warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")
        check_cnt = sum(1 for x in issues if x.get("severity") == "CHECK")
        if err_cnt == 0:
            self.status_var.set(f"검사 완료: 오류 없음 (경고 {warn_cnt}건, 확인 {check_cnt}건)")
        else:
            self.status_var.set(f"검사 완료: 오류 {err_cnt}건 / 경고 {warn_cnt}건 / 확인 {check_cnt}건")

        # 기본으로 "전체" 탭 보여주기
        try:
            self.nb.select(0)
        except Exception:
            pass

    def _w(self, tab, text, tag="INFO"):
        txt = self.text_widgets.get(tab)
        if not txt:
            txt = self.text_widgets.get("기타")
        txt.insert("end", text, tag)

    def _print_summary(self, summary, issues):
        tab = "전체"
        txt = self.text_widgets[tab]
        txt.delete("1.0", "end")

        self._w(tab, "[검사 개요]\n", "HEADER")
        self._w(tab, f"- 파일: {self.xlsx_path}\n", "INFO")

        targets = summary.get("targets") or {}
        self._w(tab, "- 시트 확인:\n", "INFO")
        for y in (2026, 2025, 2024):
            s = targets.get(y)
            if s:
                self._w(tab, f"  · {y}: {s}\n", "INFO")
            else:
                self._w(tab, f"  · {y}: (없음)\n", "WARNING")

        hidden = summary.get("hidden_sheet")
        cnt = summary.get("hidden_course_count", 0)
        if hidden:
            self._w(tab, f"- 지침 시트: {hidden} (과목 {cnt}개)\n", "INFO")
        else:
            self._w(tab, "- 지침 시트: (없음)\n", "ERROR")

        err_cnt = sum(1 for x in issues if x.get("severity") == "ERROR")
        warn_cnt = sum(1 for x in issues if x.get("severity") == "WARNING")
        check_cnt = sum(1 for x in issues if x.get("severity") == "CHECK")
        self._w(tab, f"- 총계: 오류 {err_cnt}건 / 경고 {warn_cnt}건 / 확인 {check_cnt}건\n\n", "INFO")

        self._w(tab, "[시트별 안내]\n", "HEADER")
        self._w(tab, "- 각 탭에서 해당 시트의 문제상황만 확인할 수 있습니다.\n", "INFO")
        self._w(tab, "- '기타' 탭에는 파일/시트 누락 등 특정 시트에 귀속되지 않는 오류가 표시됩니다.\n\n", "INFO")

    def _print_issues_per_sheet(self, issues):
        # 다른 탭들은 내용만 초기화(전체는 summary가 있으므로 유지)
        for name, txt in self.text_widgets.items():
            if name == "전체":
                continue
            txt.delete("1.0", "end")

        if not issues:
            self._w("전체", "문제 없음.\n", "INFO")
            return

        # 그룹핑
        groups = {}
        for it in issues:
            sheet = it.get("sheet", "-") or "-"
            groups.setdefault(sheet, []).append(it)

        sev_rank = {"ERROR": 0, "WARNING": 1, "CHECK": 2, "INFO": 3}

        def sort_key(x):
            row = x.get("row", "-")
            try:
                row_n = int(row)
            except Exception:
                row_n = 10**9
            return (sev_rank.get(x.get("severity", "INFO"), 9), row_n)

        # 각 시트 탭에 출력
        for sheet, items in groups.items():
            tab = sheet if sheet in self.text_widgets else "기타"
            self._w(tab, "[문제 목록]\n", "HEADER")
            for it in sorted(items, key=sort_key):
                sev = it.get("severity", "INFO")
                row = it.get("row", "-")
                msg = it.get("message", "")
                self._w(tab, f"- [{sev}] 행 {row}: {msg}\n", sev if sev in ("ERROR", "WARNING", "CHECK") else "INFO")

            err_cnt = sum(1 for x in items if x.get("severity") == "ERROR")
            warn_cnt = sum(1 for x in items if x.get("severity") == "WARNING")
            check_cnt = sum(1 for x in items if x.get("severity") == "CHECK")
            self._w(tab, "\n", "INFO")
            self._w(tab, f"[요약] 오류 {err_cnt}건, 경고 {warn_cnt}건, 확인 {check_cnt}건\n", "HEADER")

        # 전체 탭에도 전체 지침 간단 요약(원하면 제거 가능)
        self._w("전체", "[전체 문제 요약(시트별)]\n", "HEADER")
        for sheet, items in sorted(groups.items(), key=lambda kv: kv[0]):
            err_cnt = sum(1 for x in items if x.get("severity") == "ERROR")
            warn_cnt = sum(1 for x in items if x.get("severity") == "WARNING")
            check_cnt = sum(1 for x in items if x.get("severity") == "CHECK")
            label = sheet if sheet != "-" else "기타"
            self._w("전체", f"- {label}: 오류 {err_cnt} / 경고 {warn_cnt} / 확인 {check_cnt}\n", "INFO")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
